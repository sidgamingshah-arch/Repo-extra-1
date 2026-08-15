"""Dump the current template's on-screen output structure to Excel, for review and revision.

Built from ``_build_statement`` — the same function the Workspace grid renders from — so the sheet
cannot describe a layout different from the one on screen. Every row the screen shows appears here in
screen order, including the template lines a filing never populated, because those are exactly the
ones a reviser needs to see: a line nobody extracted is still a line the template asks for.

Calculated lines carry their formula from the TEMPLATE's rollup declaration rather than from the run,
so the derivation is shown whether or not this particular filing had the inputs to compute it.

    python scripts/template_output_structure.py [--out PATH] [--basis consolidated|standalone]
"""
from __future__ import annotations

import argparse
import sys

_STATEMENTS = ("balance_sheet", "profit_and_loss", "cash_flow", "changes_in_equity")
_SHEET = {"balance_sheet": "Balance Sheet", "profit_and_loss": "Profit & Loss",
          "cash_flow": "Cash Flow", "changes_in_equity": "Changes in Equity"}
_HEAD = ["#", "Kind", "Label (as shown)", "Canonical key", "Note", "Current", "Prior",
         "Status", "Calculated", "Formula (from the template)", "Confidence", "Source"]


def _rollups(template: dict) -> dict[str, dict]:
    """canonical_key -> rollup, for every node that declares one, at any depth."""
    out: dict[str, dict] = {}

    def walk(nodes: list[dict]) -> None:
        for n in nodes or []:
            if n.get("canonical_key") and n.get("rollup"):
                out[n["canonical_key"]] = n["rollup"]
            walk(n.get("children") or [])

    for stmt in template.get("statements") or []:
        walk(stmt.get("sections") or [])
    return out


def _formula(rollup: dict | None) -> str:
    if not rollup:
        return ""
    op = str(rollup.get("op") or "sum")
    kids = list(rollup.get("children") or [])
    if not kids:
        return op
    joiner = " + " if op == "sum" else " − "
    return f"{op.upper()}( {joiner.join(kids)} )"


def _declared_rows(template: dict, stype: str) -> list[dict]:
    """The statement's rows straight from the template, in declaration order.

    ``_build_statement`` serves NOTHING for a statement the filing did not present — right for the
    screen, which should not show an empty profit and loss for a balance-sheet-only filing, but wrong
    for a structure review, where the unpresented statements are precisely what needs checking. The
    shape matches what the grid renders, so a sheet built this way reads identically; only the
    figures are absent, and the cover says which sheets came from which source.
    """
    out: list[dict] = []

    def walk(nodes: list[dict]) -> None:
        for n in nodes or []:
            key, role = n.get("canonical_key"), str(n.get("role") or "")
            if key:
                kind = role if role in ("subtotal", "total") else "item"
                out.append({"id": key, "label": n.get("label") or key, "kind": kind})
            else:
                out.append({"id": f"sec_{n.get('node_id') or ''}",
                            "label": n.get("label") or "", "kind": "section"})
            walk(n.get("children") or [])

    for stmt in template.get("statements") or []:
        if stmt.get("type") == stype:
            walk(stmt.get("sections") or [])
    return out


def _source(row: dict) -> str:
    prov = ((row.get("inspector") or {}).get("src")
            or (row.get("source") or {}).get("label") or "")
    return str(prov or "")


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--out", default="template_output_structure.xlsx")
    ap.add_argument("--basis", default="consolidated", choices=("consolidated", "standalone"))
    args = ap.parse_args()

    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    from app.api.routes.documents import _build_statement, _template_for_run
    from app.db.base import SessionLocal
    from app.db.models import ExtractionRun

    with SessionLocal() as session:
        run = (session.query(ExtractionRun)
               .filter(ExtractionRun.result.isnot(None))
               .order_by(ExtractionRun.created_at.desc()).first())
        if run is None:
            print("no extraction run with a result is stored — extract a document first")
            return 1
        template = _template_for_run(session, run)
        if not template:
            print(f"run {run.id} has no template attached, so it has no output structure")
            return 1
        rows = run.result.get("rows", [])
        rollups = _rollups(template)

        wb = openpyxl.Workbook()
        cover = wb.active
        cover.title = "How to read this"
        head_fill = PatternFill("solid", fgColor="1F3A5F")
        sect_fill = PatternFill("solid", fgColor="EEF1F5")
        calc_fill = PatternFill("solid", fgColor="FFF7E6")

        declared = {str(s.get("type")) for s in (template.get("statements") or [])}
        counts: list[tuple[str, int, int, str]] = []
        for stype in _STATEMENTS:
            if stype not in declared:
                continue                      # the template does not declare this statement at all
            st = _build_statement(rows, template, stype, run.document_id or "", args.basis, "en",
                                  run.result.get("units"), None, "pdf", 0,
                                  run.result.get("netting") or [])
            srows = st.get("rows") or []
            origin = "rendered from the run"
            if not srows:
                # This filing did not present the statement, so the grid serves nothing for it. The
                # template still declares it, and that structure is what a revision needs.
                srows, origin = _declared_rows(template, stype), "template declaration only"
            if not srows:
                continue
            ws = wb.create_sheet(_SHEET.get(stype, stype)[:31])
            ws.append(_HEAD)
            for c in range(1, len(_HEAD) + 1):
                cell = ws.cell(1, c)
                cell.font = Font(bold=True, color="FFFFFF", size=9)
                cell.fill = head_fill
                cell.alignment = Alignment(vertical="center", wrap_text=True)

            populated = 0
            for i, r in enumerate(srows, start=1):
                kind = str(r.get("kind") or "item")
                key = str(r.get("id") or "")
                rollup = rollups.get(key)
                conf = r.get("confidence") or {}
                if r.get("v1") is not None or r.get("v2") is not None:
                    populated += 1
                ws.append([
                    i, kind, str(r.get("label") or ""),
                    "" if kind == "section" else key,
                    str(r.get("note") or ""),
                    r.get("v1"), r.get("v2"),
                    str(r.get("status") or ("" if kind == "section" else "extracted"
                                            if r.get("v1") is not None else "")),
                    "yes" if rollup else "", _formula(rollup),
                    f"{conf.get('pct')}%" if conf.get("pct") is not None else "",
                    _source(r),
                ])
                row_i = ws.max_row
                if kind == "section":
                    for c in range(1, len(_HEAD) + 1):
                        ws.cell(row_i, c).fill = sect_fill
                    ws.cell(row_i, 3).font = Font(bold=True, size=10)
                elif kind in ("subtotal", "total"):
                    ws.cell(row_i, 3).font = Font(bold=True)
                    if rollup:
                        ws.cell(row_i, 10).fill = calc_fill
            for col, width in zip("ABCDEFGHIJKL",
                                  (5, 10, 46, 42, 6, 14, 14, 11, 11, 70, 11, 14)):
                ws.column_dimensions[col].width = width
            ws.freeze_panes = "C2"
            counts.append((_SHEET.get(stype, stype), len(srows), populated, origin))

        # The cover explains what the reader is looking at, and is honest about the figures.
        cover["A1"] = "Template output structure — as rendered on screen"
        cover["A1"].font = Font(bold=True, size=14)
        lines = [
            "",
            f"Template:      {template.get('name')}  ({template.get('template_key')})",
            f"Built from:    extraction run {run.id}",
            f"Basis column:  {args.basis}",
            "",
            "One sheet per statement. Rows are in SCREEN ORDER and include every line the template",
            "declares — including lines this filing never populated, because a line nobody extracted",
            "is still a line the template asks for, and that is what a revision needs to see.",
            "",
            "Columns",
            "  Kind             section / item / subtotal / total, as the grid groups them.",
            "  Label            exactly the text the screen shows for that row.",
            "  Canonical key    the identifier the ontology maps a printed caption onto.",
            "  Current / Prior  the extracted figures for the basis above. Blank = not extracted.",
            "  Status           blank when extracted; 'missing' when the template asks for a line",
            "                   this filing did not yield.",
            "  Calculated       'yes' where the template DERIVES the figure rather than reading it.",
            "  Formula          that derivation, from the template's own rollup declaration — shown",
            "                   whether or not this filing had the inputs to compute it.",
            "",
            "Statements in this workbook",
        ]
        for name, total, populated, origin in counts:
            lines.append(f"  {name:22} {total:4} rows,  {populated:3} carrying a figure   "
                         f"({origin})")
        missing = [s for s in _STATEMENTS if s not in declared]
        if missing:
            lines += ["",
                      "This template declares no " + ", ".join(_SHEET.get(m, m) for m in missing)
                      + " — so the product has no output structure for it. Worth knowing if you",
                      "expected one."]
        lines += [
            "",
            "NOTE ON THE FIGURES: they come from the most recent extraction stored, which is a thin",
            "sample filing — most lines are therefore blank. The STRUCTURE is complete; the numbers",
            "are only there to show where they land.",
        ]
        for i, text in enumerate(lines, start=2):
            cover.cell(i, 1, text).font = Font(size=10, name="Consolas" if text.startswith("  ")
                                               else "Calibri")
        cover.column_dimensions["A"].width = 100

        wb.save(args.out)
        print(f"wrote {args.out}")
        for name, total, populated, origin in counts:
            print(f"  {name:22} {total:4} rows, {populated:3} with a figure  ({origin})")
    return 0


if __name__ == "__main__":
    sys.exit(main())
