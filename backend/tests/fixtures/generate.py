"""Synthetic fixture generators with known ground truth.

These build documents whose contents we control exactly, so extraction accuracy,
routing, and integrity checks can be asserted automatically without any real
(sensitive) financial statements. reportlab/openpyxl are dev-only dependencies.
"""
from __future__ import annotations

import io


def make_native_pdf(title: str = "Balance Sheet") -> bytes:
    """A single-page native (text-layer) PDF containing a small balance sheet."""
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    width, height = A4
    y = height - 72
    c.setFont("Helvetica-Bold", 14)
    c.drawString(72, y, title)
    c.setFont("Helvetica", 10)
    rows = [
        ("Cash and cash equivalents", "Note 14", "1,204"),
        ("Trade receivables", "Note 15", "3,410"),
        ("Property, plant and equipment", "Note 5", "12,800"),
        ("Total assets", "", "17,414"),
    ]
    for label, note, value in rows:
        y -= 24
        c.drawString(72, y, label)
        c.drawString(320, y, note)
        c.drawRightString(500, y, value)
    c.showPage()
    c.save()
    return buf.getvalue()


def make_dual_basis_pdf() -> bytes:
    """A native PDF whose columns are a two-level Consolidated | Standalone header, each with
    a current + prior period — for testing consolidated+standalone extraction in one pass."""
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    _, height = A4
    y = height - 72
    c.setFont("Helvetica-Bold", 13)
    c.drawString(72, y, "Balance Sheet")
    # Two-level column header: Consolidated over cols ~300/370, Standalone over ~440/510.
    c.setFont("Helvetica-Bold", 9)
    y -= 22
    c.drawString(300, y, "Consolidated")
    c.drawString(445, y, "Standalone")
    c.setFont("Helvetica", 9)
    y -= 14
    c.drawRightString(330, y, "2025")
    c.drawRightString(400, y, "2024")
    c.drawRightString(475, y, "2025")
    c.drawRightString(545, y, "2024")
    c.setFont("Helvetica", 10)
    rows = [
        ("Trade receivables", ("3,410", "2,900", "3,100", "2,700")),
        ("Cash and cash equivalents", ("1,204", "980", "1,050", "900")),
    ]
    for label, (cc, cp, sc, sp) in rows:
        y -= 22
        c.drawString(72, y, label)
        c.drawRightString(330, y, cc)
        c.drawRightString(400, y, cp)
        c.drawRightString(475, y, sc)
        c.drawRightString(545, y, sp)
    c.showPage()
    c.save()
    return buf.getvalue()


def make_group_company_pdf() -> bytes:
    """A native PDF headed the way an HKEX filing heads its balance sheet: "Group" over one pair
    of dated columns and "Company" over another, with the scale declared in the statement header.

    This is the layout the engine used to read as single-basis — the Company's figures were filed
    as the Group's and added to them — because the detector only knew the words "Consolidated" and
    "Standalone"."""
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    _, height = A4
    y = height - 72
    c.setFont("Helvetica-Bold", 13)
    c.drawString(72, y, "Consolidated Statement of Financial Position")
    c.setFont("Helvetica-Bold", 9)
    y -= 24
    c.drawString(300, y, "Group")
    c.drawString(445, y, "Company")
    c.setFont("Helvetica", 9)
    y -= 14
    for x, text in ((330, "2024"), (400, "2023"), (475, "2024"), (545, "2023")):
        c.drawRightString(x, y, text)
    y -= 12
    c.drawRightString(330, y, "RMB'000")
    c.setFont("Helvetica", 10)
    rows = [
        ("Investments in subsidiaries", ("", "", "8,000", "7,500")),
        ("Trade receivables", ("3,410", "2,900", "310", "270")),
        ("Cash and cash equivalents", ("1,204", "980", "105", "90")),
        ("Total assets", ("4,614", "3,880", "8,415", "7,860")),
    ]
    for label, cells in rows:
        y -= 22
        c.drawString(72, y, label)
        for x, text in zip((330, 400, 475, 545), cells):
            if text:
                c.drawRightString(x, y, text)
    c.showPage()
    c.save()
    return buf.getvalue()


def make_group_and_company_note_pdf() -> bytes:
    """A NOTES page whose prose names both entities in one sentence — "The Group and the Company
    had no material contingent liabilities" — above a small table.

    The negative case for basis banding: a sentence that mentions both is not a two-basis column
    header, and reading it as one splits the comparative so last year's figures are reported as
    this year's for a second entity."""
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    _, height = A4
    y = height - 72
    c.setFont("Helvetica-Bold", 12)
    c.drawString(72, y, "34. Contingent liabilities")
    c.setFont("Helvetica", 10)
    y -= 22
    c.drawString(72, y, "The Group and the Company had no material contingent liabilities as at")
    y -= 16
    c.drawString(72, y, "31 December 2024 other than the guarantees set out below.")
    y -= 24
    for x, text in ((430, "2024"), (510, "2023")):
        c.drawRightString(x, y, text)
    for label, cur, pri in (("Guarantees to banks", "1,200", "900"),
                            ("Guarantees to suppliers", "450", "500"),
                            ("Other guarantees", "30", "40"),
                            ("Total guarantees", "1,680", "1,440")):
        y -= 22
        c.drawString(72, y, label)
        c.drawRightString(430, y, cur)
        c.drawRightString(510, y, pri)
    c.showPage()
    c.save()
    return buf.getvalue()


def make_company_limited_header_pdf() -> bytes:
    """A single-basis face page whose RUNNING HEADER names a "Group" on the left and ends in
    "Company Limited" on the right, where the value columns are.

    The other negative case, and the one that isolates the geometric test: both words a two-basis
    header needs are on one line, in separate runs, with no amount on it — and the page still has
    one basis, because a basis caption bands the figures it stands OVER and "Group" here stands
    over the caption column. Banding on the filer's own name would split the comparative, so last
    year's figures would be reported as this year's for a second entity."""
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    _, height = A4
    y = height - 50
    c.setFont("Helvetica", 8)
    c.drawString(72, y, "Sunrise Group Holdings Limited")
    c.drawRightString(523, y, "Sunrise Company Limited")
    c.setFont("Helvetica-Bold", 13)
    y -= 26
    c.drawString(72, y, "Statement of Financial Position")
    c.setFont("Helvetica", 9)
    y -= 22
    c.drawRightString(430, y, "2024")
    c.drawRightString(510, y, "2023")
    c.setFont("Helvetica", 10)
    for label, cur, pri in (("Inventories", "2,000", "1,800"),
                            ("Trade receivables", "3,410", "2,900"),
                            ("Cash and cash equivalents", "1,204", "980"),
                            ("Total assets", "6,614", "5,680")):
        y -= 22
        c.drawString(72, y, label)
        c.drawRightString(430, y, cur)
        c.drawRightString(510, y, pri)
    c.showPage()
    c.save()
    return buf.getvalue()


def make_rich_pdf() -> bytes:
    """A fuller native PDF: BS totals + P&L headline + a notes page with qualitative
    disclosures — so derived ratios, the disclosure scan, and free-form notes all populate."""
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    _, height = A4

    def rows(items, y):
        c.setFont("Helvetica", 10)
        for label, cur, pri in items:
            c.drawString(72, y, label)
            c.drawRightString(430, y, cur)
            c.drawRightString(510, y, pri)
            y -= 22
        return y

    c.setFont("Helvetica-Bold", 13); c.drawString(72, height - 60, "Balance Sheet")
    y = rows([
        ("Inventories", "2,000", "1,800"),
        ("Trade receivables", "3,410", "2,900"),
        ("Cash and cash equivalents", "1,204", "980"),
        ("Total current assets", "6,614", "5,680"),
        ("Total current liabilities", "3,300", "3,100"),
        ("Total non-current liabilities", "1,200", "1,400"),
        ("Total equity", "9,114", "7,180"),
        ("Total assets", "13,614", "11,680"),
    ], height - 90)
    c.showPage()

    c.setFont("Helvetica-Bold", 13); c.drawString(72, height - 60, "Statement of Profit or Loss")
    rows([
        ("Revenue from operations", "20,000", "18,000"),
        ("Operating profit", "3,200", "2,700"),
        ("Profit for the year", "2,400", "1,950"),
    ], height - 90)
    c.showPage()

    c.setFont("Helvetica-Bold", 12); c.drawString(72, height - 60, "Notes to the Financial Statements")
    c.setFont("Helvetica", 9)
    for i, line in enumerate([
        "The auditor has issued a qualified opinion in respect of inventory valuation.",
        "Contingent liabilities: the Group is subject to legal proceedings estimated at 500.",
        "The Company has provided financial guarantees to subsidiaries totalling 1,200.",
        "Related party transactions with associates are disclosed in note 28.",
        "Subsequent events: a dividend was declared after the reporting period.",
    ]):
        c.drawString(72, height - 90 - i * 18, line)
    c.showPage()
    c.save()
    return buf.getvalue()


def make_multipage_pdf() -> bytes:
    """A 2-page native PDF: face on page 0, notes on page 1."""
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    width, height = A4

    c.setFont("Helvetica-Bold", 14)
    c.drawString(72, height - 72, "Statement of Financial Position")
    c.setFont("Helvetica", 10)
    c.drawString(72, height - 100, "Cash and cash equivalents      Note 14      1,204")
    c.showPage()

    c.setFont("Helvetica-Bold", 14)
    c.drawString(72, height - 72, "Notes to the Financial Statements")
    c.setFont("Helvetica", 10)
    c.drawString(72, height - 100, "Note 14: Cash and cash equivalents")
    c.drawString(72, height - 120, "Cash on hand      204")
    c.drawString(72, height - 140, "Balances with banks      1,000")
    c.showPage()
    c.save()
    return buf.getvalue()


def make_annual_report_pdf() -> bytes:
    """A realistic HKEX/IFRS-shaped annual report: cover, an auditor's report (which mentions
    face phrases in prose), three *consolidated* face statements, then a notes section that
    opens WITHOUT a 'Notes to…' banner — straight into '1 General information' — followed by a
    numbered continuation note that also mentions 'profit or loss' in prose. Ground truth:
    the three statement pages are FACE, the auditor page is not FACE, and the two note pages
    are NOTES (so their tables are extracted, not skipped)."""
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    _, height = A4

    def page(title, body_lines, title_size=14):
        c.setFont("Helvetica-Bold", title_size)
        c.drawString(72, height - 72, title)
        c.setFont("Helvetica", 10)
        y = height - 104
        for ln in body_lines:
            c.drawString(72, y, ln); y -= 20
        c.showPage()

    # 0: cover / contents (OTHER)
    page("Annual Report 2024", ["Contents", "Corporate information", "Chairman's statement"])
    # 1: independent auditor's report — prose mentions "profit or loss" / "cash flows" but is
    #    NOT a statement face (and precedes the statements).
    page("Independent Auditor's Report", [
        "In our opinion the consolidated financial statements give a true and fair view.",
        "We audited the statement of profit or loss and the statement of cash flows.",
        "Key audit matters were addressed in forming our opinion."])
    # 2-4: the three consolidated face statements (FACE)
    page("Consolidated Statement of Profit or Loss", [
        "Revenue      Note 5      45,230",
        "Cost of sales      (28,110)",
        "Profit for the year      6,120"])
    page("Consolidated Statement of Financial Position", [
        "Property, plant and equipment      Note 12      88,400",
        "Trade receivables      Note 15      12,300",
        "Cash and cash equivalents      Note 18      9,870",
        "Total assets      143,900"])
    page("Consolidated Statement of Cash Flows", [
        "Net cash from operating activities      14,200",
        "Net cash used in investing activities      (9,100)"])
    # 5: notes section opens with a numbered heading, NO 'Notes to…' banner (exercises the
    #    'first numbered note after the face statements' start rule).
    page("1 General information", [
        "The Company is incorporated in the Cayman Islands with limited liability.",
        "Its shares are listed on The Stock Exchange of Hong Kong Limited."])
    # 6: a numbered continuation note that mentions a face phrase in prose (must stay NOTES).
    page("18 Cash and cash equivalents", [
        "Cash at banks and on hand      9,870",
        "Amounts are measured at amortised cost; see the statement of profit or loss for interest.",
        "Short-term deposits      2,400"])
    c.save()
    return buf.getvalue()


def make_hk_running_header_report_pdf() -> bytes:
    """Reproduces the structure of a real HK/PRC-listed annual report that trips naive
    classifiers: a bilingual RUNNING HEADER on every page, a Financial Highlights page that
    quotes 'Summary of Statement of Profit or Loss' (must NOT be a face), an auditor's report
    that mentions statements in prose, statement titles SPLIT across two lines, a 'Notes to
    Consolidated Financial Statements' banner, and a Five Year Financial Summary in back-matter.
    Ground truth: highlights + auditor = OTHER, the three statements = FACE, the two note pages =
    NOTES, the five-year summary = OTHER."""
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    _, height = A4
    n = [0]

    def page(title_lines, body_lines):
        y = height - 56
        c.setFont("Helvetica", 8)
        c.drawString(60, y, f"{n[0]:02d}")                       # page number
        c.drawString(90, y, "Acme Holdings Limited / Annual Report 2024")   # running header EN
        c.setFont("Helvetica", 8)
        c.drawString(90, y - 12, "艾克美控股有限公司 / 二零二四年年報")          # running header ZH
        y -= 40
        c.setFont("Helvetica-Bold", 13)
        for tl in title_lines:                                    # the real page title (may split)
            c.drawString(72, y, tl); y -= 18
        c.setFont("Helvetica", 10)
        y -= 6
        for bl in body_lines:
            c.drawString(72, y, bl); y -= 18
        c.showPage(); n[0] += 1

    page(["Annual Report 2024"], ["Corporate information", "Contents"])                    # 0 cover
    page(["Financial Highlights", "Summary of Statement of Profit or Loss"],               # 1 highlights (OTHER)
         ["Revenue      45,230", "Profit for the year      6,120"])
    page(["Independent Auditor's Report"],                                                 # 2 auditor (OTHER)
         ["In our opinion the consolidated financial statements give a true and fair view.",
          "We audited the consolidated statement of profit or loss and the statement of cash flows."])
    page(["Consolidated Statement of", "Profit or Loss"],                                  # 3 FACE (split title)
         ["Revenue      Note 5      45,230", "Profit for the year      6,120"])
    page(["Consolidated Statement of", "Financial Position"],                              # 4 FACE (split title)
         ["Property, plant and equipment      Note 12      88,400",
          "Cash and cash equivalents      Note 18      9,870", "Total assets      143,900"])
    page(["Consolidated Statement of", "Cash Flows"],                                      # 5 FACE (split title)
         ["Net cash from operating activities      14,200"])
    page(["Notes to Consolidated", "Financial Statements"],                                # 6 NOTES (banner)
         ["1. Corporate and group information",
          "The Company is incorporated in the Cayman Islands with limited liability."])
    page(["18 Cash and cash equivalents"],                                                 # 7 NOTES (continuation)
         ["Cash at banks and on hand      9,870",
          "See the statement of profit or loss for related interest income."])
    page(["Five Year Financial Summary"],                                                  # 8 back-matter (OTHER)
         ["Revenue      45,230      40,110      38,900      35,000      31,200"])
    c.save()
    return buf.getvalue()


def make_xlsx() -> bytes:
    """A workbook with a negative-number format and a hidden sheet."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"
    ws["A1"] = "Cash and cash equivalents"
    ws["B1"] = 1204
    ws["A2"] = "Accumulated depreciation"
    ws["B2"] = -500
    ws["B2"].number_format = "#,##0;(#,##0)"

    hidden = wb.create_sheet("Working")
    hidden.sheet_state = "hidden"
    hidden["A1"] = "scratch"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def make_comparative_pdf() -> bytes:
    """A two-column comparative balance sheet — this year beside last year.

    Deliberately includes a line printed for the PRIOR YEAR ONLY ("Pledged deposits", released
    during the year). That row is the one a positional reader mis-files as the current year, and
    it is also the case a reviewer has to be able to click through to last year's page for.
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    _, height = A4
    y = height - 72
    c.setFont("Helvetica-Bold", 13)
    c.drawString(72, y, "Consolidated Statement of Financial Position")
    y -= 20
    c.setFont("Helvetica-Bold", 9)
    c.drawRightString(430, y, "31 December 2024")
    c.drawRightString(520, y, "31 December 2023")
    c.setFont("Helvetica", 10)
    rows = [
        ("Property, plant and equipment", "Note 5", "12,800", "11,400"),
        ("Inventories", "Note 8", "2,150", "1,980"),
        ("Trade receivables", "Note 15", "3,410", "3,120"),
        ("Cash and cash equivalents", "Note 14", "1,204", "990"),
        ("Pledged deposits", "Note 14", "", "2,031"),          # prior year only
        ("Total assets", "", "19,564", "21,521"),
    ]
    for label, note, cur, prior in rows:
        y -= 24
        c.drawString(72, y, label)
        c.drawString(300, y, note)
        if cur:
            c.drawRightString(430, y, cur)
        if prior:
            c.drawRightString(520, y, prior)
    c.showPage()
    c.save()
    return buf.getvalue()
