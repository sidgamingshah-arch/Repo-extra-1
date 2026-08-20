"""The eight analyst buckets: every face row and every note lands in exactly one of them.

WHAT MAKES THIS STORE WORTH HAVING, and the property each test below defends: an analyst can ask
"show me everything this filing says about current liabilities" and get the face rows AND the notes
behind them, with nothing counted twice and nothing quietly dropped. Two ways that goes wrong and
neither is visible in a total:

* a row in two buckets — the buckets then sum to more than the statement;
* a row in no bucket, or swept into Others unnoticed — the buckets sum to less, and the reader has
  no way to tell that from a filing that simply did not state the section.

So the invariants are partition invariants, and ``unresolved_face_item_ids`` is the measurement that
stops "everything is placed" being achieved by placing everything in Others.
"""
from __future__ import annotations

import json
import pathlib
from decimal import Decimal

import pytest

from app.core.models.document import DocumentModel, PageSource
from app.core.models.enums import Basis, LineRole, PageKind
from app.core.models.geometry import Provenance
from app.core.models.line_item import (
    ExtractedValue, LineItem, NoteItem, NoteRef, NotesTable,
)
from app.core.stage import PipelineContext
from app.schemas.loader import load_ontology
from app.services.buckets import (
    BUCKET_KEYS, bucket_of, section_token, segment_source, statement_of_section,
)
from app.stages.segment import SegmentStage

_SAMPLES = pathlib.Path(__file__).resolve().parent.parent / "app" / "sample" / "templates"


@pytest.fixture(scope="module")
def ontology():
    raw = json.loads((_SAMPLES / "hkfrs_hk_china_ontology.json").read_text())
    return load_ontology(raw, resolve=True)


def _li(ordinal: int, label: str, key: str | None, value: int, page: int = 0,
        role: LineRole = LineRole.LINE, notes: list[str] | None = None) -> LineItem:
    item = LineItem(source_label=label, canonical_key=key, ordinal=ordinal, role=role)
    if notes:
        item.note_refs = [NoteRef(raw=",".join(notes), numbers=list(notes))]
    item.set_value(ExtractedValue(
        value=Decimal(value), value_raw=Decimal(value), basis=Basis.CONSOLIDATED,
        period_label="current", provenance=Provenance(page_index=page)))
    return item


def _doc(items: list[LineItem], pages: dict[int, tuple[str | None, PageKind]],
         notes: list[NotesTable] | None = None) -> DocumentModel:
    doc = DocumentModel(filename="f.pdf", locale="en")
    doc.pages = [PageSource(index=i, statement=stmt, kind=kind)
                 for i, (stmt, kind) in sorted(pages.items())]
    doc.line_items = items
    doc.notes = notes or []
    return doc


def _note(number: str, title: str, page: int, keys: list[str | None]) -> NotesTable:
    return NotesTable(note_number=number, title=title, source_pages=[page],
                      items=[NoteItem(raw_label=f"{title} {i}", canonical_key=k, ordinal=i)
                             for i, k in enumerate(keys)])


# --- the resolution rule -----------------------------------------------------------------------

def test_the_five_balance_sheet_sections_each_get_their_own_bucket():
    """The reason the segmentation cannot be done at page classification, stated as a test: these
    five are printed on ONE page, so the page's statement separates none of them and only the row's
    own resolved section can."""
    assert bucket_of("bs_s1_non_current_assets", "balance_sheet")[0] == "non_current_assets"
    assert bucket_of("bs_s2_current_assets", "balance_sheet")[0] == "current_assets"
    assert bucket_of("bs_s3_current_liabilities", "balance_sheet")[0] == "current_liabilities"
    assert bucket_of("bs_s4_non_current_liabilities", "balance_sheet")[0] == (
        "non_current_liabilities")
    assert bucket_of("bs_s5_equity", "balance_sheet")[0] == "equity"


def test_a_whole_statement_resolves_without_consulting_its_sections():
    """P&L and cash flow are ONE bucket each, so every section of them lands in the same place and
    the taxonomy does not need to know that ``pl_s4_exceptional_items`` exists."""
    for section in ("pl_s1_income", "pl_s4_exceptional_items", "pl_top_level"):
        assert bucket_of(section, "profit_and_loss") == ("profit_and_loss", "statement")
    for section in ("cf_s1_cash_flow_from_operating_activities", "cf_top_level"):
        assert bucket_of(section, "cash_flow") == ("cash_flow", "statement")
    # The statement of changes in equity is that section's movement, not a ninth bucket.
    assert bucket_of(None, "equity_changes") == ("equity", "statement")


def test_the_balance_sheets_own_totals_are_others_and_say_so():
    """Total assets spans two buckets and total equity-and-liabilities spans three, so no section
    bucket can hold them without being wrong. Others is the right answer and ``statement_total`` is
    what distinguishes it from a row nothing could place."""
    assert bucket_of("bs_top_level", "balance_sheet") == ("others", "statement_total")
    assert bucket_of(None, None) == ("others", "unresolved")


def test_a_section_this_taxonomy_does_not_know_is_named_not_swallowed():
    """A template with a sixth balance-sheet section would otherwise put a whole section of a filing
    into Others, and the buckets would still sum to the statement — the failure is invisible in
    every total. The reason is carried out so the stage can report the section by name."""
    assert bucket_of("bs_s9_revaluation_surplus", "balance_sheet") == ("others", "unknown_section")


def test_the_section_bucket_edge_is_derived_from_the_id_not_tabulated(ontology):
    """Every section the SHIPPED rulebook declares resolves to a real bucket, and the balance-sheet
    ones resolve by their own section phrase — the same phrase ``mapping.HEADING_ROW_SECTIONS`` uses
    to recognise a printed banner. That shared vocabulary is what stops the rulebook and this layer
    drifting into two different ideas of what "current assets" means."""
    from app.services.rollups import section_members

    unknown = []
    for section in section_members(ontology):
        bucket, reason = bucket_of(section, statement_of_section(section))
        assert bucket in BUCKET_KEYS
        if reason == "unknown_section":
            unknown.append(section)
    assert unknown == [], f"the shipped rulebook declares sections with no bucket: {unknown}"
    assert section_token("bs_s2_current_assets") == "current_assets"


# --- the partition -----------------------------------------------------------------------------

def _mixed_filing() -> DocumentModel:
    """One balance-sheet page carrying four buckets, plus a P&L page, a cash-flow page and two
    notes pages — the shape that makes a page-level segmentation useless."""
    items = [
        _li(0, "Property, plant and equipment",
            "bs_non_current_assets__property_plant_and_equipment", 5000, page=0, notes=["14"]),
        _li(1, "Inventories", "bs_current_assets__inventories", 1200, page=0),
        _li(2, "Trade and bills payables",
            "bs_current_liabilities__current_trade_payables", 900, page=0, notes=["22"]),
        _li(3, "Interest-bearing bank borrowings",
            "bs_non_current_liabilities__non_current_borrowings", 2000, page=0),
        _li(4, "Share capital", "bs_equity__share_capital", 100, page=0),
        _li(5, "Total assets", "bs_total_assets", 6200, page=0, role=LineRole.TOTAL),
        _li(6, "A caption nothing placed", None, 42, page=0),
        _li(7, "Revenue", "pl_income__revenue_from_operations", 9000, page=1),
        _li(8, "Cost of goods sold", "pl_expenses__cost_of_goods_sold", -6000, page=1),
        _li(9, "Net cash from operating activities",
            "cf_cash_flow_from_operating_activities__net_cash_from_operating_activities",
            1500, page=2),
    ]
    notes = [
        _note("14", "Property, plant and equipment", 3,
              ["bs_non_current_assets__property_plant_and_equipment"]),
        _note("22", "Trade and other payables", 4,
              ["bs_current_liabilities__current_trade_payables"]),
    ]
    return _doc(items, {
        0: ("balance_sheet", PageKind.FACE),
        1: ("profit_and_loss", PageKind.FACE),
        2: ("cash_flow", PageKind.FACE),
        3: (None, PageKind.NOTES),
        4: (None, PageKind.NOTES),
    }, notes)


def test_every_face_row_lands_in_exactly_one_bucket(ontology):
    """The partition invariant. Summed over the buckets, the face rows equal the filing's face rows
    — no row counted twice, no row lost."""
    doc = _mixed_filing()
    store = segment_source(doc, ontology)

    placed = [i for seg in store.segments for i in seg.face_item_ids]
    assert len(placed) == len(set(placed)), "a row is in two buckets"
    assert set(placed) == {str(li.id) for li in doc.line_items}


def test_one_balance_sheet_page_is_split_across_four_buckets_and_equity(ontology):
    """The whole point of the change: a single page's rows reach five different buckets, which no
    page-level classification could do."""
    store = segment_source(_mixed_filing(), ontology)
    counts = {seg.bucket: len(seg.face_item_ids) for seg in store.segments}

    assert counts == {
        "non_current_assets": 1, "current_assets": 1, "non_current_liabilities": 1,
        "current_liabilities": 1, "equity": 1,
        "profit_and_loss": 2, "cash_flow": 1,
        # Total assets (spans sections) and the caption nothing placed.
        "others": 2,
    }


def test_a_row_nothing_placed_is_measurable_and_not_hidden_in_others(ontology):
    """Others holds two rows for two different reasons, and only one of them is a coverage failure.
    Without this list a filing whose sections were half-read would look exactly like one whose
    balance sheet simply prints its own totals."""
    doc = _mixed_filing()
    store = segment_source(doc, ontology)
    unplaced = {li.source_label for li in doc.line_items
                if str(li.id) in set(store.unresolved_face_item_ids)}

    assert unplaced == {"A caption nothing placed"}
    # …and the statement total is in Others WITHOUT being called unresolved.
    total = next(li for li in doc.line_items if li.source_label == "Total assets")
    assert str(total.id) in set(store.segment("others").face_item_ids)
    assert str(total.id) not in set(store.unresolved_face_item_ids)


def test_a_note_is_filed_under_the_bucket_whose_face_row_cites_it(ontology):
    """The face citation is the strongest signal there is: a note titled "Trade and other payables"
    is current liabilities because the current-liabilities line points at it, whatever its own rows
    did or did not map to."""
    store = segment_source(_mixed_filing(), ontology)

    assert store.segment("non_current_assets").note_numbers == ["14"]
    assert store.segment("current_liabilities").note_numbers == ["22"]
    assert store.segment("others").note_numbers == []
    # Notes pages are attributed to the bucket that owns the note, not left on the filing.
    assert store.segment("non_current_assets").note_pages == [3]
    assert store.segment("current_liabilities").note_pages == [4]


def test_a_note_two_buckets_cite_is_filed_in_both_and_marked_shared(ontology):
    """Borrowings split across current and non-current routinely share one note, and BOTH sections
    need it: an analyst reading current liabilities cannot be sent to another bucket to find the
    breakdown of a figure printed in front of them.

    The cost is that the note's figures now appear twice across the store, so ``shared_notes`` marks
    it in every bucket holding it — a caller adding the buckets up can subtract the overlap, and the
    duplication is a stated fact rather than a silent one.
    """
    items = [
        _li(0, "Bank borrowings (non-current)",
            "bs_non_current_liabilities__non_current_borrowings", 2000, page=0, notes=["25"]),
        _li(1, "Bank borrowings (current)",
            "bs_current_liabilities__current_borrowings", 500, page=0, notes=["25"]),
        _li(2, "Lease liabilities (current)",
            "bs_current_liabilities__current_lease_liabilities", 80, page=0, notes=["25"]),
    ]
    doc = _doc(items, {0: ("balance_sheet", PageKind.FACE), 1: (None, PageKind.NOTES)},
               [_note("25", "Interest-bearing bank borrowings", 1, [None])])
    store = segment_source(doc, ontology)

    # In both, however lopsided the citation count is — two current rows against one non-current.
    assert store.segment("current_liabilities").note_numbers == ["25"]
    assert store.segment("non_current_liabilities").note_numbers == ["25"]
    # …and marked in both, which is what makes the overlap subtractable.
    assert store.segment("current_liabilities").shared_notes == ["25"]
    assert store.segment("non_current_liabilities").shared_notes == ["25"]
    # The note's page reaches both buckets too, so either one can open it.
    assert store.segment("current_liabilities").note_pages == [1]
    assert store.segment("non_current_liabilities").note_pages == [1]
    # Two filings of one note: the notes side is deliberately NOT a partition.
    assert sum(len(s.note_numbers) for s in store.segments) == 2


def test_a_note_only_one_bucket_cites_is_not_marked_shared(ontology):
    """The other side, so ``shared_notes`` means something: a note with one citing bucket carries no
    overlap to subtract, and marking it would make every note look duplicated."""
    store = segment_source(_mixed_filing(), ontology)

    assert store.segment("non_current_assets").note_numbers == ["14"]
    assert store.segment("non_current_assets").shared_notes == []
    assert store.segment("current_liabilities").shared_notes == []


def test_a_note_no_face_row_cites_is_placed_from_its_own_rows(ontology):
    """The fallback, and it has to reach the P&L and cash-flow buckets too — a note on operating
    expenses is printed on a notes page, so there is no statement to read and the section token
    alone ("expenses") answers only for the balance sheet."""
    doc = _doc([], {0: (None, PageKind.NOTES)}, [
        _note("8", "Other operating expenses", 0, ["pl_expenses__other_operating_costs"]),
        _note("17", "Inventories", 0, ["bs_current_assets__inventories"]),
        _note("31", "A note nothing mapped", 0, [None]),
    ])
    store = segment_source(doc, ontology)

    assert store.segment("profit_and_loss").note_numbers == ["8"]
    assert store.segment("current_assets").note_numbers == ["17"]
    assert store.segment("others").note_numbers == ["31"]
    # …and the one that reached Others by failure, not by belonging, is named.
    assert store.unresolved_note_numbers == ["31"]


def test_a_row_printed_inside_a_note_is_not_also_a_face_row(ontology):
    """``residual._sweep_notes`` synthesises face rows FROM note items, and a note's own breakdown
    reaching the face store as well would count the note's figures twice — once in the note, once
    as a row of the section it belongs to."""
    inside = _li(0, "Buildings", "bs_non_current_assets__property_plant_and_equipment", 3000,
                 page=3)
    inside.note_number = "14"
    doc = _doc([inside], {3: (None, PageKind.NOTES)},
               [_note("14", "Property, plant and equipment", 3, [None])])
    store = segment_source(doc, ontology)

    assert [i for seg in store.segments for i in seg.face_item_ids] == []


# --- the stage ---------------------------------------------------------------------------------

def test_the_stage_records_the_segmentation_and_reports_what_it_could_not_place(ontology):
    doc = _mixed_filing()
    ctx = PipelineContext(raw_bytes=b"")
    ctx.ontology = ontology
    SegmentStage().run(doc, ctx)

    assert doc.buckets is not None
    assert sum(len(s.face_item_ids) for s in doc.buckets.segments) == len(doc.line_items)
    assert any("segment:unresolved(1 face rows" in m for m in ctx.logs), ctx.logs
    assert any("segment:current_liabilities(1 rows, 1 notes)" == m for m in ctx.logs), ctx.logs


def test_the_stage_survives_a_run_with_no_rulebook(ontology):
    """Sections come from the rulebook. Without one, every row falls back to its page's statement —
    P&L and cash flow still resolve, the balance sheet's five buckets cannot, and the rows say so
    instead of being silently distributed."""
    doc = _mixed_filing()
    ctx = PipelineContext(raw_bytes=b"")
    SegmentStage().run(doc, ctx)

    counts = {s.bucket: len(s.face_item_ids) for s in doc.buckets.segments}
    assert counts["profit_and_loss"] == 2 and counts["cash_flow"] == 1
    # Every balance-sheet row is in Others, and every one of them is reported unresolved.
    assert counts["others"] == 7
    assert len(doc.buckets.unresolved_face_item_ids) == 7


def test_no_fixture_in_this_module_names_a_concept_the_rulebook_does_not_have(ontology):
    """Written because two of the fixtures above did exactly that, and the symptom was misleading
    rather than loud: an invented key resolves to no section, so the row reported as "nothing placed
    it" and the counts looked like a defect in the resolver instead of a typo in the test."""
    known = {m.canonical_key for m in ontology.mappings}
    used = {li.canonical_key for li in _mixed_filing().line_items if li.canonical_key}
    for note in _mixed_filing().notes:
        used |= {i.canonical_key for i in note.items if i.canonical_key}
    assert used <= known, f"fixtures name concepts the rulebook does not have: {sorted(used - known)}"


# --- the endpoints -----------------------------------------------------------------------------

def _seed_run(result: dict, filename: str = "buckets.pdf") -> str:
    """A document with one stored succeeded run carrying `result`.

    Seeded rather than extracted: the endpoints read the run out of the database, and what is under
    test is the JOIN between the stored segmentation and the stored rows — not the extraction that
    produced either.
    """
    import uuid

    from app.db.base import SessionLocal, init_db
    from app.db.models import Document, ExtractionRun

    init_db()
    with SessionLocal() as session:
        doc = Document(filename=filename, fmt="pdf", byte_size=1, page_count=5,
                       content_hash=uuid.uuid4().hex, object_key="k", owner="admin",
                       status="extracted")
        session.add(doc)
        session.flush()
        session.add(ExtractionRun(document_id=doc.id, status="succeeded", options={},
                                  result={"filename": filename, **result}))
        session.commit()
        return doc.id


def _segmented_result(ontology) -> dict:
    """The mixed filing, segmented, in the shape a real run stores."""
    doc = _mixed_filing()
    store = segment_source(doc, ontology)
    rows = [{"id": str(li.id), "source_label": li.source_label,
             "canonical_key": li.canonical_key, "role": li.role.value, "values": []}
            for li in doc.line_items]
    notes = [{"no": n.note_number, "title": n.title, "page": n.source_pages[0] + 1, "rows": []}
             for n in doc.notes]
    return {"rows": rows, "note_details": notes, "buckets": store.model_dump(mode="json")}


def test_the_index_serves_all_eight_buckets_in_reading_order(client, ontology):
    """Eight rows, always, in the order a filing is read — a bucket a filing happens not to state
    is served as zero rather than omitted, so the screen has a stable shape and an empty section is
    visibly empty instead of missing."""
    doc_id = _seed_run(_segmented_result(ontology))
    body = client.get(f"/api/v1/documents/{doc_id}/buckets").json()

    assert body["segmented"] is True
    assert [b["bucket"] for b in body["buckets"]] == list(BUCKET_KEYS)
    assert [b["label"] for b in body["buckets"]][:2] == ["Non-current assets", "Current assets"]
    counts = {b["bucket"]: b["face_rows"] for b in body["buckets"]}
    assert counts["current_liabilities"] == 1 and counts["profit_and_loss"] == 2
    # The measurement that stops Others reading as coverage.
    assert body["unresolved_face_rows"] == 1
    assert body["unknown_sections"] == []


def test_the_face_rows_summed_over_the_buckets_equal_the_runs_own_rows(client, ontology):
    """The partition, asserted through the API rather than in-process — a serialisation that
    dropped or duplicated a membership would not otherwise show up here."""
    result = _segmented_result(ontology)
    doc_id = _seed_run(result)
    body = client.get(f"/api/v1/documents/{doc_id}/buckets").json()

    assert sum(b["face_rows"] for b in body["buckets"]) == len(result["rows"])


def test_a_buckets_detail_serves_its_own_rows_and_notes(client, ontology):
    doc_id = _seed_run(_segmented_result(ontology))
    body = client.get(f"/api/v1/documents/{doc_id}/buckets/current_liabilities").json()

    assert body["label"] == "Current liabilities"
    assert [r["source_label"] for r in body["rows"]] == ["Trade and bills payables"]
    assert [n["no"] for n in body["notes"]] == ["22"]
    assert body["sections"] == ["bs_s3_current_liabilities"]
    # 1-based, like every page number this API serves.
    assert body["face_pages"] == [1] and body["note_pages"] == [5]


def test_the_detail_marks_a_row_that_reached_others_by_failing_to_place(client, ontology):
    """Others holds a statement total and a row nothing placed. A reader has to be able to tell
    them apart on the row, not just in a count, or the bucket looks like a rag-bag either way."""
    doc_id = _seed_run(_segmented_result(ontology))
    body = client.get(f"/api/v1/documents/{doc_id}/buckets/others").json()

    marked = {r["source_label"]: r["unresolved"] for r in body["rows"]}
    assert marked == {"Total assets": False, "A caption nothing placed": True}


def _shared_note_result(ontology) -> dict:
    """A borrowings note cited from both the current and non-current liability rows."""
    items = [
        _li(0, "Bank borrowings (non-current)",
            "bs_non_current_liabilities__non_current_borrowings", 2000, page=0, notes=["25"]),
        _li(1, "Bank borrowings (current)",
            "bs_current_liabilities__current_borrowings", 500, page=0, notes=["25"]),
    ]
    doc = _doc(items, {0: ("balance_sheet", PageKind.FACE), 1: (None, PageKind.NOTES)},
               [_note("25", "Interest-bearing bank borrowings", 1, [None])])
    store = segment_source(doc, ontology)
    rows = [{"id": str(li.id), "source_label": li.source_label,
             "canonical_key": li.canonical_key, "role": li.role.value, "values": []}
            for li in doc.line_items]
    notes = [{"no": "25", "title": "Interest-bearing bank borrowings", "page": 2,
              "rows": [{"label": "Bank loans", "values": []}]}]
    return {"rows": rows, "note_details": notes, "buckets": store.model_dump(mode="json")}


def test_both_buckets_serve_the_shared_notes_own_content(client, ontology):
    """The point of filing it in both: each bucket serves the note's rows, not a pointer. An analyst
    on either section reads the breakdown without leaving the section."""
    doc_id = _seed_run(_shared_note_result(ontology), "shared-note.pdf")

    for bucket in ("current_liabilities", "non_current_liabilities"):
        body = client.get(f"/api/v1/documents/{doc_id}/buckets/{bucket}").json()
        assert [n["no"] for n in body["notes"]] == ["25"], bucket
        assert body["notes"][0]["rows"] == [{"label": "Bank loans", "values": []}], bucket
        assert body["shared_notes"] == ["25"], bucket


def test_the_index_reconciles_the_duplicated_note_with_a_distinct_count(client, ontology):
    """Adding the per-bucket note counts is the WRONG total once a note is filed twice, so the index
    serves the filing-level answer beside them. Without it a screen showing "2 notes" for a filing
    with one note has no way to be right."""
    doc_id = _seed_run(_shared_note_result(ontology), "shared-note-count.pdf")
    body = client.get(f"/api/v1/documents/{doc_id}/buckets").json()

    assert sum(b["notes"] for b in body["buckets"]) == 2      # the duplicated filing
    assert body["distinct_notes"] == 1                        # what the filing actually has
    assert {b["bucket"] for b in body["buckets"] if b["shared_notes"]} == {
        "current_liabilities", "non_current_liabilities"}


def test_an_unknown_bucket_is_a_404_not_an_empty_bucket(client, ontology):
    doc_id = _seed_run(_segmented_result(ontology))
    assert client.get(f"/api/v1/documents/{doc_id}/buckets/goodwill").status_code == 404


def test_a_run_from_before_the_segment_stage_says_so_instead_of_serving_eight_empties(client):
    """A filing whose sections were never segmented and one whose sections are all empty are
    different answers, and only the first is fixed by re-running the extraction."""
    doc_id = _seed_run({"rows": [{"id": "x", "source_label": "Inventories"}]})

    index = client.get(f"/api/v1/documents/{doc_id}/buckets").json()
    assert index["segmented"] is False and index["buckets"] == []
    detail = client.get(f"/api/v1/documents/{doc_id}/buckets/current_assets").json()
    assert detail["segmented"] is False and detail["rows"] == []


# --- end to end --------------------------------------------------------------------------------

def _shipped_rulebook(client) -> str:
    """The SHIPPED rulebook's latest version id, chosen by key rather than by list position.

    ``ontologies[0]`` worked alone and failed in the suite: other modules publish their own
    rulebooks, so position selected a two-concept test fixture that mapped nothing — and the run
    still reported ``applied: true``, so the failure surfaced as "every row unplaced" rather than as
    "wrong rulebook".
    """
    rows = [o for o in client.get("/api/v1/ontologies").json()
            if o.get("ontology_key") == "hkfrs_hk_china"]
    assert rows, "the shipped rulebook is not seeded"
    return max(rows, key=lambda o: o.get("version") or 0)["id"]


def _extract(client, name: str, ontology_version_id: str | None = None) -> tuple[str, str]:
    from tests.fixtures.generate import make_native_pdf

    doc_id = client.post("/api/v1/documents",
                         files={"file": (name, make_native_pdf(),
                                         "application/pdf")}).json()["id"]
    body = {"ontology_version_id": ontology_version_id} if ontology_version_id else {}
    started = client.post(f"/api/v1/documents/{doc_id}/extractions", json=body)
    assert started.status_code == 202, started.text
    run_id = started.json()["run_id"]
    for _ in range(200):
        status = client.get(f"/api/v1/extractions/{run_id}").json()["status"]
        if status in {"succeeded", "failed"}:
            break
    assert status == "succeeded", client.get(f"/api/v1/extractions/{run_id}").json()
    return doc_id, run_id


def test_a_real_extraction_reaches_the_buckets_endpoint_with_its_rows_in_the_right_sections(client):
    """The whole path, once: upload a native PDF, run the actual pipeline with the seeded rulebook,
    read the actual endpoint.

    Everything above works on a hand-built document model, which proves the resolver and the join but
    not that the stage is WIRED — a segmentation the pipeline never runs, or one the run never
    stores, looks identical to a correct one from inside a unit test. This is also the test that
    caught the real defect: ``note_number`` was read as "this row lives in a note" when it actually
    holds the note a row CITES, so the four-row balance sheet below placed exactly one row.

    The fixture's balance sheet prints no section banners, so the sections here come only from the
    concepts the mapper resolved — the path a condensed filing takes.
    """
    doc_id, run_id = _extract(client, "buckets-e2e.pdf", _shipped_rulebook(client))

    index = client.get(f"/api/v1/documents/{doc_id}/buckets").json()
    assert index["segmented"] is True, "the segment stage did not run, or the run did not store it"
    assert [b["bucket"] for b in index["buckets"]] == list(BUCKET_KEYS)

    served = client.get(f"/api/v1/extractions/{run_id}").json()["result"]
    assert served["rulebook"]["applied"] is True, "the run mapped nothing; placement is untestable"
    assert sum(b["face_rows"] for b in index["buckets"]) == len(served["rows"])
    assert index["unresolved_face_rows"] == 0

    def labels(bucket: str) -> set[str]:
        body = client.get(f"/api/v1/documents/{doc_id}/buckets/{bucket}").json()
        return {r["source_label"] for r in body["rows"]}

    assert "Property, plant and equipment" in labels("non_current_assets")
    assert {"Trade receivables"} <= labels("current_assets")
    assert any("Cash and cash equivalents" in x for x in labels("current_assets"))
    # The balance sheet's own total spans both asset buckets, so it is in Others by design.
    assert "Total assets" in labels("others")


def test_an_extraction_that_pins_no_rulebook_still_uses_the_one_in_force(client):
    """The reason the buckets fill at all on a plain run, and it was not true until now.

    A run naming no rulebook used to map against NOTHING — the worker left ``ontology = None``, so no
    caption resolved to a concept, no concept carried a section, and every row landed unplaced. The
    upload screen sends exactly that request, so in practice every extraction produced a filing with
    nothing recognised in it while the comment beside the pin claimed "a run naming no rulebook is
    read by the shipped default".

    The rulebook in force is resolved when the run is created and PINNED on it, which is what makes
    this a default rather than the silent substitution this codebase removed for templates: the run
    stores the id, so a later reader is told which rulebook produced the figures instead of having to
    reconstruct "whatever was in force at the time".
    """
    doc_id, run_id = _extract(client, "buckets-default-rulebook.pdf")

    served = client.get(f"/api/v1/extractions/{run_id}").json()["result"]
    assert served["rulebook"]["applied"] is True
    assert served["rulebook"]["ontology_key"] == "hkfrs_hk_china"
    assert served["rulebook"]["ontology_version_id"], "the run must name the rulebook it used"

    index = client.get(f"/api/v1/documents/{doc_id}/buckets").json()
    counts = {b["bucket"]: b["face_rows"] for b in index["buckets"]}
    assert counts["current_assets"] == 2 and counts["non_current_assets"] == 1
    assert index["unresolved_face_rows"] == 0


def test_the_worker_reads_the_options_the_run_stores(client):
    """The defect behind the one above, and the more general one: the worker was handed
    ``body.model_dump()`` while the row stored a different dict, so anything settled at run creation
    was recorded on the run and then not used to produce its figures. A run that says which rulebook
    it read the filing against and did not read it is worse than one that says nothing.
    """
    from app.db.base import SessionLocal
    from app.db.models import ExtractionRun

    _doc_id, run_id = _extract(client, "buckets-one-options-dict.pdf")
    with SessionLocal() as session:
        run = session.get(ExtractionRun, run_id)
        stored = run.options.get("ontology_version_id")
        assert stored, "the run did not pin a rulebook"
        # The row's column and its options agree, and the figures were produced by that rulebook.
        assert run.ontology_version_id == stored
        assert run.result["rulebook"]["ontology_version_id"] == stored
        assert run.result["rulebook"]["applied"] is True


def test_a_caller_pinning_a_rulebook_still_gets_the_one_it_asked_for(client):
    """The default fills in an ABSENT pin and never overrides a present one — reproducing an earlier
    spread against a superseded rulebook is a legitimate request, and this is the assertion that
    stops the fallback quietly taking it over."""
    pinned = _shipped_rulebook(client)
    _doc_id, run_id = _extract(client, "buckets-pinned.pdf", pinned)

    served = client.get(f"/api/v1/extractions/{run_id}").json()["result"]
    assert served["rulebook"]["ontology_version_id"] == pinned


# --- the duplication is confined to the bucket store -------------------------------------------

def test_a_shared_note_is_still_stored_once_outside_the_buckets(client, ontology):
    """THE AUDIT, as an assertion. Filing a note in two buckets duplicates it in the bucket store
    and nowhere else, and that is what keeps every other consumer correct.

    ``note_details`` is the flat list the run stores and the one every note consumer reads — the
    export's note sheet, the ``/notes`` index, the ``/notes/{no}`` detail, the §20 reconciliation.
    They all key by note number, so a note filed in two buckets must still appear ONCE here. If the
    duplication ever leaked into this list, an exported workbook would print the borrowings note
    twice and the notes count would overstate the filing, with nothing on either saying why.
    """
    result = _shared_note_result(ontology)
    doc_id = _seed_run(result, "shared-note-isolation.pdf")

    numbers = [n["no"] for n in result["note_details"]]
    assert numbers == ["25"], "the flat note list must not carry the bucket store's duplication"

    # Both buckets serve it…
    for bucket in ("current_liabilities", "non_current_liabilities"):
        served = client.get(f"/api/v1/documents/{doc_id}/buckets/{bucket}").json()
        assert [n["no"] for n in served["notes"]] == ["25"]
    # …while the filing-level index still reports one note.
    assert client.get(f"/api/v1/documents/{doc_id}/notes").json()["count"] == 1


def test_the_export_prints_a_shared_note_once(client, ontology):
    """The consumer that matters most, because the workbook is what leaves the building. The export
    reads ``note_details`` and never the bucket store, so a note two sections share is written as
    one block — asserted here rather than argued, since a double-printed note in a deliverable is
    not something a reader can be expected to spot.
    """
    import io

    from openpyxl import load_workbook

    from app.services.export import build_rows_xlsx  # noqa: F401  (import guard: openpyxl present)
    from app.services.export import build_statement_workbook

    result = _shared_note_result(ontology)
    template = json.loads((_SAMPLES / "hkfrs_hk_china_template.json").read_text())
    data = build_statement_workbook(result["rows"], template, filename="shared-note",
                                    note_details=result["note_details"])
    wb = load_workbook(io.BytesIO(data))
    headings = [c.value for sheet in wb.worksheets for row in sheet.iter_rows()
                for c in row
                if isinstance(c.value, str) and c.value.startswith("Note 25")]

    assert len(headings) == 1, f"the shared note was written {len(headings)} times: {headings}"
