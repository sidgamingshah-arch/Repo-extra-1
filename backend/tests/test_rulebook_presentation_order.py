"""The ``mappings`` array is in the order a statement reads, and the workbook inherits it.

WHY THIS IS A TEST AND NOT A TIDY-UP. A reviewer reads the rulebook as the Concepts sheet of the
ontology workbook, and that sheet is generated straight from the ``mappings`` array — so the array
IS the presentation. It had drifted into insertion order: gross profit sat forty-fourth among the
profit-and-loss concepts, after the profit-attribution lines, and every top-level subtotal was
appended at the end, well below the sections whose figures they total. Nothing was wrong with the
extraction; the document a reviewer has to check line by line simply did not read like a statement,
and "cost of sales appears after gross profit" is indistinguishable, on the page, from a mapping
error.

The order is DERIVED from the template rather than maintained a second time, because the template
already is the authority on presentation — it is what the Workspace grid and the Excel export are
built from. A hand-kept second opinion would be one more thing to drift.

The array order must also stay MEANINGLESS to the engine, and that is the other half of this file:
re-ordering presentation must never move a figure. ``test_residual_framework`` holds the one place
where it did — two residual buckets claiming one section used to resolve by file position.
"""
from __future__ import annotations

import json
import pathlib

import pytest

from app.services.ontology_xlsx import build_ontology_xlsx

_SAMPLES = pathlib.Path(__file__).resolve().parent.parent / "app" / "sample" / "templates"
ONTOLOGY = _SAMPLES / "hkfrs_hk_china_ontology.json"
TEMPLATE = _SAMPLES / "hkfrs_hk_china_template.json"


@pytest.fixture(scope="module")
def ontology() -> dict:
    return json.loads(ONTOLOGY.read_text(encoding="utf-8"))


@pytest.fixture(scope="module")
def template() -> dict:
    return json.loads(TEMPLATE.read_text(encoding="utf-8"))


def _template_order(template: dict) -> list[str]:
    """Every canonical_key the template names, depth-first, in presentation order."""
    order: list[str] = []

    def walk(nodes: list[dict]) -> None:
        for node in nodes:
            if node.get("canonical_key"):
                order.append(node["canonical_key"])
            walk(node.get("children") or [])

    for statement in template.get("statements", []):
        walk(statement.get("sections", []))
    return order


def test_gross_profit_comes_after_the_cost_of_sales_it_is_struck_from(ontology):
    """The reviewer's requirement, in their own example: "gross profit should come [after] cost of
    sales on P&L". Stated on its own because it is the one an analyst checks by eye."""
    keys = [c["canonical_key"] for c in ontology["mappings"]]
    assert (keys.index("pl_income__revenue_from_operations")
            < keys.index("pl_expenses__cost_of_goods_sold")
            < keys.index("pl_expenses__purchases_of_stock_in_trade")
            < keys.index("pl_gross_profit")
            < keys.index("pl_expenses__taxes_and_surcharges")
            < keys.index("pl_expenses__total_operating_cost")
            < keys.index("pl_operating_profit_ebit")
            < keys.index("pl_profit_before_tax")
            < keys.index("pl_profit_for_the_year"))


def test_the_rulebook_lists_its_concepts_in_exactly_the_templates_order(ontology, template):
    """The general rule behind that example, so a new concept cannot be filed in the wrong place
    without failing here.

    It is an EQUALITY, not merely "consistently ordered", and that is worth stating: every one of
    the rulebook's concepts is presented somewhere by the template, so the template's own walk
    determines the whole array and there is no second order to keep. The template names 19 more keys
    than the rulebook has concepts — its section headers, which carry no figure and no rules.
    """
    keys = [c["canonical_key"] for c in ontology["mappings"]]
    concepts = set(keys)
    assert keys == [k for k in _template_order(template) if k in concepts]


def test_every_concept_is_presented_somewhere_by_the_template(ontology, template):
    """What makes the equality above derivable rather than a coincidence — and a check worth having
    on its own: a concept the template never presents has no row in the Workspace grid and no cell
    in the Excel export, so a filing could map to it and the figure would never be shown."""
    placed = set(_template_order(template))
    orphans = [c["canonical_key"] for c in ontology["mappings"] if c["canonical_key"] not in placed]
    assert not orphans, f"the template presents no row for {orphans}"


def test_the_file_counts_its_own_concepts_correctly(ontology):
    """``metadata.concept_count`` is served to the upload screen as the size of the rulebook a run
    will use. Retiring two concepts left it declaring 185 over an array of 183 — a number a reader
    has no way to check and every reason to trust."""
    assert ontology["metadata"]["concept_count"] == len(ontology["mappings"])


def test_the_workbook_a_reviewer_reads_is_in_the_same_order(ontology):
    """The point of all of the above: the Concepts sheet is the artefact the ordering is FOR. If the
    builder ever sorted or grouped rows on its own, the array could be in statement order and the
    sheet still would not be."""
    from openpyxl import load_workbook
    import io

    wb = load_workbook(io.BytesIO(build_ontology_xlsx(ontology)))
    ws = wb["Concepts"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = header.index("Canonical key") + 1
    sheet_keys = [ws.cell(row=r, column=col).value for r in range(2, ws.max_row + 1)]
    assert sheet_keys == [c["canonical_key"] for c in ontology["mappings"]]
