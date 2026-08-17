"""A calculated total is a real, selectable node on the Template & Ontology screen — and a rulebook
that cannot be validated is refused.

Defects of one family are pinned here: a surface that answers confidently about the wrong thing, and
a gate that degrades to silence instead of failing.

THE TREE WALK. ``get_template_detail`` used to treat every entry in a statement's ``sections[]`` as
a heading and look for line items only among that entry's ``children``. The shipped template
deliberately declares its calculated lines as CHILDLESS top-level sections — Gross Profit sits
between cost of sales and operating expenses because that is where the statement prints it — so each
of them came out as an inert heading row with no ``node_config`` entry, and the screen resolved the
click to whichever concept its fallback chain reached. An analyst could then edit Property, Plant and
Equipment's aliases under a heading reading "Gross profit". The walk now branches on ``role``, the
way ``services.export._emit_nodes`` does.

THE ONTOLOGY GATE. ``_publish_new_version`` validated an edit against the target template only ``if
tpl_row is not None``, so an ontology whose ``target_template_key`` matched no stored template
published unvalidated and became the rulebook in force while mapping onto keys no template declares.

WHAT THE RULEBOOK DOES NOT MAP. Serving the calculated totals as selectable lines widened the class
of node the editor opens by one key the rulebook in force has no concept for
(``bs_liabilities__total_liabilities``, beside the older ``bs_equity__equity_attributable_to_owners``).
The editor opened fully enabled on it and Save came back 404, and the key was offered in the
confusable-with and netting pickers, which answer 422. ``node_config`` now carries ``mapped``, and
these tests hold that flag to the refusals themselves rather than to a transcribed list of keys.

A SPACER IS NOT A LINE. ``LineRole.SPACER`` is a presentational gap. The detail walk and
``export._emit_nodes`` both tested ``role == "header"`` exactly, so a spacer fell through to the
figure branch of each: a selectable concept on screen, and a row in the workbook with a label, a
value column per period and a canonical key an extracted figure could attach to. No shipped template
declares one, so it is pinned with a template published here that does.

Order is asserted against the shipped file rather than a transcribed list: a second spelling of the
template's own order is a second thing that has to keep agreeing with it.
"""
from __future__ import annotations

import json
from pathlib import Path

import pytest

API = "/api/v1"
SEEDED_TEMPLATE = "hkfrs_hk_china_v1"

_DIR = Path(__file__).resolve().parent.parent / "app" / "sample" / "templates"
TEMPLATE = json.loads((_DIR / "hkfrs_hk_china_template.json").read_text())


def _statement(stype: str) -> dict:
    return next(s for s in TEMPLATE["statements"] if s["type"] == stype)


def _row_id(node: dict) -> str:
    """The tree id the endpoint gives this node: headings are addressed by section, lines by key."""
    return f"sec:{node['node_id']}" if node.get("role") == "header" else node["canonical_key"]


def _statement_level_lines() -> list[str]:
    """Every line the template prints at statement level — the childless non-heading sections.

    Mostly the calculated totals (Gross profit, Net assets, Closing cash), plus the odd extracted
    line the statement prints outside any section. Derived from the file so a line added to a future
    template is covered without editing a list here: the whole CLASS of node has to stay selectable,
    not just the seventeen shipped today.
    """
    return [sec["canonical_key"]
            for stmt in TEMPLATE["statements"]
            for sec in stmt.get("sections", [])
            if sec.get("role") != "header" and not sec.get("children")]


def _template_line_items() -> int:
    """The nodes of the shipped template that CARRY A FIGURE — the count ``POST /templates`` reports.

    Spelled independently of the endpoint (which asks ``review_lines.is_statement_line``) so this
    stays a check and not a restatement. A heading and a spacer are captions; everything else is a
    line. The shipped template declares no spacer, so the number is the same either way.
    """
    from app.schemas.loader import load_template

    return len([n for n in load_template(TEMPLATE).all_nodes()
                if n.role.value not in ("header", "spacer")])


def _detail(client, locale: str = "en") -> dict:
    tpl = next(r for r in client.get(f"{API}/templates").json()
               if r["template_key"] == SEEDED_TEMPLATE)
    r = client.get(f"{API}/templates/{tpl['id']}/detail?locale={locale}")
    assert r.status_code == 200, r.text
    return r.json()


def _statement_rows(tree: list[dict], stype: str) -> list[dict]:
    """The rows between this statement's own heading and the next one."""
    start = next(i for i, n in enumerate(tree) if n["id"] == f"stmt:{stype}")
    rest = tree[start + 1:]
    end = next((i for i, n in enumerate(rest) if str(n["id"]).startswith("stmt:")), len(rest))
    return rest[:end]


# --- a calculated total is a selectable line -------------------------------------------------

def test_gross_profit_is_a_selectable_line_not_a_heading(client):
    """The row the analyst clicks and the concept the editor opens must be the same thing."""
    detail = _detail(client)
    row = next((n for n in detail["tree"] if n["id"] == "pl_gross_profit"), None)
    assert row is not None, (
        "Gross profit reached the tree only as 'sec:pl_gross_profit' — a heading id no "
        "node_config key can match, which is what sent the click to another concept")
    assert not row.get("head"), "a calculated total carries a figure; it is a line, not a heading"
    cfg = detail["node_config"].get("pl_gross_profit")
    assert cfg is not None, "a selectable line must carry the rules the editor edits"
    assert cfg["canonical_key"] == "pl_gross_profit"
    assert cfg["label"] == "Gross profit"
    # A total printed at statement level has no section above it, so it breadcrumbs to the
    # statement rather than borrowing the heading of whichever section happens to precede it.
    assert cfg["breadcrumb"] == "Profit And Loss"
    assert cfg["aggregation"] == "Sum of children"


@pytest.mark.parametrize("key", _statement_level_lines())
def test_every_statement_level_line_is_editable(client, key):
    """Not just Gross Profit: the whole class, in all three statements."""
    detail = _detail(client)
    assert key in detail["node_config"], f"{key} is a line the screen cannot answer about"
    row = next((n for n in detail["tree"] if n["id"] == key), None)
    assert row is not None and not row.get("head")


def test_a_totals_label_localizes_like_any_other_line(client):
    """A total reaches `_loc` on the same path as a line — it must not fall back to English."""
    detail = _detail(client, locale="zh")
    zh = _statement("profit_and_loss")["sections"]
    expected = next(s for s in zh if s["canonical_key"] == "pl_gross_profit")["label_i18n"]["zh"]
    assert detail["node_config"]["pl_gross_profit"]["label"] == expected


# --- position: the user's item 4, held from the Template-screen side --------------------------

def test_the_pl_tree_is_in_the_templates_own_order(client):
    """Serving a total as a real node must not move it. The template's order IS the statement's
    order, and a calculated line printed mid-statement that renders at the end is a spread no
    analyst can read against the filing."""
    rows = _statement_rows(_detail(client)["tree"], "profit_and_loss")
    top = [n["id"] for n in rows if n["lvl"] == 1]
    assert top == [_row_id(s) for s in _statement("profit_and_loss")["sections"]]


def test_gross_profit_sits_between_cost_of_sales_and_operating_expenses(client):
    """The concrete placement, named, because "in file order" is only reassuring if you know what
    the file says: Gross profit follows the cost-of-sales block and precedes operating expenses.

    This is the reviewer's positioning requirement stated as an assertion — "gross profit should
    come [after] cost of sales on P&L". It reads the cost lines directly now that the intermediate
    total-cost-of-sales subtotal is retired, so there is nothing between the last cost line and the
    margin it produces.
    """
    ids = [n["id"] for n in _statement_rows(_detail(client)["tree"], "profit_and_loss")]
    assert (ids.index("sec:pl_s2a_cost_of_sales")
            < ids.index("pl_expenses__cost_of_goods_sold")
            < ids.index("pl_expenses__purchases_of_stock_in_trade")
            < ids.index("pl_gross_profit")
            < ids.index("sec:pl_s2_expenses")
            < ids.index("pl_expenses__taxes_and_surcharges")
            < ids.index("pl_expenses__total_operating_cost")
            < ids.index("pl_operating_profit_ebit"))


# --- genuine headings keep behaving exactly as they did --------------------------------------

def test_a_genuine_heading_is_still_a_heading_over_its_children(client):
    detail = _detail(client)
    tree = detail["tree"]
    at = next(i for i, n in enumerate(tree) if n["id"] == "sec:pl_s1_income")
    assert tree[at]["head"] is True and tree[at]["lvl"] == 1
    assert tree[at]["id"] not in detail["node_config"], (
        "a heading carries no figure and no ontology rules, so it must not be selectable")

    section = next(s for s in _statement("profit_and_loss")["sections"]
                   if s["node_id"] == "pl_s1_income")
    children = tree[at + 1:at + 1 + len(section["children"])]
    assert [c["id"] for c in children] == [c["canonical_key"] for c in section["children"]]
    assert all(c["lvl"] == 2 and not c.get("head") for c in children)
    assert all(c["id"] in detail["node_config"] for c in children)


def test_the_line_item_count_is_the_templates_real_one(client):
    """One quantity, one spelling: the screen's count and the count `POST /templates` reports on
    upload are the same number. They disagreed while the seventeen statement-level lines were
    counted as headings here (170) and as line items there (187)."""
    detail = _detail(client)
    expected = _template_line_items()
    assert detail["template"]["line_items"] == expected
    assert len(detail["node_config"]) == expected
    lines = [n for n in detail["tree"]
             if not n.get("head") and not str(n["id"]).startswith("stmt:")]
    assert len(lines) == expected


# --- what the rulebook in force does NOT map -------------------------------------------------

def _rulebook_concepts(client, detail: dict) -> set[str]:
    """The canonical keys the rulebook IN FORCE declares, read from that rulebook.

    Read back through the API rather than off the shipped file: ``mapped`` is a claim about whichever
    rulebook ``ontology_select.select_for_template`` chose for THIS database, so a test transcribing
    the file would keep agreeing with itself while the screen described another rulebook's concepts.
    """
    ont = detail["ontology"]
    assert ont, "the reference template should have a rulebook in force"
    r = client.get(f"{API}/ontologies/{ont['id']}")
    assert r.status_code == 200, r.text
    return {m.get("canonical_key") for m in (r.json()["definition"].get("mappings") or [])}


def test_every_line_says_whether_the_rulebook_in_force_maps_it(client):
    """The editor opens off ``node_config``, so ``node_config`` is where the answer has to be.

    Without it the screen cannot tell an editable concept from one whose every write the server
    refuses, and it offered the same fully-enabled editor for both.
    """
    detail = _detail(client)
    silent = [k for k, c in detail["node_config"].items() if "mapped" not in c]
    assert not silent, (
        f"{len(silent)} lines say nothing about whether the rulebook maps them (e.g. "
        f"{sorted(silent)[:3]}) — the screen cannot know which of its controls would be refused")
    declared = _rulebook_concepts(client, detail)
    wrong = sorted(k for k, c in detail["node_config"].items() if c["mapped"] != (k in declared))
    assert not wrong, f"`mapped` disagrees with the rulebook in force for {wrong}"
    assert not all(c["mapped"] for c in detail["node_config"].values()), (
        "the shipped template declares lines this rulebook has no concept for; a `mapped` that is "
        "True everywhere is the flag not being computed at all")


def test_the_calculated_total_this_round_exposed_is_the_unmapped_one(client):
    """The concrete case, named, because "some line is unmapped" is only useful if you know which.

    The rulebook maps sixteen of the seventeen lines the template prints at statement level;
    ``bs_liabilities__total_liabilities`` is the one it does not, and serving those totals as
    selectable lines is what first put it in front of an analyst. If a later rulebook gains the
    concept, this test is the record of what changed — the derived test above keeps the class covered.
    """
    detail = _detail(client)
    key = "bs_liabilities__total_liabilities"
    assert key in detail["node_config"], "the class of node this is about must stay selectable"
    assert key not in _rulebook_concepts(client, detail)
    assert detail["node_config"][key]["mapped"] is False
    assert detail["node_config"]["pl_gross_profit"]["mapped"] is True, (
        "the other sixteen ARE mapped: an editor read-only for a concept the rulebook has is the "
        "same defect facing the other way")


def test_the_controls_the_screen_withholds_are_the_ones_the_server_refuses(client):
    """``mapped`` is not an opinion about tidiness — it is the answer to "would this be refused?".

    All three writes the screen used to offer for an unmapped key are asked here, so the flag and the
    server's answer cannot drift apart: the concept editor's Save (404, not in this ontology), the
    confusable-with picker and the netting pickers (422, unknown concept). None of the three changes
    stored state when it is refused.
    """
    detail = _detail(client)
    ont_id = detail["ontology"]["id"]
    unmapped = sorted(k for k, c in detail["node_config"].items() if not c["mapped"])
    assert unmapped, "no unmapped line in the shipped template would make this test prove nothing"
    editable = next(k for k, c in detail["node_config"].items() if c["mapped"])
    for key in unmapped:
        r = client.patch(f"{API}/ontologies/{ont_id}/mappings",
                         json={"canonical_key": key, "aliases": ["Anything at all"]})
        assert r.status_code == 404, f"an alias edit on {key} was accepted: {r.text}"
        r = client.patch(f"{API}/ontologies/{ont_id}/mappings",
                         json={"canonical_key": editable, "confusable_with": [key]})
        assert r.status_code == 422, f"{key} was accepted as a confusable_with target: {r.text}"
        r = client.patch(f"{API}/ontologies/{ont_id}/netting-rules",
                         json={"id": "unmapped_probe", "target_key": key, "subtract_keys": []})
        assert r.status_code == 422, f"{key} was accepted as a netting target: {r.text}"


_UNRESOLVABLE_TPL_KEY = "unresolvable_probe_tpl"
_UNRESOLVABLE_ONT_KEY = "unresolvable_probe_ont"
_UNRESOLVABLE_TEMPLATE = {
    "template_key": _UNRESOLVABLE_TPL_KEY,
    "name": "Unresolvable rulebook probe",
    "statements": [{
        "type": "balance_sheet",
        "sections": [{"node_id": "cash", "canonical_key": "probe_cash", "label": "Cash",
                      "role": "line"}],
    }],
}
_UNRESOLVABLE_ONTOLOGY = {
    "ontology_key": _UNRESOLVABLE_ONT_KEY,
    "target_template_key": _UNRESOLVABLE_TPL_KEY,
    "schema_version": 2,
    # A section layer, so the fold runs at all…
    "section_defaults": {"bs_s1": {"statement": "balance_sheet", "section_scope": ["bs_s1"]}},
    # …and an `inherits` naming an entry that is not in it. The definition VALIDATES, and
    # `loader.resolve_inherits` raises on it — so `get_template_detail`'s `load_ontology(resolve=True)`
    # fails and its `except` serves the screen no rules at all. The concept is declared regardless,
    # and the declaration is what every editing endpoint looks for.
    "mappings": [{"canonical_key": "probe_cash", "label": "Cash", "aliases": ["Cash"],
                  "inherits": "no_such_section"}],
}


@pytest.fixture
def unresolvable_rulebook(client):
    """A stored rulebook the detail cannot LOAD, targeting a template of one line.

    Inserted straight into the database because ``POST /ontologies`` refuses an unresolvable
    ``inherits`` (422) — a row like this predates that gate, which is the state the ``except`` in
    ``get_template_detail`` exists for.
    """
    from app.db.models import OntologyVersion, TemplateVersion

    from app.db.base import SessionLocal

    r = client.post(f"{API}/templates", json={"definition": _UNRESOLVABLE_TEMPLATE})
    assert r.status_code == 201, r.text
    with SessionLocal() as session:
        session.add(OntologyVersion(ontology_key=_UNRESOLVABLE_ONT_KEY,
                                    target_template_key=_UNRESOLVABLE_TPL_KEY,
                                    version=1, definition=_UNRESOLVABLE_ONTOLOGY))
        session.commit()
    try:
        yield r.json()
    finally:
        _drop(OntologyVersion, OntologyVersion.ontology_key, _UNRESOLVABLE_ONT_KEY)
        _drop(TemplateVersion, TemplateVersion.template_key, _UNRESOLVABLE_TPL_KEY)


def test_a_rulebook_that_cannot_be_loaded_still_says_which_concepts_it_declares(
        client, unresolvable_rulebook):
    """``mapped`` is read off the STORED definition, because that is what the writes are checked on.

    Taken from the loaded rulebook, one unresolvable ``inherits`` — and every concept of the shipped
    rulebook uses ``inherits`` — turned the flag inside out: the screen locked every line and said
    the rulebook declared no concept for any of them, while the edit endpoint went on accepting the
    key it had just disabled.
    """
    r = client.get(f"{API}/templates/{unresolvable_rulebook['id']}/detail")
    assert r.status_code == 200, r.text
    detail = r.json()
    assert detail["node_config"]["probe_cash"]["mapped"] is True, (
        "the rulebook declares this concept; only its RULES could not be loaded")
    r = client.patch(f"{API}/ontologies/{detail['ontology']['id']}/mappings",
                     json={"canonical_key": "probe_cash", "aliases": ["Cash at bank"]})
    assert r.status_code == 200, (
        f"the write `mapped` is a claim about: {r.text}")


# --- an unvalidatable rulebook is refused ----------------------------------------------------

_PROBE_KEY = "orphan_probe_tpl"
_PROBE_ONT_KEY = "orphan_probe_ont"
_PROBE_TEMPLATE = {
    "template_key": _PROBE_KEY,
    "name": "Orphan probe",
    "statements": [{
        "type": "balance_sheet",
        "sections": [{"node_id": "cash", "canonical_key": "cash", "label": "Cash",
                      "role": "line"}],
    }],
}
_PROBE_ONTOLOGY = {
    "ontology_key": _PROBE_ONT_KEY,
    "target_template_key": _PROBE_KEY,
    "mappings": [{"canonical_key": "cash", "label": "Cash", "aliases": ["Cash"]}],
}


def _drop(model, key_column, key: str) -> None:
    from sqlalchemy import delete

    from app.db.base import SessionLocal

    with SessionLocal() as session:
        session.execute(delete(model).where(key_column == key))
        session.commit()


@pytest.fixture
def probe_pair(client):
    """A throwaway template and a rulebook published against it, removed again afterwards.

    The ``client`` fixture is session-scoped and so is its database: a probe left behind would sit
    in every later test's `/templates` and `/ontologies` listing.
    """
    from app.db.models import OntologyVersion, TemplateVersion

    r = client.post(f"{API}/templates", json={"definition": _PROBE_TEMPLATE})
    assert r.status_code == 201, r.text
    tpl = r.json()
    r = client.post(f"{API}/ontologies", json={"definition": _PROBE_ONTOLOGY})
    assert r.status_code == 201, r.text
    try:
        yield tpl, r.json()
    finally:
        _drop(OntologyVersion, OntologyVersion.ontology_key, _PROBE_ONT_KEY)
        _drop(TemplateVersion, TemplateVersion.template_key, _PROBE_KEY)


@pytest.fixture
def orphaned_ontology(probe_pair):
    """The same rulebook, with its target template GONE — a renamed key, or a deleted template.

    Orphaned by dropping the template row rather than by publishing against a key that never
    existed, because the create path has always refused that: the hole was on the edit path.
    """
    from app.db.models import TemplateVersion

    _drop(TemplateVersion, TemplateVersion.template_key, _PROBE_KEY)
    return probe_pair[1]["id"]


def test_an_edit_to_a_rulebook_with_no_target_template_is_refused(client, orphaned_ontology):
    """`if tpl_row is not None` skipped validation entirely when the target template was missing,
    so this edit published — and a rulebook nothing had checked became the one in force."""
    r = client.patch(f"{API}/ontologies/{orphaned_ontology}/mappings",
                     json={"canonical_key": "cash", "aliases": ["Cash at bank"]})
    assert r.status_code == 422, r.text
    assert _PROBE_KEY in json.dumps(r.json()), (
        "the refusal has to name the template it could not find, or the author cannot fix it")


def test_a_netting_edit_takes_the_same_gate(client, orphaned_ontology):
    """Both inline-edit endpoints publish through `_publish_new_version`, so neither can be the
    one path where validation is skipped."""
    r = client.patch(f"{API}/ontologies/{orphaned_ontology}/netting-rules",
                     json={"id": "probe", "target_key": "cash", "subtract_keys": []})
    assert r.status_code == 422, r.text
    assert _PROBE_KEY in json.dumps(r.json())


def test_a_new_rulebook_naming_no_stored_template_is_refused(client):
    body = {"definition": {"ontology_key": "no_such_target_ont",
                           "target_template_key": "template_that_was_never_published",
                           "mappings": []}}
    r = client.post(f"{API}/ontologies", json=body)
    assert r.status_code == 422, r.text
    assert "template_that_was_never_published" in json.dumps(r.json())


# --- and the record says WHICH template version it was checked against -----------------------

def test_a_published_rulebook_states_the_template_version_it_was_checked_against(probe_pair):
    """The check runs against whichever template version is newest at publish time, which is not
    necessarily the version a run pins. Saying so on the response is what lets a reader tell,
    rather than assume, what a rulebook was held to."""
    tpl, ont = probe_pair
    assert ont.get("validated_against_template") == {
        "id": tpl["id"], "template_key": tpl["template_key"], "version": tpl["version"]}


def test_an_inline_edit_states_it_the_same_way(client, probe_pair):
    """Create and edit report it with one spelling, so neither path is the one you have to guess
    about."""
    tpl, ont = probe_pair
    r = client.patch(f"{API}/ontologies/{ont['id']}/mappings",
                     json={"canonical_key": "cash", "aliases": ["Cash at bank"]})
    assert r.status_code == 200, r.text
    against = r.json().get("validated_against_template")
    assert against == {"id": tpl["id"], "template_key": _PROBE_KEY, "version": tpl["version"]}
    assert against == ont["validated_against_template"]


def test_an_edit_to_the_shipped_rulebook_names_the_shipped_template(client):
    """Not only for a probe: the rulebook actually in force reports it too."""
    ont = _detail(client)["ontology"]
    assert ont, "the reference template should have a rulebook in force"
    r = client.patch(f"{API}/ontologies/{ont['id']}/mappings",
                     json={"canonical_key": "pl_gross_profit", "aliases": ["Gross profit"]})
    assert r.status_code == 200, r.text
    against = r.json().get("validated_against_template")
    assert against, "an edit that says nothing about what validated it cannot be audited"
    assert against["template_key"] == SEEDED_TEMPLATE
    assert against["version"] == max(r["version"] for r in client.get(f"{API}/templates").json()
                                     if r["template_key"] == SEEDED_TEMPLATE)
    assert against["id"]


# --- a spacer is a presentational gap, not a line ---------------------------------------------

_SPACER_KEY = "spacer_probe_tpl"
_GAP_LABEL = "— gap —"
_SPACER_TEMPLATE = {
    "template_key": _SPACER_KEY,
    "name": "Spacer probe",
    "statements": [{
        "type": "balance_sheet",
        "sections": [{
            "node_id": "ca", "canonical_key": "ca_head", "label": "Current assets",
            "role": "header",
            "children": [
                {"node_id": "cash", "canonical_key": "cash", "label": "Cash", "role": "line"},
                # Carrying a canonical_key because the schema requires one, which is exactly what
                # made the old treatment costly: the gap was a keyed row an extracted figure could
                # land on.
                {"node_id": "gap", "canonical_key": "spacer_gap", "label": _GAP_LABEL,
                 "role": "spacer"},
                {"node_id": "inv", "canonical_key": "inv", "label": "Inventories", "role": "line"},
            ],
        }],
    }],
}


@pytest.fixture
def spacer_template(client):
    """A published template that declares a spacer — no shipped one does — removed afterwards.

    Removed for the reason ``probe_pair`` is: the ``client`` fixture's database is session-scoped, so
    a probe left behind sits in every later test's `/templates` listing.
    """
    from app.db.models import TemplateVersion

    r = client.post(f"{API}/templates", json={"definition": _SPACER_TEMPLATE})
    assert r.status_code == 201, r.text
    try:
        yield r.json()
    finally:
        _drop(TemplateVersion, TemplateVersion.template_key, _SPACER_KEY)


def test_a_spacer_is_no_line_item_on_publish_or_on_the_screen(client, spacer_template):
    """A gap is not a concept: it has no aliases, no sign convention and no figure to extract.

    Counted as a line item it also breaks the one quantity `_publish` and this screen both report.
    """
    assert spacer_template["line_items"] == 2, (
        "the template declares two lines and one gap; a spacer carries no figure")
    r = client.get(f"{API}/templates/{spacer_template['id']}/detail")
    assert r.status_code == 200, r.text
    detail = r.json()
    assert detail["template"]["line_items"] == spacer_template["line_items"]
    assert sorted(detail["node_config"]) == ["cash", "inv"], (
        "a presentational gap must not be offered as a concept to alias and sign")
    ids = [n["id"] for n in detail["tree"]]
    assert ids == ["stmt:balance_sheet", "sec:ca", "cash", "sec:gap", "inv"], (
        "the gap keeps its place in the statement, addressed like a caption: `sec:gap` matches no "
        "node_config key, so clicking it cannot resolve to another concept's rules")
    gap = next(n for n in detail["tree"] if n["id"] == "sec:gap")
    assert gap["head"] is True and gap["lvl"] == 2, "a gap is not selectable"


def test_the_export_writes_a_spacer_as_a_blank_row_not_a_figure_row():
    """The workbook's equivalent of a presentational gap is an empty row, not a line item.

    Emitted through the figure branch, the gap arrived as a row with a label, a note cell and a value
    column per period — and, because it carries a canonical_key, a row `by_key` could attach an
    extracted figure to.
    """
    pytest.importorskip("openpyxl")
    import io

    import openpyxl

    from app.services.export import build_statement_workbook

    def _row(key: str, label: str, value: str) -> dict:
        return {"canonical_key": key, "source_label": label,
                "values": [{"basis": "consolidated", "period_label": "current", "value": value}]}

    rows = [_row("cash", "Cash", "10"), _row("spacer_gap", _GAP_LABEL, "4242"),
            _row("inv", "Inventories", "20")]
    wb = openpyxl.load_workbook(io.BytesIO(
        build_statement_workbook(rows, _SPACER_TEMPLATE, filename="f.pdf")))
    ws = wb["Balance Sheet"]
    col_a = [ws.cell(r, 1).value for r in range(1, ws.max_row + 1)]
    assert _GAP_LABEL not in col_a, f"the gap was written out as a line: {col_a}"
    at = col_a.index("Cash")
    assert col_a[at + 1] is None, f"the gap's row is the gap, so it stays empty: {col_a}"
    assert col_a[at + 2] == "Inventories", f"the lines around it keep their order: {col_a}"
    gap_row = at + 2                                   # col_a[i] is row i + 1
    assert all(ws.cell(gap_row, c).value in (None, "")
               for c in range(1, ws.max_column + 1)), "the gap row carries no cell of its own"
    written = {ws.cell(r, c).value
               for r in range(1, ws.max_row + 1) for c in range(1, ws.max_column + 1)}
    assert 4242 not in written and "4242" not in written, (
        "a figure keyed to the gap reached the sheet — a presentational row with a number in it")
