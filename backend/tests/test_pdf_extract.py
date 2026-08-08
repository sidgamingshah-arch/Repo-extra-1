"""PDF extraction: native text layer (PyMuPDF) and the scanned/OCR path — both producing
line items with page + normalized-bbox provenance."""
from __future__ import annotations

import time

import pytest

pytest.importorskip("fitz")

from tests.fixtures.generate import make_native_pdf


def _await_run(client, doc_id: str) -> dict:
    """Extraction is a background job now; poll the run until it finishes and return its
    result (raises if it fails or never completes)."""
    for _ in range(100):
        r = client.get(f"/api/v1/documents/{doc_id}/run")
        if r.status_code == 200 and r.json().get("status") == "succeeded":
            return r.json()["result"]
        if r.status_code == 200 and r.json().get("status") == "failed":
            raise AssertionError("extraction failed")
        time.sleep(0.05)
    raise AssertionError("extraction did not finish")


def test_native_pdf_extracts_values_with_bbox_provenance():
    from app.core.models.enums import Basis
    from app.services.documents import run_extraction

    doc, ctx = run_extraction(make_native_pdf(), filename="bs.pdf")
    assert doc.fmt.value == "pdf"
    labels = {li.source_label for li in doc.line_items}
    assert any("Trade receivables" in lbl for lbl in labels)

    tr = next(li for li in doc.line_items if "Trade receivables" in li.source_label)
    val = tr.get_value(Basis.CONSOLIDATED, period_label="current")
    assert val is not None and int(val.value) == 3410      # "3,410" parsed
    assert val.provenance is not None and val.provenance.page_index == 0
    assert val.provenance.bbox is not None                  # real click-to-source geometry
    assert val.provenance.source_kind == "native"
    assert tr.note_number == "15"                           # "Note 15" captured, not a value


def test_extraction_api_returns_rows_with_provenance(client):
    """Upload → run extraction → the run result carries view-ready rows + provenance."""
    doc_id = client.post(
        "/api/v1/documents", files={"file": ("bs.pdf", make_native_pdf(), "application/pdf")}
    ).json()["id"]
    r = client.post(f"/api/v1/documents/{doc_id}/extractions", json={})
    assert r.status_code == 202, r.text
    assert r.json()["status"] == "running"                  # async: kicked off, poll for result
    rows = _await_run(client, doc_id)["rows"]
    assert rows and any(row["values"] for row in rows)
    tr = next(row for row in rows if "Trade receivables" in row["source_label"])
    prov = tr["values"][0]["provenance"]
    assert prov["source_kind"] == "native" and prov["page_index"] == 0 and prov["bbox"]
    assert tr["note"] == "15"


def test_document_integrity_endpoint_shapes_real_report(client):
    """A real uploaded file surfaces its own pre-flight integrity in the screen's shape."""
    doc_id = client.post(
        "/api/v1/documents", files={"file": ("bs.pdf", make_native_pdf(), "application/pdf")}
    ).json()["id"]
    r = client.get(f"/api/v1/documents/{doc_id}/integrity")
    assert r.status_code == 200, r.text
    body = r.json()
    assert 0 <= body["score"] <= 100
    assert body["grade"] and body["summary"]
    assert len(body["stats"]) == 4 and any(s["label"] == "Pages" for s in body["stats"])
    assert body["issues"]  # at least the "no issues" row for a clean native PDF
    assert all({"title", "detail", "pages", "note", "status", "severity"} <= set(i) for i in body["issues"])


def test_real_review_and_export_from_extraction(client):
    """A real uploaded file → extract → review queue + Excel/JSON export, all from the
    document's own run (no demo data)."""
    doc_id = client.post(
        "/api/v1/documents", files={"file": ("bs.pdf", make_native_pdf(), "application/pdf")}
    ).json()["id"]
    client.post(f"/api/v1/documents/{doc_id}/extractions", json={})
    assert _await_run(client, doc_id)["rows"]

    # Latest run is fetchable for the export preview/counts.
    run = client.get(f"/api/v1/documents/{doc_id}/run")
    assert run.status_code == 200 and run.json()["result"]["rows"]

    # Review queue derives from the real rows.
    rev = client.get(f"/api/v1/documents/{doc_id}/review").json()
    assert set(rev) == {"checks", "tabs", "summary"}
    assert rev["tabs"][0]["label"] == "All"
    assert rev["summary"]["open"] + rev["summary"]["passed"] >= 1

    # JSON export contains the real line items with source locations.
    j = client.get(f"/api/v1/documents/{doc_id}/export", params={"fmt": "json"})
    assert j.status_code == 200 and j.headers["content-type"].startswith("application/json")
    body = j.json()
    assert body["source_document"] == "bs.pdf" and body["line_item_count"] >= 1
    assert any("Trade receivables" in (li["source_label"] or "") for li in body["line_items"])

    # Excel export is a real workbook.
    x = client.get(f"/api/v1/documents/{doc_id}/export", params={"fmt": "excel"})
    assert x.status_code == 200
    assert x.content[:2] == b"PK" and len(x.content) > 200          # xlsx = zip


def test_real_pages_and_statement_from_document(client):
    """Page Scope + Workspace read the real document: page classification (pre-extraction)
    and the extracted statement grouped by the template."""
    doc_id = client.post(
        "/api/v1/documents", files={"file": ("bs.pdf", make_native_pdf(), "application/pdf")}
    ).json()["id"]

    # Pages are available before extraction (confirm-scope step).
    pg = client.get(f"/api/v1/documents/{doc_id}/pages").json()
    assert pg["total"] >= 1 and pg["total"] == len(pg["pages"])
    assert pg["focused"] + pg["skipped"] == pg["total"]
    assert all({"no", "cls", "conf", "included", "scan"} <= set(p) for p in pg["pages"])

    # Attach the seeded HK ontology/template so mapping produces canonical keys (as the UI does).
    onts = client.get("/api/v1/ontologies").json()
    ont = next((o for o in onts if o["ontology_key"] == "hkfrs_hk_china_v1"), onts[0])
    tpls = client.get("/api/v1/templates").json()
    tpl = next((tt for tt in tpls if tt["template_key"] == ont["target_template_key"]), tpls[0])
    client.post(f"/api/v1/documents/{doc_id}/extractions",
                json={"ontology_version_id": ont["id"], "template_version_id": tpl["id"]})
    st = client.get(f"/api/v1/documents/{doc_id}/statement", params={"statement": "balance_sheet"}).json()
    assert st["statement"] == "balance_sheet" and st["viewer"]["company"] == "bs.pdf"
    items = [r for r in st["rows"] if r["kind"] == "item"]
    assert items and any("Trade receivables" in (r.get("source_label") or "") for r in items)
    assert all(r["v1"] is None or isinstance(r["v1"], (int, float)) for r in st["rows"])

    # Standalone basis (not extracted) → valid but empty grid, never demo data.
    sa = client.get(f"/api/v1/documents/{doc_id}/statement",
                    params={"statement": "balance_sheet", "basis": "standalone"}).json()
    assert sa["rows"] == []


def test_dual_basis_extracted_in_one_pass():
    """A Consolidated | Standalone two-level header → both bases extracted for each line."""
    from app.core.models.enums import Basis
    from app.services.documents import run_extraction
    from tests.fixtures.generate import make_dual_basis_pdf

    doc, _ = run_extraction(make_dual_basis_pdf(), filename="dual.pdf")
    tr = next(li for li in doc.line_items if "Trade receivables" in li.source_label)
    cc = tr.get_value(Basis.CONSOLIDATED, period_label="current")
    cp = tr.get_value(Basis.CONSOLIDATED, period_label="prior")
    sc = tr.get_value(Basis.STANDALONE, period_label="current")
    sp = tr.get_value(Basis.STANDALONE, period_label="prior")
    assert cc and int(cc.value) == 3410 and cp and int(cp.value) == 2900
    assert sc and int(sc.value) == 3100 and sp and int(sp.value) == 2700


def test_dual_basis_statement_selects_by_basis(client):
    from tests.fixtures.generate import make_dual_basis_pdf

    doc_id = client.post(
        "/api/v1/documents", files={"file": ("dual.pdf", make_dual_basis_pdf(), "application/pdf")}
    ).json()["id"]
    onts = client.get("/api/v1/ontologies").json()
    ont = next((o for o in onts if o["ontology_key"] == "hkfrs_hk_china_v1"), onts[0])
    tpls = client.get("/api/v1/templates").json()
    tpl = next((tt for tt in tpls if tt["template_key"] == ont["target_template_key"]), tpls[0])
    client.post(f"/api/v1/documents/{doc_id}/extractions",
                json={"ontology_version_id": ont["id"], "template_version_id": tpl["id"]})
    _await_run(client, doc_id)

    def tr_value(basis):
        st = client.get(f"/api/v1/documents/{doc_id}/statement",
                        params={"statement": "balance_sheet", "basis": basis}).json()
        row = next(x for x in st["rows"]
                   if x["kind"] == "item" and "Trade receivables" in (x.get("source_label") or ""))
        return row["v1"]

    assert tr_value("consolidated") == 3410
    assert tr_value("standalone") == 3100        # different basis → different value, one pass


def _extract_with_ontology(client, filename="bs.pdf"):
    doc_id = client.post(
        "/api/v1/documents", files={"file": (filename, make_native_pdf(), "application/pdf")}
    ).json()["id"]
    onts = client.get("/api/v1/ontologies").json()
    ont = next((o for o in onts if o["ontology_key"] == "hkfrs_hk_china_v1"), onts[0])
    tpls = client.get("/api/v1/templates").json()
    tpl = next((tt for tt in tpls if tt["template_key"] == ont["target_template_key"]), tpls[0])
    client.post(f"/api/v1/documents/{doc_id}/extractions",
                json={"ontology_version_id": ont["id"], "template_version_id": tpl["id"]})
    _await_run(client, doc_id)
    return doc_id


def test_real_line_item_edit_persists_to_statement(client):
    doc_id = _extract_with_ontology(client)
    rows = client.get(f"/api/v1/documents/{doc_id}/run").json()["result"]["rows"]
    key = next(r["canonical_key"] for r in rows if r.get("canonical_key"))

    r = client.patch(f"/api/v1/documents/{doc_id}/line-items/{key}",
                     json={"value": 12345, "formula": "=A+B"})
    assert r.status_code == 200, r.text
    assert r.json()["status"] == "edited" and r.json()["value"] == "12345"

    st = client.get(f"/api/v1/documents/{doc_id}/statement", params={"statement": "balance_sheet"}).json()
    row = next(x for x in st["rows"] if x["id"] == key)
    assert row["v1"] == 12345 and row["status"] == "edited" and row["formula"] == "=A+B"
    assert row["editable"] is True


def test_formatted_statement_export_is_template_driven(client):
    """The default Excel export mirrors the run's template — one sheet per statement, with
    the template's sections/subtotals/totals and localized labels — for ANY such template."""
    import openpyxl
    import io as _io

    doc_id = _extract_with_ontology(client)

    # Statement layout (default) → a workbook with a Balance Sheet sheet built from the template.
    x = client.get(f"/api/v1/documents/{doc_id}/export", params={"fmt": "excel", "layout": "statement"})
    assert x.status_code == 200 and x.content[:2] == b"PK"
    wb = openpyxl.load_workbook(_io.BytesIO(x.content))
    assert "Balance Sheet" in wb.sheetnames
    ws = wb["Balance Sheet"]
    joined = " | ".join(str(v) for row in ws.iter_rows(values_only=True) for v in row if v)
    assert "Current assets" in joined                 # a template section header
    assert "Total current assets" in joined           # a template subtotal (structural anchor)
    assert "Trade receivables" in joined              # an extracted, template-labelled line

    # Localized (zh) labels come from the template's label_i18n.
    zh = client.get(f"/api/v1/documents/{doc_id}/export",
                    params={"fmt": "excel", "layout": "statement", "locale": "zh"})
    wbz = openpyxl.load_workbook(_io.BytesIO(zh.content))
    zjoined = " | ".join(str(v) for row in wbz["Balance Sheet"].iter_rows(values_only=True)
                         for v in row if v)
    assert "流动资产" in zjoined and "应收账款" in zjoined       # section + line in Chinese

    # Flat layout is still available.
    flat = client.get(f"/api/v1/documents/{doc_id}/export", params={"fmt": "excel", "layout": "flat"})
    assert flat.status_code == 200 and flat.content[:2] == b"PK"


def test_edit_then_revert_restores_original(client):
    doc_id = _extract_with_ontology(client)
    st0 = client.get(f"/api/v1/documents/{doc_id}/statement", params={"statement": "balance_sheet"}).json()
    item = next(r for r in st0["rows"] if r["kind"] == "item")
    key, original = item["id"], item["v1"]

    client.patch(f"/api/v1/documents/{doc_id}/line-items/{key}", json={"value": 99999, "formula": "=x"})
    st1 = client.get(f"/api/v1/documents/{doc_id}/statement", params={"statement": "balance_sheet"}).json()
    edited = next(r for r in st1["rows"] if r["id"] == key)
    assert edited["v1"] == 99999 and edited["status"] == "edited"

    r = client.delete(f"/api/v1/documents/{doc_id}/line-items/{key}")
    assert r.status_code == 200 and r.json()["reverted"] is True
    st2 = client.get(f"/api/v1/documents/{doc_id}/statement", params={"statement": "balance_sheet"}).json()
    reverted = next(r for r in st2["rows"] if r["id"] == key)
    assert reverted["v1"] == original and reverted["status"] is None and reverted["formula"] is None


def test_real_notes_index_and_detail(client):
    doc_id = _extract_with_ontology(client)
    notes = client.get(f"/api/v1/documents/{doc_id}/notes").json()
    assert notes["count"] >= 1 and notes["linked"] >= 1
    assert any(n["no"] == 15 for n in notes["notes"])        # make_native_pdf cites "Note 15"

    detail = client.get(f"/api/v1/documents/{doc_id}/notes/15").json()
    assert detail["no"] == 15 and detail["linked_label"]
    assert detail["rows"] and detail["reconciliation"] is None


def test_real_integrity_and_review_localized(client):
    doc_id = _extract_with_ontology(client)
    zh = client.get(f"/api/v1/documents/{doc_id}/integrity", params={"locale": "zh"}).json()
    assert any(s["label"] == "页数" for s in zh["stats"])       # "Pages" localized
    assert zh["grade"] and zh["grade"] != "Ready to extract"

    rev = client.get(f"/api/v1/documents/{doc_id}/review", params={"locale": "fr"}).json()
    assert rev["tabs"][0]["label"] == "Tous"                    # "All" localized (fr)


def test_can_access_enforces_ownership():
    from app.api.routes.documents import _can_access
    from app.security import Principal, Role

    class _Doc:
        owner = "alice"

    bob = Principal(role=Role.ANALYST, username="bob", name="Bob", via="session")
    alice = Principal(role=Role.ANALYST, username="alice", name="Alice", via="session")
    reviewer = Principal(role=Role.REVIEWER, username="rev", name="Rev", via="session")
    assert not _can_access(_Doc(), bob)          # another analyst's document
    assert _can_access(_Doc(), alice)            # the owner
    assert _can_access(_Doc(), reviewer)         # reviewer works the whole queue


def test_document_ownership_scoping(anon_client, auth):
    up = anon_client.post(
        "/api/v1/documents", files={"file": ("own.pdf", make_native_pdf(), "application/pdf")},
        headers=auth("analyst"),
    )
    doc_id = up.json()["id"]
    # The analyst sees their own upload; a reviewer (elevated) can open it too.
    mine = anon_client.get("/api/v1/documents", headers=auth("analyst")).json()["documents"]
    assert any(d["id"] == doc_id for d in mine)
    assert anon_client.get(f"/api/v1/documents/{doc_id}", headers=auth("reviewer")).status_code == 200
    assert anon_client.get(f"/api/v1/documents/{doc_id}/integrity", headers=auth("reviewer")).status_code == 200
    # Unauthenticated is rejected outright.
    assert anon_client.get(f"/api/v1/documents/{doc_id}").status_code == 401


def test_review_and_run_404_before_extraction(client):
    doc_id = client.post(
        "/api/v1/documents", files={"file": ("x.pdf", make_native_pdf(), "application/pdf")}
    ).json()["id"]
    # No run yet → /run and /export 404; /review returns an empty (but valid) queue.
    assert client.get(f"/api/v1/documents/{doc_id}/run").status_code == 404
    assert client.get(f"/api/v1/documents/{doc_id}/export", params={"fmt": "json"}).status_code == 404
    rev = client.get(f"/api/v1/documents/{doc_id}/review").json()
    assert rev["summary"] == {"open": 0, "passed": 0}


def test_missing_integrity_report_is_not_shown_as_clean(client):
    """A document with no stored integrity report must NOT be reported as a clean all-clear."""
    from app.api.routes.documents import _serialize_document_integrity

    class _Row:
        integrity_report = None
        page_count = 3
    body = _serialize_document_integrity(_Row())
    assert body["grade"] == "Not analyzed" and body["score"] == 0
    assert all(i["severity"] != "ok" for i in body["issues"])


def test_page_image_endpoint_renders_png(client):
    doc_id = client.post(
        "/api/v1/documents", files={"file": ("bs.pdf", make_native_pdf(), "application/pdf")}
    ).json()["id"]
    r = client.get(f"/api/v1/documents/{doc_id}/pages/0/image")
    assert r.status_code == 200
    assert r.headers["content-type"] == "image/png"
    assert r.content[:8] == b"\x89PNG\r\n\x1a\n" and len(r.content) > 100


def test_reference_ontology_seeded_and_attached_run_maps(client):
    """The shipped HK ontology is seeded; attaching it to a run populates canonical keys
    (deterministic here — the alias tier maps offline without an LLM)."""
    from app.config import get_settings
    from app.services import settings_state  # noqa: F401 (import parity with other tests)

    onts = client.get("/api/v1/ontologies").json()
    ont = next(o for o in onts if o["ontology_key"] == "hkfrs_hk_china_v1")
    tpls = client.get("/api/v1/templates").json()
    tpl = next(t for t in tpls if t["template_key"] == ont["target_template_key"])

    doc_id = client.post(
        "/api/v1/documents", files={"file": ("bs.pdf", make_native_pdf(), "application/pdf")}
    ).json()["id"]

    s = get_settings()
    prev = s.extraction.llm_mapping
    s.extraction.llm_mapping = False  # force the deterministic ensemble (no network) for the test
    try:
        r = client.post(f"/api/v1/documents/{doc_id}/extractions",
                        json={"ontology_version_id": ont["id"], "template_version_id": tpl["id"]})
        assert r.status_code == 202, r.text
        rows = _await_run(client, doc_id)["rows"]
        tr = next(row for row in rows if "Trade receivables" in row["source_label"])
        assert tr["canonical_key"], "expected the caption to map to a canonical concept"
    finally:
        s.extraction.llm_mapping = prev


def test_scanned_page_routes_through_ocr_port():
    """A scanned page is rasterized and sent to the configured OCR provider; its words
    (already normalized) reconstruct into line items with source_kind 'ocr'."""
    from app.config import get_settings
    from app.core.models.document import DocumentModel, PageSource
    from app.core.models.enums import DocFormat, PageKind, PageSourceKind
    from app.core.models.geometry import BBox
    from app.core.stage import PipelineContext
    from app.ports.registry import registry
    from app.services.pdf_extract import extract_pdf

    class _FakeOcr:
        id = "fakeocr"

        def recognize(self, image_bytes, *, lang="en"):
            return {"words": [
                {"text": "Cash", "bbox": BBox(x0=0.10, y0=0.20, x1=0.20, y1=0.23), "confidence": 0.99},
                {"text": "and", "bbox": BBox(x0=0.21, y0=0.20, x1=0.26, y1=0.23), "confidence": 0.99},
                {"text": "equivalents", "bbox": BBox(x0=0.27, y0=0.20, x1=0.40, y1=0.23), "confidence": 0.99},
                {"text": "1,204", "bbox": BBox(x0=0.80, y0=0.20, x1=0.90, y1=0.23), "confidence": 0.98},
            ], "angle": 0.0}

        def detect_orientation(self, image_bytes):
            return 0.0

    registry.register("ocr", "fakeocr", _FakeOcr)
    settings = get_settings()
    prev = settings.ocr.engine
    settings.ocr.engine = "fakeocr"
    try:
        doc = DocumentModel(filename="scan.pdf", content_hash="h1")
        doc.fmt = DocFormat.PDF
        doc.pages = [PageSource(index=0, source_kind=PageSourceKind.SCANNED, kind=PageKind.FACE)]
        ctx = PipelineContext(raw_bytes=make_native_pdf())
        added = extract_pdf(make_native_pdf(), doc, ctx)
        assert added >= 1
        item = doc.line_items[0]
        assert "Cash" in item.source_label
        ev = next(iter(item.values.values()))
        assert int(ev.value) == 1204
        assert ev.provenance.source_kind == "ocr" and ev.provenance.bbox is not None
    finally:
        settings.ocr.engine = prev
