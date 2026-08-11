"""Excel extraction reaches parity with the PDF path: a dedicated Note column becomes the
row's note reference, and a Consolidated / Standalone header band is extracted in one pass
(Req 15, 13)."""
from __future__ import annotations

import io

from app.core.models.enums import Basis


def _dual_basis_notes_xlsx() -> bytes:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"
    # Row 1: basis band; Row 2: Note col + period headers; Rows 3-4: data.
    ws["B1"] = "Consolidated"
    ws["D1"] = "Standalone"
    ws["A2"] = "Line item"; ws["B2"] = "Note"
    ws["C2"] = "Current"; ws["D2"] = "Prior"
    # Wait: band above C/D (consolidated) and E/F (standalone) — lay columns out explicitly.
    wb.remove(ws)

    ws = wb.create_sheet("Balance Sheet")
    # A=label B=Note C,D=Consolidated cur/prior  E,F=Standalone cur/prior
    ws["C1"] = "Consolidated"; ws["E1"] = "Standalone"
    ws["B2"] = "Note"; ws["C2"] = "Current"; ws["D2"] = "Prior"; ws["E2"] = "Current"; ws["F2"] = "Prior"
    ws["A3"] = "Trade receivables"; ws["B3"] = 15
    ws["C3"] = 3410; ws["D3"] = 2900; ws["E3"] = 3100; ws["F3"] = 2700
    ws["A4"] = "Cash and cash equivalents"; ws["B4"] = 14
    ws["C4"] = 1204; ws["D4"] = 980; ws["E4"] = 1050; ws["F4"] = 900
    buf = io.BytesIO(); wb.save(buf)
    return buf.getvalue()


def test_excel_note_column_and_dual_basis():
    from app.services.excel_extract import extract_workbook

    items = extract_workbook(_dual_basis_notes_xlsx(), document_id="h")
    tr = next(li for li in items if "Trade receivables" in li.source_label)
    assert tr.note_number == "15"                                   # Note column captured (m5)

    cc = tr.get_value(Basis.CONSOLIDATED, period_label="current")
    cp = tr.get_value(Basis.CONSOLIDATED, period_label="prior")
    sc = tr.get_value(Basis.STANDALONE, period_label="current")
    sp = tr.get_value(Basis.STANDALONE, period_label="prior")
    assert cc and int(cc.value) == 3410 and cp and int(cp.value) == 2900     # consolidated
    assert sc and int(sc.value) == 3100 and sp and int(sp.value) == 2700     # standalone (m3)
    # The note number is not swallowed as a value.
    assert all(int(ev.value) != 15 for ev in tr.values.values())
