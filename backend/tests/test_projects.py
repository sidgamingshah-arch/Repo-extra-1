from __future__ import annotations

from decimal import Decimal

from app.sample.demo import BALANCE_SHEET
from app.services import checks as ce


def test_project_and_statement(client):
    r = client.get("/api/v1/projects/demo")
    assert r.status_code == 200
    assert r.json()["project"]["entity"] == "Reliance Industries Ltd"

    r = client.get("/api/v1/projects/demo/statements/balance_sheet?basis=consolidated")
    body = r.json()
    ppe = next(x for x in body["rows"] if x["id"] == "ppe")
    assert ppe["v1"] == 423180
    assert ppe["confidence"]["pct"] == 96
    assert ppe["note"] == "3"


def test_localized_statement_labels(client):
    # Chinese output labels; source_label stays English (input→output parity).
    zh = client.get("/api/v1/projects/demo/statements/balance_sheet?locale=zh").json()
    ppe = next(x for x in zh["rows"] if x["id"] == "ppe")
    assert ppe["label"] == "物业、厂房及设备"
    assert ppe["source_label"] == "Property, plant and equipment"
    assert zh["label"] == "资产负债表"  # statement name localized too

    # Arabic + French resolve; unknown locale falls back to English.
    ar = client.get("/api/v1/projects/demo/statements/balance_sheet?locale=ar").json()
    assert next(x for x in ar["rows"] if x["id"] == "tot_assets")["label"] == "إجمالي الأصول"
    fr = client.get("/api/v1/projects/demo/statements/balance_sheet?locale=fr").json()
    assert next(x for x in fr["rows"] if x["id"] == "trade_recv")["label"] == "Créances clients"
    en = client.get("/api/v1/projects/demo/statements/balance_sheet?locale=en").json()
    assert next(x for x in en["rows"] if x["id"] == "ppe")["label"] == "Property, plant and equipment"


def test_standalone_scaling():
    from app.api.routes.projects import _scale
    assert _scale(423180, "standalone") == round(423180 * 0.88)
    assert _scale(423180, "consolidated") == 423180


def test_edit_override_roundtrip(client):
    r = client.patch("/api/v1/projects/demo/line-items/ppe",
                     json={"value": 430000, "formula": "Note3.net_block + 6820"})
    assert r.status_code == 200
    body = client.get("/api/v1/projects/demo/statements/balance_sheet").json()
    ppe = next(x for x in body["rows"] if x["id"] == "ppe")
    assert ppe["v1"] == 430000
    assert ppe["status"] == "edited"
    assert ppe["formula"] == "Note3.net_block + 6820"
    # revert
    client.delete("/api/v1/projects/demo/line-items/ppe")
    body = client.get("/api/v1/projects/demo/statements/balance_sheet").json()
    ppe = next(x for x in body["rows"] if x["id"] == "ppe")
    assert ppe["v1"] == 423180


def test_localized_dynamic_content(client):
    # Review checks, integrity issues, and note detail localize with ?locale.
    zh_review = client.get("/api/v1/projects/demo/review?locale=zh").json()
    assert zh_review["checks"][0]["title"] == "资产负债表不平衡"
    fr_integ = client.get("/api/v1/projects/demo/integrity?locale=fr").json()
    assert any(i["title"] == "Pages pivotées" for i in fr_integ["issues"])
    ar_note = client.get("/api/v1/projects/demo/notes/12?locale=ar").json()
    assert ar_note["reconciliation"] and "الإيضاح" in ar_note["reconciliation"]


def test_review_notes_endpoints(client):
    review = client.get("/api/v1/projects/demo/review").json()
    assert len(review["checks"]) == 4

    note = client.get("/api/v1/projects/demo/notes/12").json()
    assert note["title"] == "Trade Receivables"
    assert note["linked_line"] == "trade_recv"


def test_a_note_and_a_statement_cannot_label_the_same_figures_differently(client):
    """The finding, as an assertion: the Notes screen hardcoded "FY25"/"FY24" over the same two
    columns the Workspace labelled from the project's real periods. Both now read one list."""
    note = client.get("/api/v1/projects/demo/notes/12").json()
    stmt = client.get("/api/v1/projects/demo/statements/balance_sheet").json()
    assert note["periods"] == stmt["periods"]
    assert len(note["periods"]) == 2 and all(note["periods"])


def test_every_sample_viewer_carries_exactly_one_active_chip(client):
    """A chip is a LABEL naming what the viewer is showing. Each statement used to carry a second
    `active: False` note chip — which is a tab control — and this viewer has no second tab: the
    sample source pane is a rendered paper mock with no pages behind it. One chip is also the shape
    the real routes serve, so the two paths cannot drift apart again."""
    for statement in ("balance_sheet", "profit_and_loss", "cash_flow"):
        viewer = client.get(f"/api/v1/projects/demo/statements/{statement}").json()["viewer"]
        assert len(viewer["chips"]) == 1, statement
        assert viewer["chips"][0]["active"] is True
        # The dropped note references are not lost — the callout names them in prose.
        assert "Note" in viewer["callout"]


def test_the_sample_export_checklist_serves_only_options_the_workbook_reads(client):
    """An option the builder never reads is a switch wired to nothing. `build_xlsx` honours
    `confidence` and `notes_sheet` and no others, so those are the only two offered."""
    options = client.get("/api/v1/projects/demo/export-options").json()["options"]
    assert {o["key"] for o in options} == {"confidence", "notes_sheet"}


def test_the_sample_export_checklist_gates_the_file(client):
    import io

    import openpyxl

    def workbook(include: dict):
        r = client.post("/api/v1/projects/demo/export",
                        json={"format": "excel", "include": include})
        assert r.status_code == 200, r.text
        return openpyxl.load_workbook(io.BytesIO(r.content))

    on = workbook({"confidence": True, "notes_sheet": True})
    assert "All notes" in on.sheetnames
    assert "Conf." in [c.value for c in on[on.sheetnames[0]][1]]

    off = workbook({"confidence": False, "notes_sheet": False})
    assert "All notes" not in off.sheetnames
    assert "Conf." not in [c.value for c in off[off.sheetnames[0]][1]]


def test_sample_counts_describe_the_lists_they_head(client):
    """Every count the sample serves is counted from the rows served with it.

    This replaces an assertion that pinned the defect: `summary["open"] == 12` was true of a
    stored literal and false of the four checks beside it, so the screen read "12 open" over a
    list of four. The same three literals headed all three screens — 84 pages over ten cards,
    48 notes over twelve, 12 findings over four — so the invariant, not the number, is the test.
    """
    review = client.get("/api/v1/projects/demo/review").json()
    assert review["summary"]["open"] == len(review["checks"])
    by_label = {t["label"]: t["count"] for t in review["tabs"]}
    assert by_label["All"] == len(review["checks"])
    # Each tab counts its own type, so the per-type tabs sum to the whole list exactly once.
    assert sum(c for label, c in by_label.items() if label != "All") == len(review["checks"])
    # …and each tab's count is the count of what it SELECTS, which is what the client filters by.
    for tab in review["tabs"]:
        selected = (review["checks"] if tab["types"] is None
                    else [c for c in review["checks"] if c["type"] in tab["types"]])
        assert tab["count"] == len(selected), tab["label"]

    # The Export footer's two figures, from the same seeded dataset the Review header counts.
    progress = client.get("/api/v1/projects/demo").json()["project"]["progress"]
    statements = {name: client.get(f"/api/v1/projects/demo/statements/{name}").json()
                  for name in ("balance_sheet", "profit_and_loss", "cash_flow")}
    items = sum(1 for s in statements.values() for r in s["rows"] if r.get("kind") == "item")
    assert progress["line_items"] == items

    # The header's third tile says "lines with no finding", so `passed` counts the statement LINES
    # named by no served finding — the same quantity over the same POPULATION the real route serves
    # under the same tile. It used to be `items - len(checks)`, which assumed one finding per line
    # item; then it was the ITEM rows less the item rows a finding names, which is a THIRD population:
    # the real path counts subtotals and totals as lines (a total is what the balance card names), so
    # the sample was answering 31 over its 33 item rows while serving 6 subtotals and 4 totals beside
    # them, 8 of which no finding names. Read off `names` — the field the real route serves for the
    # same purpose — not re-derived here.
    named = {n for c in review["checks"] for n in c["names"]}
    assert named == {c["target"] for c in review["checks"]}
    rows = [r for s in statements.values() for r in s["rows"]]
    lines = [r for r in rows if r.get("kind") not in ("section", "subhead")]
    indicted = [r for r in lines if r.get("id") in named]
    assert review["summary"]["passed"] == len(lines) - len(indicted)
    # THE ASSERTION THAT FAILS WITH THE DEFECT RESTORED: the two populations are provably different
    # numbers on this very data, so the item-only count cannot pass as the served one.
    assert len(lines) > items
    assert review["summary"]["passed"] != items - len(
        [r for r in indicted if r.get("kind") == "item"])
    assert 0 < len(indicted) <= len(review["checks"])
    # …and a subtotal/total IS in the population: findings name two of them here, which is exactly the
    # row class the item-only count dropped from both sides of the subtraction.
    assert len([r for r in indicted if r.get("kind") != "item"]) == 2
    assert progress["in_review"] == review["summary"]["open"] == len(review["checks"])
    # `pct` is gone rather than derived: "how far through the workflow" has no source in the
    # sample, and 72 was a literal. A number with no source is not served under any name.
    assert "pct" not in progress
    assert set(progress) == {"line_items", "in_review"}

    notes = client.get("/api/v1/projects/demo/notes").json()
    assert notes["count"] == len(notes["notes"])
    # `linked` is statement lines citing one of those notes — real lines, so it cannot exceed them.
    assert 0 < notes["linked"] <= 60

    pages = client.get("/api/v1/projects/demo/pages").json()
    assert pages["total"] == len(pages["pages"])
    assert pages["focused"] == sum(1 for p in pages["pages"] if p["included"])
    assert pages["skipped"] == pages["total"] - pages["focused"]
    # The chips are matched to page kinds POSITIONALLY by the client: all / face / notes / other.
    chips = pages["filters"]
    assert [c["label"] for c in chips] == ["All pages", "Face", "Notes", "Other"]
    assert chips[0]["count"] == len(pages["pages"])
    for i, kind in enumerate(("face", "notes", "other"), start=1):
        assert chips[i]["count"] == sum(1 for p in pages["pages"] if p["kind"] == kind)


def test_sample_page_filter_labels_localize(client):
    """The renamed chips still translate — a derived label is no use if it only exists in en."""
    zh = client.get("/api/v1/projects/demo/pages?locale=zh").json()
    assert [c["label"] for c in zh["filters"]] == ["全部页面", "表内", "附注", "其他"]


def test_export_xlsx_and_json(client):
    r = client.post("/api/v1/projects/demo/export", json={"format": "json"})
    assert r.status_code == 200 and r.headers["content-type"].startswith("application/json")
    assert b"Trade receivables" in r.content

    r = client.post("/api/v1/projects/demo/export", json={"format": "excel"})
    assert r.status_code == 200
    assert r.content[:2] == b"PK"  # xlsx is a zip container


# --- generic checks engine (used for real extractions) ---

def test_balance_check_passes_on_balanced_sheet():
    rows = [{"id": r["id"], "label": r["label"], "kind": r.get("kind", "item"),
             "v1": r.get("v1"), "v2": r.get("v2")} for r in BALANCE_SHEET]
    results = ce.check_balance(rows)
    assert results and results[0].status == "pass"
    assert results[0].delta == Decimal(0)


def test_balance_check_fails_on_imbalance():
    rows = [{"id": "tot_assets", "kind": "total", "v1": 1268100},
            {"id": "tot_eq", "kind": "total", "v1": 1266860}]
    r = ce.check_balance(rows)[0]
    assert r.status == "fail" and r.delta == Decimal(1240)


def test_subtotal_rollup_detects_mismatch():
    rows = [
        {"id": "sh", "kind": "subhead", "label": "X"},
        {"id": "a", "kind": "item", "label": "a", "v1": 100},
        {"id": "b", "kind": "item", "label": "b", "v1": 200},
        {"id": "sub", "kind": "subtotal", "label": "Sub", "v1": 310},  # should be 300
    ]
    checks = ce.check_subtotals(rows)
    assert checks[0].status == "fail" and checks[0].delta == Decimal(10)


def test_sign_anomaly_flags_positive_expense():
    rows = [{"id": "fin", "kind": "item", "label": "Finance costs", "v1": 18400}]
    checks = ce.check_signs(rows)
    assert checks and checks[0].type == "sign" and checks[0].target == "fin"
