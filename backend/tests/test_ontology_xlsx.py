"""The rulebook workbook: a round trip that loses nothing, and refuses what it cannot read."""
from __future__ import annotations

import json
from pathlib import Path

import pytest

from app.services.ontology_xlsx import (
    EMPTY_MARKER,
    SHEET_CONCEPTS,
    SHEET_CONFIG,
    OntologySheetError,
    build_ontology_xlsx,
    parse_ontology_xlsx,
)

pytest.importorskip("openpyxl")

SHIPPED = (Path(__file__).resolve().parents[1] / "app" / "sample" / "templates"
           / "hkfrs_hk_china_ontology.json")


def _shipped() -> dict:
    return json.loads(SHIPPED.read_text())


def test_the_shipped_rulebook_round_trips_byte_for_byte():
    """THE GUARANTEE. Not a fixture — the real 183-concept rulebook: export it, read it back, and the
    definition must be the SAME definition. An exporter that quietly drops a block would otherwise
    hand an admin a file whose upload silently publishes a smaller rulebook than the one they edited,
    which is the failure this codebase spends most of its guards on.
    """
    original = _shipped()
    rebuilt = parse_ontology_xlsx(build_ontology_xlsx(original))
    assert rebuilt == original


def test_every_concept_and_field_survives():
    """Stated separately from equality so a failure says WHAT went missing rather than dumping two
    183-entry structures side by side."""
    original = _shipped()
    rebuilt = parse_ontology_xlsx(build_ontology_xlsx(original))
    assert len(rebuilt["mappings"]) == len(original["mappings"]) == 183
    assert [m["canonical_key"] for m in rebuilt["mappings"]] \
        == [m["canonical_key"] for m in original["mappings"]]
    for a, b in zip(original["mappings"], rebuilt["mappings"]):
        assert a == b, a["canonical_key"]
    assert set(rebuilt) == set(original)
    for key in original:
        assert rebuilt[key] == original[key], key


def test_a_deliberately_empty_list_is_not_confused_with_an_unset_field():
    """THE DISTINCTION THE FORMAT RESTS ON. Thirteen residual buckets carry ``aliases: []`` because
    they match nothing and are filled by the sweep; if blank meant "empty" there would be no way to
    say "unset", and if blank meant "unset" those thirteen declarations would vanish on the first
    round trip — turning alias matching back on for every residual.
    """
    original = _shipped()
    residuals = [m for m in original["mappings"] if m.get("value_scope") == "exclusive_residual"]
    assert len(residuals) == 13
    assert all(m.get("aliases") == [] for m in residuals), "premise moved"

    rebuilt = parse_ontology_xlsx(build_ontology_xlsx(original))
    by_key = {m["canonical_key"]: m for m in rebuilt["mappings"]}
    for m in residuals:
        got = by_key[m["canonical_key"]]
        assert "aliases" in got and got["aliases"] == [], m["canonical_key"]
    # …and a field that genuinely is not set stays absent rather than arriving as an empty list.
    unset = next(m for m in original["mappings"] if "regex_hints" not in m)
    assert "regex_hints" not in by_key[unset["canonical_key"]]


def test_the_workbook_the_admin_gets_has_the_sheets_the_readme_names():
    import openpyxl
    wb = openpyxl.load_workbook(__import__("io").BytesIO(build_ontology_xlsx(_shipped())))
    assert wb.sheetnames == [SHEET_CONCEPTS, "Section defaults", "Netting rules",
                             SHEET_CONFIG, "README"]
    ws = wb[SHEET_CONCEPTS]
    assert ws.max_row == 184                      # 183 concepts + the header
    assert ws["A1"].value == "Canonical key"
    assert ws.freeze_panes == "B2"                # the key column stays visible while scrolling
    # Per-locale alias columns, derived from what the rulebook supports.
    headers = [c.value for c in ws[1]]
    assert "Aliases (en)" in headers and "Aliases (zh)" in headers
    readme = "\n".join(str(r[0].value or "") for r in wb["README"].iter_rows())
    assert EMPTY_MARKER in readme and "BLANK versus EMPTY" in readme


def test_an_edited_alias_reaches_the_rebuilt_rulebook():
    """The point of the exercise: a change made in Excel is the change that is published."""
    import io

    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(build_ontology_xlsx(_shipped())))
    ws = wb[SHEET_CONCEPTS]
    headers = [c.value for c in ws[1]]
    key_col = headers.index("Canonical key") + 1
    alias_col = headers.index("Aliases") + 1
    target_row = next(r for r in range(2, ws.max_row + 1)
                      if ws.cell(r, key_col).value == "bs_current_assets__inventories")
    ws.cell(target_row, alias_col).value = "Inventories\nStocks\nStock-in-trade"
    buf = io.BytesIO()
    wb.save(buf)

    rebuilt = parse_ontology_xlsx(buf.getvalue())
    m = next(x for x in rebuilt["mappings"]
             if x["canonical_key"] == "bs_current_assets__inventories")
    assert m["aliases"] == ["Inventories", "Stocks", "Stock-in-trade"]


def test_it_refuses_what_it_cannot_read_rather_than_guessing():
    import io

    import openpyxl

    # Not a workbook at all.
    with pytest.raises(OntologySheetError, match="readable .xlsx"):
        parse_ontology_xlsx(b"this is not a workbook")

    # A workbook with none of the expected sheets.
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(OntologySheetError, match="Missing the 'Concepts' sheet"):
        parse_ontology_xlsx(buf.getvalue())

    # A renamed header — the reader is header-driven, so it says so instead of reading the wrong
    # column and publishing a rulebook whose definitions landed in the aliases.
    wb2 = openpyxl.load_workbook(io.BytesIO(build_ontology_xlsx(_shipped())))
    wb2[SHEET_CONCEPTS]["E1"] = "Synonyms"
    b2 = io.BytesIO()
    wb2.save(b2)
    with pytest.raises(OntologySheetError, match="missing column"):
        parse_ontology_xlsx(b2.getvalue())

    # Two rows claiming one concept.
    wb3 = openpyxl.load_workbook(io.BytesIO(build_ontology_xlsx(_shipped())))
    ws3 = wb3[SHEET_CONCEPTS]
    ws3.cell(3, 1).value = ws3.cell(2, 1).value
    b3 = io.BytesIO()
    wb3.save(b3)
    with pytest.raises(OntologySheetError, match="appears more than once"):
        parse_ontology_xlsx(b3.getvalue())


def test_the_rebuilt_definition_passes_the_same_gate_as_a_json_upload():
    """A workbook must not be a second, looser door. The rebuilt definition has to satisfy the
    loader AND the undeclared-key check the JSON upload applies."""
    from app.schemas.loader import load_ontology, resolve_inherits, unknown_keys

    rebuilt = parse_ontology_xlsx(build_ontology_xlsx(_shipped()))
    ont = load_ontology(rebuilt)
    assert len(ont.mappings) == 183
    assert unknown_keys(rebuilt, ont, limit=500) == []
    resolve_inherits(rebuilt)          # every `inherits` still names a real section


def test_a_field_this_module_has_never_heard_of_still_round_trips():
    """WHY THE CATCH-ALL EXISTS. The round-trip test above caught three netting fields absent from
    this module's column list — ``evidence_required``, ``on_apply``, ``decompose_into`` — which would
    have been dropped on the first export. Columns were added for those, but the lesson is that any
    column list goes stale as the schema grows, and a rulebook must not need this module edited before
    it can survive a round trip. Anything uncovered travels in the catch-all column instead.
    """
    original = _shipped()
    original["mappings"][0]["some_future_field"] = {"nested": ["value"]}
    original["netting_rules"][0]["another_new_one"] = 7
    rebuilt = parse_ontology_xlsx(build_ontology_xlsx(original))
    assert rebuilt["mappings"][0]["some_future_field"] == {"nested": ["value"]}
    assert rebuilt["netting_rules"][0]["another_new_one"] == 7
    assert rebuilt == original


def test_one_field_claimed_by_two_cells_is_refused():
    """A key in its own column AND inside the catch-all is ambiguous. Picking a winner would discard
    an edit silently, which is the whole failure mode this format is trying to avoid."""
    import io

    import openpyxl
    from app.services.ontology_xlsx import OTHER_COLUMN

    wb = openpyxl.load_workbook(io.BytesIO(build_ontology_xlsx(_shipped())))
    ws = wb[SHEET_CONCEPTS]
    headers = [c.value for c in ws[1]]
    ws.cell(2, headers.index(OTHER_COLUMN) + 1).value = '{"aliases": ["clash"]}'
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(OntologySheetError, match="appears both in its own column"):
        parse_ontology_xlsx(buf.getvalue())


def test_the_workbook_downloads_and_uploads_through_the_api(client):
    """End to end through the routes: download the rulebook in force, upload it back, and a new
    version is published with the same concept count — the workbook is a real authoring path, not a
    read-only report."""
    from app.db.base import SessionLocal
    from app.db.models import OntologyVersion
    from app.services.ontology_select import select_for_template

    with SessionLocal() as s:
        row = select_for_template(s, "hkfrs_hk_china_v1")
        oid, before_version = row.id, row.version

    got = client.get(f"/api/v1/ontologies/{oid}/xlsx")
    assert got.status_code == 200, got.text
    assert "spreadsheetml" in got.headers["content-type"]
    assert ".xlsx" in got.headers["content-disposition"]

    posted = client.post("/api/v1/ontologies/xlsx",
                         files={"file": ("edited.xlsx", got.content,
                                         "application/vnd.openxmlformats-officedocument."
                                         "spreadsheetml.sheet")})
    assert posted.status_code == 201, posted.text
    body = posted.json()
    assert body["mappings"] == 183
    assert body["version"] == before_version + 1        # a new version, nothing overwritten
    try:
        # And it validated against the template, exactly as a JSON upload would.
        assert body["validated_against_template"]["template_key"] == "hkfrs_hk_china_v1"
    finally:
        with SessionLocal() as s:
            s.delete(s.get(OntologyVersion, body["id"]))
            s.commit()


def test_a_workbook_that_would_publish_a_broken_rulebook_is_refused(client):
    """The workbook must not be a looser door than the JSON upload. A canonical key the template does
    not declare is refused here too — with a 422, not a published rulebook that maps onto nothing."""
    import io

    import openpyxl
    from app.db.base import SessionLocal
    from app.services.ontology_select import select_for_template

    with SessionLocal() as s:
        oid = select_for_template(s, "hkfrs_hk_china_v1").id
    data = client.get(f"/api/v1/ontologies/{oid}/xlsx").content
    wb = openpyxl.load_workbook(io.BytesIO(data))
    wb[SHEET_CONCEPTS].cell(2, 1).value = "not_a_key_the_template_declares"
    buf = io.BytesIO()
    wb.save(buf)
    posted = client.post("/api/v1/ontologies/xlsx",
                         files={"file": ("bad.xlsx", buf.getvalue(),
                                         "application/vnd.openxmlformats-officedocument."
                                         "spreadsheetml.sheet")})
    assert posted.status_code == 422, posted.text
