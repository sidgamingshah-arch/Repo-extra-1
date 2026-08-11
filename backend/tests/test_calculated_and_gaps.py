"""Calculated lines carry their COMPUTED figure, and a gap gets one chance to be explained.

A subtotal printed on the page is a fourth opinion alongside the lines it is meant to be the sum
of. These tests pin the consequences of preferring the computation: what the face shows, what the
printed figure is kept for, and what happens when the two disagree — including the LLM layer that
asks whether a line the mapper could not place belongs in the section that is short by its amount.
"""
from __future__ import annotations

import pytest

from app.api.routes.documents import _accounting_checks, _build_statement
from app.services.rollups import evaluate, evaluate_rows


def _v(period, value, *, basis="consolidated", page=3):
    return {"basis": basis, "period_label": period, "value": str(value),
            "provenance": {"source_kind": "native", "page_index": page,
                           "bbox": {"x0": 0.6, "y0": 0.2, "x1": 0.7, "y1": 0.21}}}


def _row(key, label, cur=None, prior=None, **kw):
    vals = []
    if cur is not None:
        vals.append(_v("current", cur, **kw))
    if prior is not None:
        vals.append(_v("prior", prior, **kw))
    return {"canonical_key": key, "source_label": label, "values": vals}


# A section with two lines, a residual bucket, and a subtotal calculated from all three.
TEMPLATE = {
    "schema_version": 1, "template_key": "t", "name": "T",
    "statements": [{"type": "balance_sheet", "sections": [
        {"node_id": "sec", "canonical_key": "bs_ca", "label": "Current assets", "role": "header",
         "children": [
             {"node_id": "n1", "canonical_key": "bs_ca__inventories", "label": "Inventories",
              "role": "line", "rollup": None},
             {"node_id": "n2", "canonical_key": "bs_ca__cash", "label": "Cash", "role": "line",
              "rollup": None},
             {"node_id": "n3", "canonical_key": "bs_ca__others", "label": "Others",
              "role": "line", "rollup": None},
             {"node_id": "n4", "canonical_key": "bs_ca__total", "label": "Total current assets",
              "role": "subtotal",
              "rollup": {"op": "sum", "children": ["bs_ca__inventories", "bs_ca__cash",
                                                   "bs_ca__others"]}},
         ]},
        {"node_id": "sec2", "canonical_key": "bs_cl", "label": "Current liabilities",
         "role": "header", "children": [
             {"node_id": "m1", "canonical_key": "bs_cl__payables", "label": "Payables",
              "role": "line", "rollup": None},
             {"node_id": "m2", "canonical_key": "bs_cl__total", "label": "Total current liabilities",
              "role": "subtotal",
              "rollup": {"op": "sum", "children": ["bs_cl__payables"]}},
         ]},
        # A net figure: first term minus the rest.
        {"node_id": "net", "canonical_key": "bs_net_current", "label": "Net current assets",
         "role": "total", "children": [],
         "rollup": {"op": "diff", "children": ["bs_ca__total", "bs_cl__total"]}},
    ]}],
}


def _stmt(rows, **kw):
    return _build_statement(rows, TEMPLATE, "balance_sheet", "f.pdf", **kw)


def _row_of(d, key):
    return next(r for r in d["rows"] if r["id"] == key)


# --------------------------------------------------------------------------------------
# What the face shows
# --------------------------------------------------------------------------------------
def test_a_subtotal_shows_its_components_not_the_printed_figure():
    # The document printed 999 for a section its own lines say is 130. Showing 999 would put a
    # figure on the face that the lines beneath it contradict.
    rows = [_row("bs_ca__inventories", "Inventories", 100, 90),
            _row("bs_ca__cash", "Cash", 30, 20),
            _row("bs_ca__total", "Total current assets", 999, 888)]
    row = _row_of(_stmt(rows), "bs_ca__total")
    assert (row["v1"], row["v2"]) == (130, 110)
    assert row["origin"] == "calculated"
    # The printed figure is retained — it is what the divergence is measured against.
    assert (row["reported1"], row["reported2"]) == (999, 888)


def test_a_nested_total_rolls_up_the_computed_subtotals_not_the_printed_ones():
    rows = [_row("bs_ca__inventories", "Inventories", 100),
            _row("bs_ca__cash", "Cash", 30),
            _row("bs_ca__total", "Total current assets", 999),      # printed, wrong
            _row("bs_cl__payables", "Payables", 50),
            _row("bs_cl__total", "Total current liabilities", 777)]  # printed, wrong
    net = _row_of(_stmt(rows), "bs_net_current")
    # 130 − 50, not 999 − 777. A total that summed the printed subtotals would inherit both errors.
    assert net["v1"] == 80
    assert net["origin"] == "calculated"


def test_a_calculated_line_the_document_never_printed_is_still_filled_in():
    rows = [_row("bs_ca__inventories", "Inventories", 100),
            _row("bs_ca__cash", "Cash", 30)]
    row = _row_of(_stmt(rows), "bs_ca__total")
    assert row["v1"] == 130
    assert row["origin"] == "calculated"
    assert row["status"] is None            # not a gap: the template says what it is made of


def test_a_subtotal_with_no_extracted_components_falls_back_to_the_printed_figure():
    rows = [_row("bs_ca__total", "Total current assets", 999)]
    row = _row_of(_stmt(rows), "bs_ca__total")
    assert row["v1"] == 999
    assert row["origin"] == "reported_uncomputed"


def test_a_computed_subtotal_lists_the_components_it_came_from_with_their_pages():
    rows = [_row("bs_ca__inventories", "Inventories", 100, 90, page=4),
            _row("bs_ca__cash", "Cash", 30, 20, page=5)]
    row = _row_of(_stmt(rows), "bs_ca__total")
    contribs = {c["label"]: c for c in row["contributions"]}
    assert contribs["Inventories"]["v1"] == 100 and contribs["Inventories"]["v2"] == 90
    assert contribs["Inventories"]["source"]["page_index"] == 4
    assert contribs["Cash"]["source"]["page_index"] == 5
    # A component the document never yielded is listed as absent, not silently dropped.
    assert contribs["Others"]["v1"] is None and contribs["Others"]["residual"] is True


def test_the_prior_column_lists_the_prior_periods_components():
    rows = [_row("bs_ca__inventories", "Inventories", 100, 90),
            _row("bs_ca__cash", "Cash", 30, 20)]
    row = _row_of(_stmt(rows), "bs_ca__total")
    assert [c["v2"] for c in row["contributions"]] == [90, 20, None]


def test_a_diff_rollup_subtracts_every_term_after_the_first():
    reported = {"a": 500.0, "b": 120.0, "c": 30.0}.get
    tpl = {"statements": [{"type": "balance_sheet", "sections": [
        {"canonical_key": "net", "node_id": "net", "label": "Net", "role": "total",
         "rollup": {"op": "diff", "children": ["a", "b", "c"]}}]}]}
    assert evaluate(tpl, reported)["net"].value == 350.0


def test_a_rollup_cycle_is_reported_rather_than_recursed_into():
    tpl = {"statements": [{"type": "balance_sheet", "sections": [
        {"canonical_key": "a", "node_id": "a", "label": "A", "role": "total",
         "rollup": {"op": "sum", "children": ["b"]}},
        {"canonical_key": "b", "node_id": "b", "label": "B", "role": "total",
         "rollup": {"op": "sum", "children": ["a"]}}]}]}
    out = evaluate(tpl, lambda k: None)          # must return, not blow the stack
    assert out["a"].cycle and out["b"].cycle


# --------------------------------------------------------------------------------------
# A manual value outranks the arithmetic
# --------------------------------------------------------------------------------------
def test_a_manual_value_outranks_the_computed_one():
    rows = [_row("bs_ca__inventories", "Inventories", 100),
            _row("bs_ca__cash", "Cash", 30),
            {**_row("bs_ca__total", "Total current assets", 555),
             "edited": True, "edited_slots": ["consolidated/current"]}]
    row = _row_of(_stmt(rows), "bs_ca__total")
    assert row["v1"] == 555 and row["origin"] == "manual"
    # …and the analyst can still see what they overrode.
    assert row["calculated1"] == 130


def test_correcting_a_component_moves_every_subtotal_above_it():
    rows = [{**_row("bs_ca__inventories", "Inventories", 100),
             "edited": True, "edited_slots": ["consolidated/current"]},
            _row("bs_ca__cash", "Cash", 30),
            _row("bs_cl__payables", "Payables", 50)]
    d = _stmt(rows)
    assert _row_of(d, "bs_ca__total")["v1"] == 130
    assert _row_of(d, "bs_net_current")["v1"] == 80


# --------------------------------------------------------------------------------------
# The review queue is built from the difference
# --------------------------------------------------------------------------------------
def test_a_printed_subtotal_that_differs_from_its_components_becomes_a_review_item():
    rows = [_row("bs_ca__inventories", "Inventories", 100),
            _row("bs_ca__cash", "Cash", 30),
            _row("bs_ca__total", "Total current assets", 999)]
    checks = _accounting_checks(rows, [], "en", [], TEMPLATE)
    item = next(c for c in checks if c["type"] == "calculated_mismatch")
    assert item["target"] == "bs_ca__total"
    assert item["delta"] == "-869"                 # computed 130 vs printed 999
    labels = [row[0] for row in item["calc"]]
    assert "Printed in the document" in labels and "Computed from components" in labels
    assert "Inventories" in labels                 # the arithmetic travels with the finding


def test_a_subtotal_that_agrees_with_its_components_raises_nothing():
    rows = [_row("bs_ca__inventories", "Inventories", 100),
            _row("bs_ca__cash", "Cash", 30),
            _row("bs_ca__total", "Total current assets", 130)]
    checks = _accounting_checks(rows, [], "en", [], TEMPLATE)
    assert [c for c in checks if c["type"] == "calculated_mismatch"] == []


def test_a_printed_subtotal_with_no_components_is_flagged_as_unverified():
    rows = [_row("bs_ca__total", "Total current assets", 999)]
    checks = _accounting_checks(rows, [], "en", [], TEMPLATE)
    item = next(c for c in checks if c["type"] == "uncomputed")
    assert item["target"] == "bs_ca__total" and item["tone"] == "med"


# --------------------------------------------------------------------------------------
# The Excel export must not disagree with the screen
# --------------------------------------------------------------------------------------
def test_the_workbook_holds_the_computed_subtotal_and_notes_the_printed_one():
    pytest.importorskip("openpyxl")
    import io

    import openpyxl

    from app.services.export import build_statement_workbook

    rows = [_row("bs_ca__inventories", "Inventories", 100),
            _row("bs_ca__cash", "Cash", 30),
            _row("bs_ca__total", "Total current assets", 999)]
    wb = openpyxl.load_workbook(io.BytesIO(
        build_statement_workbook(rows, TEMPLATE, filename="f.pdf")))
    ws = wb["Balance Sheet"]
    cells = {ws.cell(r, 1).value: r for r in range(1, ws.max_row + 1)}
    r = cells["Total current assets"]
    # Column C is the first value column (consolidated / current).
    assert ws.cell(r, 3).value == 130
    note = ws.cell(r, 1).comment.text
    assert "computed 130" in note and "document printed 999" in note


# --------------------------------------------------------------------------------------
# Gap closing: arithmetic proposes, the model disposes
# --------------------------------------------------------------------------------------
def _gap_rows():
    """A section short by 40 in both periods, and three unplaced lines — one of which fits."""
    return [
        _row("bs_ca__inventories", "Inventories", 100, 90),
        _row("bs_ca__cash", "Cash", 30, 20),
        _row("bs_ca__total", "Total current assets", 170, 130),      # printed; components give 130/110
        {"canonical_key": None, "source_label": "Pledged bank deposits",
         "values": [_v("current", 40, page=8), _v("prior", 20, page=8)]},
        {"canonical_key": None, "source_label": "Number of employees",
         "values": [_v("current", 1240, page=88)]},
        {"canonical_key": None, "source_label": "Directors' emoluments",
         "values": [_v("current", 40, page=91), _v("prior", 55, page=91)]},
    ]


def test_only_groups_that_close_the_gap_in_BOTH_periods_are_offered():
    from app.services.gap_closing import find_gaps, leftovers, viable_subsets

    rows = _gap_rows()
    gap = next(g for g in find_gaps(rows, TEMPLATE, "consolidated")
               if g.target_key == "bs_ca__total")
    assert gap.current == 40 and gap.prior == 20
    options = viable_subsets(leftovers(rows, TEMPLATE, "consolidated"), gap)
    captions = [[c.label for c in o] for o in options]
    # "Pledged bank deposits" is 40/20 — it closes both. "Directors' emoluments" is 40/55: it
    # matches this year by coincidence and last year not at all, so it is never offered.
    assert ["Pledged bank deposits"] in captions
    assert ["Directors' emoluments"] not in captions


class _Provider:
    """A provider that returns a fixed decision, and records what it was asked."""
    id = "fake"

    def __init__(self, option, confidence=0.9):
        self.option, self.confidence, self.payloads = option, confidence, []

    def complete_structured(self, *, system, messages, response_schema, temperature=0.0,
                            max_tokens=2048):
        import json as _json
        self.payloads.append(_json.loads(messages[0]["content"]))
        return (response_schema(option=self.option, rationale="pledged deposits are a current "
                                "asset held on the same page as the section",
                                confidence=self.confidence),
                {"model": "fake-1", "input_tokens": 1, "output_tokens": 1})


def test_a_confirmed_routing_moves_the_line_into_others_and_the_subtotal_ties():
    from app.services.gap_closing import apply_routing, resolve_all

    rows = _gap_rows()
    provider = _Provider(option=1)
    routings = resolve_all(provider, rows, TEMPLATE)
    assert len(routings) == 1
    assert routings[0]["others_key"] == "bs_ca__others"
    assert routings[0]["labels"] == ["Pledged bank deposits"]
    assert routings[0]["model"] == "fake-1"
    # The model was shown the gap and the options, and NOT asked to do arithmetic.
    payload = provider.payloads[0]
    assert payload["difference_current"] == "40" and payload["difference_prior"] == "20"
    assert payload["would_be_placed_in"] == "bs_ca__others"

    assert apply_routing(rows, routings) == 1
    d = _stmt(rows)
    total = _row_of(d, "bs_ca__total")
    # 100 + 30 + 40 = 170, which is what the document printed. The gap is gone.
    assert total["v1"] == 170 and total["reported1"] == 170
    assert total["status"] != "recon"
    # The routed line is a listed contributor to Others, traceable to its page.
    others = _row_of(d, "bs_ca__others")
    assert others["v1"] == 40
    assert others["source"]["page_index"] == 8
    # And the reason it moved travels with the row.
    moved = rows[3]
    assert moved["mapping_method"] == "llm_gap_routing"
    assert moved["routed_to_others"]["target_key"] == "bs_ca__total"
    assert "current asset" in moved["routed_to_others"]["rationale"]


def test_the_model_declining_routes_nothing():
    from app.services.gap_closing import resolve_all

    rows = _gap_rows()
    assert resolve_all(_Provider(option=-1), rows, TEMPLATE) == []
    assert rows[3]["canonical_key"] is None


def test_an_option_outside_the_offered_set_is_refused():
    from app.services.gap_closing import resolve_all

    # A model naming option 99 has not chosen one of ours; nothing may be routed on that basis.
    assert resolve_all(_Provider(option=99), _gap_rows(), TEMPLATE) == []


def test_a_provider_that_fails_leaves_the_gap_for_review():
    from app.services.gap_closing import resolve_all

    class Broken:
        id = "broken"

        def complete_structured(self, **_kw):
            raise RuntimeError("no network")

    rows = _gap_rows()
    assert resolve_all(Broken(), rows, TEMPLATE) == []
    checks = _accounting_checks(rows, [], "en", [], TEMPLATE)
    assert any(c["type"] == "calculated_mismatch" for c in checks)


def test_one_leftover_cannot_be_spent_on_two_gaps():
    from app.services.gap_closing import resolve_all

    # Both sections are short by exactly 40, and only one unplaced line is worth 40.
    rows = [
        _row("bs_ca__inventories", "Inventories", 100, 90),
        _row("bs_ca__total", "Total current assets", 140, 110),
        _row("bs_cl__payables", "Payables", 50, 50),
        _row("bs_cl__total", "Total current liabilities", 90, 70),
        {"canonical_key": None, "source_label": "Sundry",
         "values": [_v("current", 40), _v("prior", 20)]},
    ]
    routings = resolve_all(_Provider(option=1), rows, TEMPLATE)
    assert len(routings) == 1
    assert sum(len(r["moved"]) for r in routings) == 1


def test_the_stage_is_a_no_op_without_a_provider(monkeypatch):
    """No provider means no routing — the gap stays a review item rather than being closed by an
    arithmetic coincidence nobody judged."""
    from app.config import get_settings
    from app.core.models.document import DocumentModel
    from app.core.stage import PipelineContext
    from app.stages.gap_closing import GapClosingStage

    monkeypatch.setenv("FINEX_LLM__PROVIDER", "stub")
    get_settings.cache_clear()
    try:
        doc = DocumentModel(filename="f.pdf")
        ctx = PipelineContext()
        ctx.template_def = TEMPLATE
        out = GapClosingStage().run(doc, ctx)
        assert out.gap_routings == []
    finally:
        get_settings.cache_clear()


# --------------------------------------------------------------------------------------
# An override says WHY
# --------------------------------------------------------------------------------------
def test_an_edit_carries_a_comment_that_travels_with_the_row_and_the_export(client):
    """A manual value overrides both the document and the arithmetic, so the reason has to be part
    of the record — beside the figure on screen, and in the workbook a reader opens later."""
    import io
    import uuid

    import openpyxl

    from app.db.base import SessionLocal, init_db
    from app.db.models import Document, ExtractionRun, TemplateVersion

    init_db()
    with SessionLocal() as s:
        doc = Document(filename="c.pdf", fmt="pdf", byte_size=1, page_count=1,
                       content_hash=uuid.uuid4().hex, object_key="k", owner="admin")
        s.add(doc)
        s.flush()
        tv = TemplateVersion(template_key=f"c-{uuid.uuid4().hex[:8]}", name="C", version=1,
                             definition=TEMPLATE)
        s.add(tv)
        s.flush()
        s.add(ExtractionRun(document_id=doc.id, status="succeeded",
                            options={"template_version_id": tv.id},
                            result={"rows": [_row("bs_ca__cash", "Cash", 30, 20)],
                                    "filename": "c.pdf"}))
        s.commit()
        doc_id = doc.id

    r = client.patch(f"/api/v1/documents/{doc_id}/line-items/bs_ca__cash",
                     json={"value": 44, "formula": "", "period": "current",
                           "basis": "consolidated",
                           "comment": "Agreed with management: reclassified from deposits"})
    assert r.status_code == 200, r.text
    assert "reclassified" in r.json()["comment"]

    d = client.get(f"/api/v1/documents/{doc_id}/statement",
                   params={"statement": "balance_sheet", "basis": "consolidated"}).json()
    row = next(x for x in d["rows"] if x["id"] == "bs_ca__cash")
    assert row["v1"] == 44
    assert "reclassified" in row["comments"]["current"]["text"]
    assert row["comments"]["current"]["by"] == "admin"          # who, not just what
    # The prior column was not edited, so it carries no note.
    assert "prior" not in (row["comments"] or {})

    x = client.get(f"/api/v1/documents/{doc_id}/export",
                   params={"fmt": "excel", "layout": "statement"})
    ws = openpyxl.load_workbook(io.BytesIO(x.content))["Balance Sheet"]
    cell = next(ws.cell(r_, 1) for r_ in range(1, ws.max_row + 1)
                if ws.cell(r_, 1).value == "Cash")
    assert "reclassified" in (cell.comment.text if cell.comment else "")

    # Reverting drops the note with the figure it explained.
    client.delete(f"/api/v1/documents/{doc_id}/line-items/bs_ca__cash")
    d2 = client.get(f"/api/v1/documents/{doc_id}/statement",
                    params={"statement": "balance_sheet", "basis": "consolidated"}).json()
    assert next(x for x in d2["rows"] if x["id"] == "bs_ca__cash")["comments"] is None


# --------------------------------------------------------------------------------------
# Defects an adversarial review confirmed. Each of these shipped, and each one broke a
# promise this batch had just made, so they are pinned individually.
# --------------------------------------------------------------------------------------
def test_the_printed_sum_is_never_published_as_a_re_evaluatable_formula():
    """The display rendering must not travel in `formula`.

    It did, and the consequence was the silent edit this whole batch existed to remove: the client
    prefilled its formula box with "100 + 50", sent it back with the next value edit, the server
    evaluated it, and `computed` took precedence over the typed figure — so typing 200 showed 150.
    """
    from app.services.formula import evaluate

    rows = [_row("bs_x__others", "Sundry A", 100, 90),
            _row("bs_x__others", "Sundry B", 50, 40)]
    row = _row_of(_stmt(rows), "bs_x__others")
    assert row["v1"] == 150
    assert row["arithmetic"] == "100 + 50"      # readable
    assert row["formula"] is None               # and NOT re-evaluatable
    # The rendering would evaluate cleanly if it were ever sent, which is exactly why it must not
    # be published in the field the client sends back.
    assert evaluate("100 + 50", lambda n: 0.0) == 150.0


def test_a_calculated_rows_arithmetic_is_display_only_too():
    rows = [_row("bs_ca__inventories", "Inventories", 100),
            _row("bs_ca__cash", "Cash", 30)]
    row = _row_of(_stmt(rows), "bs_ca__total")
    assert row["arithmetic"] == "100 + 30 + —"
    assert row["formula"] is None


def test_editing_one_period_leaves_the_other_periods_origin_alone():
    """Origin is decided per period. Deciding it at row level meant correcting this year flipped
    last year's total from computed back to the printed figure."""
    rows = [_row("bs_ca__inventories", "Inventories", 100, 90),
            _row("bs_ca__cash", "Cash", 30, 20),
            {**_row("bs_ca__total", "Total current assets", 555, 888),
             "edited": True, "edited_slots": ["consolidated/current"]}]
    row = _row_of(_stmt(rows), "bs_ca__total")
    assert (row["v1"], row["origin1"]) == (555, "manual")      # typed
    assert (row["v2"], row["origin2"]) == (110, "calculated")   # still its components, not 888


def test_a_period_that_cannot_be_computed_keeps_its_printed_figure():
    """Computability is per period. Entering the calculated branch because the OTHER period
    computed blanked this one out entirely."""
    rows = [_row("bs_ca__inventories", "Inventories", 100),         # current only
            _row("bs_ca__total", "Total current assets", 100, 888)]
    row = _row_of(_stmt(rows), "bs_ca__total")
    assert (row["v1"], row["origin1"]) == (100, "calculated")
    assert (row["v2"], row["origin2"]) == (888, "reported_uncomputed")


def test_a_manual_value_on_a_subtotal_reaches_the_total_above_it():
    """A rollup must stop preferring its own computation at a line the analyst answered for.

    It did not: `figure()` short-circuited to the nested computed value, so an override ON a
    subtotal was honoured for that row and ignored one row up — the spread then showed a total its
    own components contradicted.
    """
    rows = [_row("bs_ca__inventories", "Inventories", 100),
            _row("bs_ca__cash", "Cash", 30),
            {**_row("bs_ca__total", "Total current assets", 900),
             "edited": True, "edited_slots": ["consolidated/current"]},
            _row("bs_cl__payables", "Payables", 50)]
    d = _stmt(rows)
    assert _row_of(d, "bs_ca__total")["v1"] == 900
    # 900 − 50, not 130 − 50: the total is built from the figure shown beneath it.
    assert _row_of(d, "bs_net_current")["v1"] == 850


def test_two_printed_lines_on_one_off_template_concept_are_one_row():
    """Emitting a row each gave them the same id, which the client uses as its React key, its
    selection key and its edit address."""
    rows = [{"canonical_key": "commit_capital", "source_label": "Contracted for",
             "mapping_confidence": 0.8, "values": [_v("current", 900)]},
            {"canonical_key": "commit_capital", "source_label": "Authorised not contracted",
             "mapping_confidence": 0.7, "values": [_v("current", 100)]}]
    d = _build_statement(rows, TEMPLATE, "additional_items", "f.pdf")
    items = [r for r in d["rows"] if r["kind"] == "item"]
    assert len(items) == 1
    assert [r["id"] for r in items] == ["commit_capital"]
    assert items[0]["v1"] == 1000                                   # summed, as the face does
    assert [c["label"] for c in items[0]["contributions"]] == [
        "Contracted for", "Authorised not contracted"]
    assert len({r["id"] for r in d["rows"]}) == len(d["rows"])       # every id unique


def test_a_kpi_input_that_resolves_differently_per_period_still_reports_both():
    """The prior input was looked up by the key the CURRENT period resolved to, so a fallback
    keyspec that landed on a different key last year reported the prior input as absent."""
    rows = [
        # DIO: inventories / cost base, where the cost base has candidate keys tried in order.
        _row("bs_current_assets__inventories", "Inventories", 100, 80),
        _row("pl_expenses__cost_of_goods_sold", "Cost of goods sold", 500, None),
        _row("pl_expenses__total_operating_cost", "Total operating cost", None, 400),
    ]
    d = _build_statement(rows, None, "kpi", "f.pdf")
    row = _row_of(d, "kpi_dio")
    den = row["contributions"][-1]
    # Current resolved to cost of goods sold, prior to total operating cost — both reported.
    assert den["v1"] == 500 and den["v2"] == 400


def test_the_french_kpi_headings_are_in_french():
    rows = [_row("bs_current_assets__total_current_assets", "Total current assets", 300, 200),
            _row("bs_current_liabilities__total_current_liabilities",
                 "Total current liabilities", 150, 200)]
    d = _build_statement(rows, None, "kpi", "f.pdf", locale="fr")
    headings = [r["label"] for r in d["rows"] if r["kind"] == "section"]
    assert "Rentabilité" in headings
    assert "الربحية" not in headings          # the Arabic string had been copied into `fr`
