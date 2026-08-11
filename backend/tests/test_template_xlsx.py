"""The template as a workbook: it round-trips, it states which lines are calculated, and it
refuses an edit it would otherwise have to guess about.

A template decides what every extraction maps to and what every structural check recomputes, so
the reader is deliberately strict — a row it is unsure about has to fail loudly here rather than
surface months later as a figure on the wrong line.
"""
from __future__ import annotations

import io
import json
import pathlib

import pytest

pytest.importorskip("openpyxl")

from app.services.template_xlsx import (
    KIND_CALCULATED, KIND_EXTRACTED, KIND_HEADING, TemplateSheetError, build_template_xlsx,
    parse_template_xlsx)

_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _shipped() -> dict:
    p = (pathlib.Path(__file__).resolve().parents[1]
         / "app/sample/templates/hkfrs_hk_china_template.json")
    return json.loads(p.read_text())


def _sheet(data: bytes, name="Template"):
    import openpyxl

    return openpyxl.load_workbook(io.BytesIO(data))[name]


def _rows(ws) -> list[dict]:
    head = [str(c.value or "") for c in ws[1]]
    return [dict(zip(head, [c.value for c in row])) for row in ws.iter_rows(min_row=2)]


def test_the_workbook_marks_every_line_extracted_or_calculated():
    rows = _rows(_sheet(build_template_xlsx(_shipped())))
    kinds = {r["Kind"] for r in rows}
    assert kinds == {KIND_EXTRACTED, KIND_CALCULATED, KIND_HEADING}
    # A subtotal with a rollup is calculated, and says what from — that IS the check that runs.
    subtotal = next(r for r in rows
                    if r["Canonical key"] == "bs_non_current_assets__total_non_current_assets")
    assert subtotal["Kind"] == KIND_CALCULATED and subtotal["Calculation"] == "sum"
    assert "bs_non_current_assets__property_plant_and_equipment" in subtotal["Calculated from"]
    # A plain line is extracted and carries no components.
    line = next(r for r in rows
                if r["Canonical key"] == "bs_non_current_assets__property_plant_and_equipment")
    assert line["Kind"] == KIND_EXTRACTED and not line["Calculated from"]


def test_the_workbook_round_trips_the_shipped_template_without_loss():
    original = _shipped()
    back = parse_template_xlsx(build_template_xlsx(original),
                              template_key=original["template_key"], name=original["name"])

    def flat(d):
        out = []
        for st in d["statements"]:
            for sec in st["sections"]:
                out.append((st["type"], sec["canonical_key"], sec.get("role"),
                            (sec.get("rollup") or {}).get("op"),
                            tuple((sec.get("rollup") or {}).get("children") or ())))
                for c in sec.get("children", []):
                    out.append((st["type"], c["canonical_key"], c.get("role"),
                                (c.get("rollup") or {}).get("op"),
                                tuple((c.get("rollup") or {}).get("children") or ())))
        return out

    assert flat(back) == flat(original)
    # Localized labels survive the trip — output parity is a template property.
    ppe = next(c for st in back["statements"] for sec in st["sections"]
               for c in sec["children"]
               if c["canonical_key"] == "bs_non_current_assets__property_plant_and_equipment")
    assert ppe["label_i18n"]["zh"] == "物业、厂房及设备"
    # …and so do the statement identities.
    bs = next(st for st in back["statements"] if st["type"] == "balance_sheet")
    assert any(i["lhs"] == "bs_total_assets" for i in bs["identities"])


def _mini_book(rows: list[list], header=None) -> bytes:
    import openpyxl

    from app.services.template_xlsx import COLUMNS

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"
    ws.append(header or [h for _k, h in COLUMNS])
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _row(statement="Balance sheet", section="", node="", key="", label="", role="line",
         kind=KIND_EXTRACTED, op="", children="", sign="natural"):
    return [statement, section, node, key, label, "", "", "", role, kind, op, children, sign,
            "", ""]


def test_a_line_marked_calculated_must_say_what_it_is_calculated_from():
    book = _mini_book([
        _row(section="", node="sec", key="sec", label="Current assets", role="header",
             kind=KIND_HEADING),
        _row(section="sec", key="a", label="Cash"),
        _row(section="sec", key="tot", label="Total", role="subtotal", kind=KIND_CALCULATED),
    ])
    with pytest.raises(TemplateSheetError, match="Calculated from"):
        parse_template_xlsx(book, template_key="t", name="T")


def test_a_calculated_line_cannot_reference_a_key_that_is_not_in_the_template():
    book = _mini_book([
        _row(node="sec", key="sec", label="Current assets", role="header", kind=KIND_HEADING),
        _row(section="sec", key="a", label="Cash"),
        _row(section="sec", key="tot", label="Total", role="subtotal", kind=KIND_CALCULATED,
             op="sum", children="a\nnot_a_key"),
    ])
    with pytest.raises(TemplateSheetError, match="not_a_key"):
        parse_template_xlsx(book, template_key="t", name="T")


def test_an_extracted_line_with_components_is_a_contradiction_and_is_refused():
    book = _mini_book([
        _row(node="sec", key="sec", label="Current assets", role="header", kind=KIND_HEADING),
        _row(section="sec", key="a", label="Cash"),
        _row(section="sec", key="b", label="Total", children="a"),
    ])
    with pytest.raises(TemplateSheetError, match="Calculated from"):
        parse_template_xlsx(book, template_key="t", name="T")


def test_duplicate_canonical_keys_are_refused():
    book = _mini_book([
        _row(node="sec", key="sec", label="Current assets", role="header", kind=KIND_HEADING),
        _row(section="sec", key="a", label="Cash"),
        _row(section="sec", key="a", label="Cash again"),
    ])
    with pytest.raises(TemplateSheetError, match="duplicate canonical key"):
        parse_template_xlsx(book, template_key="t", name="T")


def test_a_line_before_its_section_heading_is_refused_rather_than_reparented():
    book = _mini_book([_row(section="sec", key="a", label="Cash")])
    with pytest.raises(TemplateSheetError, match="has no heading row"):
        parse_template_xlsx(book, template_key="t", name="T")


def test_an_unknown_statement_name_is_refused_not_invented():
    book = _mini_book([_row(statement="Sundry schedule", node="s", key="s", label="X",
                            role="header", kind=KIND_HEADING)])
    with pytest.raises(TemplateSheetError, match="unknown Statement"):
        parse_template_xlsx(book, template_key="t", name="T")


def test_reordered_and_extra_columns_do_not_shift_the_values():
    # People move columns and add their own notes column. Matching on header text means an edited
    # workbook still reads correctly instead of filing every value one column to the left.
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"
    ws.append(["My notes", "Kind", "Canonical key", "Label (en)", "Statement", "Section", "Role"])
    ws.append(["check this", KIND_HEADING, "sec", "Current assets", "Balance sheet", "", "header"])
    ws.append(["", KIND_EXTRACTED, "a", "Cash", "Balance sheet", "sec", "line"])
    buf = io.BytesIO()
    wb.save(buf)
    d = parse_template_xlsx(buf.getvalue(), template_key="t", name="T")
    sec = d["statements"][0]["sections"][0]
    assert sec["canonical_key"] == "sec"
    assert [c["canonical_key"] for c in sec["children"]] == ["a"]


def test_a_workbook_that_is_not_a_template_says_so():
    import openpyxl

    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(TemplateSheetError, match="no 'Template' sheet"):
        parse_template_xlsx(buf.getvalue(), template_key="t", name="T")


# ---------------------------------------------------------------------------------------
# Over HTTP: download, edit, upload → a NEW version, and role-gated
# ---------------------------------------------------------------------------------------
def _hk_template(client):
    tpls = client.get("/api/v1/templates").json()
    return next(t for t in tpls if t["template_key"] == "hkfrs_hk_china_v1")


def test_an_admin_downloads_the_template_edits_it_and_uploads_a_new_version(client):
    tpl = _hk_template(client)
    r = client.get(f"/api/v1/templates/{tpl['id']}/xlsx")
    assert r.status_code == 200 and r.headers["content-type"] == _XLSX
    assert "attachment" in r.headers["content-disposition"]

    # Turn one extracted line into a calculated one — the edit the user asked to be able to make.
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    ws = wb["Template"]
    head = [str(c.value or "") for c in ws[1]]
    ki, ci, oi, chi = (head.index(h) + 1 for h in
                       ("Kind", "Canonical key", "Calculation", "Calculated from"))
    target = next(n for n in range(2, ws.max_row + 1)
                  if ws.cell(n, ci).value == "bs_current_assets__total_current_assets")
    ws.cell(target, ki).value = KIND_CALCULATED
    ws.cell(target, oi).value = "sum"
    ws.cell(target, chi).value = "bs_current_assets__inventories"
    out = io.BytesIO()
    wb.save(out)

    up = client.post("/api/v1/templates/xlsx",
                     files={"file": ("edited.xlsx", out.getvalue(), _XLSX)},
                     data={"template_key": tpl["template_key"], "name": tpl["name"]})
    assert up.status_code == 201, up.text
    body = up.json()
    assert body["template_key"] == tpl["template_key"]
    assert body["version"] > tpl["version"]          # a new VERSION, nothing overwritten
    assert body["line_items"] > 100

    fresh = client.get(f"/api/v1/templates/{body['id']}").json()["definition"]
    node = next(c for st in fresh["statements"] for sec in st["sections"]
                for c in sec.get("children", [])
                if c["canonical_key"] == "bs_current_assets__total_current_assets")
    assert node["rollup"] == {"op": "sum", "children": ["bs_current_assets__inventories"]}
    # The version we downloaded is still there, unchanged.
    old = client.get(f"/api/v1/templates/{tpl['id']}").json()["definition"]
    old_node = next(c for st in old["statements"] for sec in st["sections"]
                    for c in sec.get("children", [])
                    if c["canonical_key"] == "bs_current_assets__total_current_assets")
    assert old_node.get("rollup", {}).get("children") != ["bs_current_assets__inventories"]


def test_a_bad_workbook_upload_explains_which_row_is_wrong(client):
    book = _mini_book([
        _row(node="sec", key="sec", label="Current assets", role="header", kind=KIND_HEADING),
        _row(section="sec", key="tot", label="Total", role="subtotal", kind=KIND_CALCULATED),
    ])
    r = client.post("/api/v1/templates/xlsx",
                    files={"file": ("bad.xlsx", book, _XLSX)},
                    data={"template_key": "", "name": "Bad"})
    assert r.status_code == 422
    assert "Row 3" in r.json()["detail"]


def test_an_analyst_cannot_publish_a_template_from_a_workbook(client, auth):
    book = _mini_book([_row(node="sec", key="sec", label="S", role="header", kind=KIND_HEADING)])
    r = client.post("/api/v1/templates/xlsx",
                    files={"file": ("t.xlsx", book, _XLSX)}, data={"name": "T"},
                    headers=auth("analyst"))
    assert r.status_code == 403


def test_the_upload_screen_can_read_the_column_contract(client):
    body = client.get("/api/v1/templates/xlsx/columns").json()
    assert [c["header"] for c in body["columns"]][:4] == [
        "Statement", "Section", "Node ID", "Canonical key"]
    assert {k["value"] for k in body["kinds"]} == {KIND_EXTRACTED, KIND_CALCULATED, KIND_HEADING}
