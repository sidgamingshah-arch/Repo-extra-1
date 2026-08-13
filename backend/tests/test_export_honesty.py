"""Export honesty (Req 8, 10): JSON carries a derived-analysis block and edited formulas,
the flat sheet has a Formula column, edited items carry a formula note in the statement
workbook, and the Include set gates the analysis sheets."""
from __future__ import annotations

import io
import time

import pytest

pytest.importorskip("fitz")

from tests.fixtures.generate import make_native_pdf, make_rich_pdf


def _await(client, doc_id):
    for _ in range(100):
        r = client.get(f"/api/v1/documents/{doc_id}/run")
        if r.status_code == 200 and r.json().get("status") == "succeeded":
            return
        time.sleep(0.05)
    raise AssertionError("did not finish")


def _extract(client, data=None, filename="bs.pdf"):
    doc_id = client.post("/api/v1/documents",
                         files={"file": (filename, data or make_native_pdf(), "application/pdf")}).json()["id"]
    onts = client.get("/api/v1/ontologies").json()
    ont = next((o for o in onts if o["ontology_key"] == "hkfrs_hk_china"), onts[0])
    tpls = client.get("/api/v1/templates").json()
    tpl = next((t for t in tpls if t["template_key"] == ont["target_template_key"]), tpls[0])
    client.post(f"/api/v1/documents/{doc_id}/extractions",
                json={"ontology_version_id": ont["id"], "template_version_id": tpl["id"]})
    _await(client, doc_id)
    return doc_id


def test_json_export_carries_analysis_and_formula(client):
    doc_id = _extract(client, data=make_rich_pdf(), filename="rich.pdf")
    # Edit a line to attach a formula.
    rows = client.get(f"/api/v1/documents/{doc_id}/run").json()["result"]["rows"]
    key = next(r["canonical_key"] for r in rows if r.get("canonical_key"))
    client.patch(f"/api/v1/documents/{doc_id}/line-items/{key}", json={"value": 123, "formula": "=A1+A2"})

    body = client.get(f"/api/v1/documents/{doc_id}/export", params={"fmt": "json"}).json()
    assert "analysis" in body and body["analysis"]["ratios"]
    assert any(d["key"] == "contingent_liabilities" for d in body["analysis"]["disclosures"])
    edited = next(li for li in body["line_items"] if li["canonical_key"] == key)
    assert edited["edited"] and edited["formula"] == "=A1+A2"


def test_flat_export_has_formula_column(client):
    import openpyxl

    doc_id = _extract(client)
    rows = client.get(f"/api/v1/documents/{doc_id}/run").json()["result"]["rows"]
    key = next(r["canonical_key"] for r in rows if r.get("canonical_key"))
    client.patch(f"/api/v1/documents/{doc_id}/line-items/{key}", json={"value": 5, "formula": "=SUM(x)"})

    x = client.get(f"/api/v1/documents/{doc_id}/export", params={"fmt": "excel", "layout": "flat"})
    wb = openpyxl.load_workbook(io.BytesIO(x.content))
    ws = wb["Extraction"]
    header = [c.value for c in ws[1]]
    assert "Formula" in header
    text = " | ".join(str(v) for row in ws.iter_rows(values_only=True) for v in row if v)
    assert "=SUM(x)" in text
    # …and it is TEXT, not a live formula cell. openpyxl promotes a leading "=" to a real formula,
    # and this expression's references are canonical line-item keys, not cell addresses, so the
    # workbook opened with #NAME? where an audit trail was intended.
    cell = next(c for row in ws.iter_rows() for c in row if c.value == "=SUM(x)")
    assert cell.data_type == "s"


def test_workbook_carries_formulas_as_notes_not_live_cells(client):
    """The Excel export does not build a live spreadsheet, and the docs now say so.

    A formula travels for AUDIT: as text in the flat sheet's Formula column, and as a cell NOTE on
    the row's label cell in the statement workbook. References are canonical line-item keys
    resolved server-side by services/formula.py and are never translated to cell addresses, so
    nothing in the file recalculates — the number in the cell is the value the server computed.
    """
    import openpyxl

    doc_id = _extract(client)
    rows = client.get(f"/api/v1/documents/{doc_id}/run").json()["result"]["rows"]
    key = next(r["canonical_key"] for r in rows if r.get("canonical_key"))
    client.patch(f"/api/v1/documents/{doc_id}/line-items/{key}",
                 json={"value": 5, "formula": "=SUM(x)"})

    x = client.get(f"/api/v1/documents/{doc_id}/export",
                   params={"fmt": "excel", "layout": "statement"})
    wb = openpyxl.load_workbook(io.BytesIO(x.content))
    assert not [c for ws in wb.worksheets for row in ws.iter_rows() for c in row
                if c.data_type == "f"]
    notes = [c.comment.text for ws in wb.worksheets for row in ws.iter_rows() for c in row
             if c.comment]
    assert any("=SUM(x)" in note for note in notes)


def test_include_gates_analysis_sheets(client):
    import openpyxl

    doc_id = _extract(client, data=make_rich_pdf(), filename="rich.pdf")
    x = client.get(f"/api/v1/documents/{doc_id}/export",
                   params={"fmt": "excel", "layout": "statement", "include": "ratios"})
    wb = openpyxl.load_workbook(io.BytesIO(x.content))
    assert "Ratios" in wb.sheetnames
    assert "Disclosures" not in wb.sheetnames and "Note details" not in wb.sheetnames


# --- the workbook says what its figures were verified against -----------------

def _banner(coverage):
    from app.services.export import validation_caption
    return validation_caption(coverage, "en")


def test_a_fully_validated_run_gets_no_warning():
    """The banner has to be able to stay silent, or it is wallpaper nobody reads."""
    assert _banner({"available": True, "alarms": [],
                    "aggregate": {"status": "PASSED", "passed": 8, "failed": 0,
                                  "evaluated": 8, "declarable": 8}}) is None


def test_a_run_where_nothing_could_be_evaluated_says_so_severely():
    """The defect this closes: a filing whose relations could not be evaluated exported as a
    workbook indistinguishable from a validated one — same figures, same formatting, nothing on the
    sheet saying the arithmetic behind them was never checked."""
    text, severe = _banner({"available": True, "alarms": [],
                            "aggregate": {"status": "UNVALIDATED", "passed": 0, "failed": 0,
                                          "evaluated": 0, "declarable": 6}})
    assert severe is True and "nothing here is validated" in text


def test_an_unenforceable_blocking_rule_is_named_even_when_the_rest_passed():
    """A blocking rule that could not be enforced is not a partial result — it is a rule the filing
    was never actually held to, and PASSED beside it would read as a clean bill of health."""
    text, severe = _banner({
        "available": True,
        "alarms": [{"code": "BLOCKING_RULE_UNENFORCEABLE", "rule_id": "guard:sign_expectation:x"}],
        "aggregate": {"status": "PASSED", "passed": 8, "failed": 0,
                      "evaluated": 8, "declarable": 8}})
    assert severe is True and "blocking rule could not be enforced" in text


def test_a_missing_coverage_report_is_not_treated_as_a_clean_one():
    """An absent report is not a passing one. Silence here would be the same defect in a new place."""
    text, severe = _banner(None)
    assert severe is True and "unknown" in text.lower()
    text, severe = _banner({"available": False, "reason": "no_template",
                            "reason_label": "No template was attached to this run."})
    assert severe is True and "No template was attached" in text


def test_the_banner_is_translated():
    from app.services.export import validation_caption
    unvalidated = {"available": True, "alarms": [],
                   "aggregate": {"status": "UNVALIDATED", "passed": 0, "failed": 0,
                                 "evaluated": 0, "declarable": 6}}
    for loc in ("zh", "ar", "fr"):
        text, _ = validation_caption(unvalidated, loc)
        assert text and not text.isascii(), f"{loc} banner fell back to English: {text!r}"


def test_the_warning_reaches_the_statement_sheet_above_the_figures():
    """Not a sheet of its own: a reader who opens the balance sheet and scrolls has to pass it."""
    import io

    import openpyxl

    from app.services.export import build_statement_workbook

    template = {"schema_version": 1, "template_key": "t", "name": "T",
                "statements": [{"type": "balance_sheet", "sections": [
                    {"node_id": "s1", "label": "Assets", "children": [
                        {"canonical_key": "bs_ca__cash", "label": "Cash", "role": "line"}]}]}]}
    rows = [{"canonical_key": "bs_ca__cash", "source_label": "Cash",
             "values": [{"basis": "consolidated", "period_label": "current", "value": "10"}]}]
    unvalidated = {"available": True, "alarms": [],
                   "aggregate": {"status": "UNVALIDATED", "passed": 0, "failed": 0,
                                 "evaluated": 0, "declarable": 6}}

    def sheet(coverage):
        wb = openpyxl.load_workbook(io.BytesIO(build_statement_workbook(
            rows, template, filename="f.pdf", coverage=coverage)))
        return wb[wb.sheetnames[0]]

    def col_a(ws, upto=10):
        return [str(ws.cell(r, 1).value or "") for r in range(1, upto)]

    warned = sheet(unvalidated)
    a = col_a(warned)
    banner = [i for i, v in enumerate(a) if v.startswith("⚠")]
    assert banner, f"no validation banner on the statement sheet: {a}"
    # Above the figures: the banner precedes the column header band, so it cannot be scrolled past
    # without being seen, and the band shifted down to make room rather than being overwritten.
    header = a.index("Line item")
    assert banner[0] < header, f"banner below the header band: {a}"
    assert a.index("Assets") == header + 1, f"section rows did not follow the band: {a}"

    # …and a clean run is not given a banner at all, so the band sits one row higher.
    clean = sheet({"available": True, "alarms": [],
                   "aggregate": {"status": "PASSED", "passed": 6, "failed": 0,
                                 "evaluated": 6, "declarable": 6}})
    b = col_a(clean)
    assert not [v for v in b if v.startswith("⚠")], b
    assert b.index("Line item") == header - 1, f"clean sheet layout shifted: {b}"
