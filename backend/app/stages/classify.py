"""Page-classification stage (face / notes / other).

Two properties matter for real annual reports, and the second is why this is a DECODE rather than a
scan. A filing runs narrative → the statement faces → the notes → back-matter, so a page's kind is
mostly settled by where it sits; but the local evidence on a page is noisy in both directions. Note
pages quote face phrases in prose ("12. Financial assets at fair value through profit or loss"), the
auditor's report lists every statement it audited in bold, and the contents page prints all of their
titles with leader dots. Judging each page alone therefore over-produces faces, and judging by
position alone cannot recover from one bad guess.

So: per-page evidence becomes an EMISSION score, document order becomes TRANSITION costs, and a
Viterbi decode picks the sequence that best explains the whole filing. One page's strong title can no
longer flip the rest of the document, and — unlike the fixed region walk this replaces — the notes
region is no longer one-way. That mattered: HK filings routinely print the Company-only statement of
financial position AFTER the consolidated notes, past note 40, and a one-way walk classified it as a
note forever. `NOTES → FACE` is now merely expensive, so a genuine face title there can win.

The lexicon carries two tiers per statement. STRONG patterns are self-anchoring — the phrase alone
identifies a statement face. WEAK patterns are generic and need an English structural anchor
elsewhere on the line, because "profit or loss" on its own is also a note-heading fragment. Chinese
patterns are all STRONG: they are multi-character and specific, so they need no anchor, and they must
NOT be end-anchored — a continuation page titled 綜合權益變動表(續) and a bilingual one-line title
both fail an end-of-line anchor.
"""
from __future__ import annotations

import io
import re
from dataclasses import dataclass, field

from app.core.models import DocumentModel, PageKind
from app.core.models.enums import DocFormat
from app.core.stage import PipelineContext

# ---------------------------------------------------------------- lexicons ---
#
# Per-character classes ([資资]) collapse Traditional, Simplified and the mixed-script typesetting HK
# filings sometimes produce into one pattern each.

_EN_ANCHOR = re.compile(
    r"statements?\s+of|balance\s+sheets?|income\s+statements?"
    r"|profit\s+and\s+loss\s+accounts?", re.I)

# (name, STRONG patterns, WEAK patterns)
_STATEMENTS: list[tuple[str, list[str], list[str]]] = [
    ("changes_in_equity", [
        r"statements?\s+of\s+changes\s+in\s+[a-z'’\s]{0,28}equity",
        r"statements?\s+of\s+(?:shareholders|owners|stockholders|equity\s+holders)"
        r"'?’?\s*equity",
        r"statements?\s+of\s+changes\s+in\s+net\s+assets",
        r"[權权]益[變变][動动][報报]?表",
    ], [
        r"changes\s+in\s+[a-z'’\s]{0,28}equity",
    ]),

    ("cash_flow", [
        r"statements?\s+of\s+cash\s?flows?",
        r"cash\s?flows?\s+statements?",
        r"[現现]金流[量動动][報报]?表",
    ], [
        r"cash\s?flows?",
    ]),

    ("balance_sheet", [
        r"statements?\s+of\s+financial\s+(?:position|condition)",
        r"balance\s+sheets?",
        r"statements?\s+of\s+assets\s+and\s+liabilities",
        r"[資资][產产][負负][債债][報报]?表",
        r"[財财][務务][狀状][況况][報报]?表",
    ], [
        r"financial\s+position",
    ]),

    # Not preferred by list order — by MATCH LENGTH at the topmost y (see _resolve_statement), so
    # "…OF PROFIT OR LOSS AND OTHER COMPREHENSIVE INCOME" beats a bare comprehensive-income hit.
    ("comprehensive_income", [
        r"statements?\s+of\s+(?:other\s+)?comprehensive\s+(?:income|loss|expenses?)",
        r"(?:其他)?全面(?:收益|收入|[虧亏][損损])[報报]?表",
        r"其他[綜综]合收益[報报]?表",
    ], [
        r"(?:other\s+)?comprehensive\s+(?:income|loss)",
    ]),

    ("profit_and_loss", [
        r"statements?\s+of\s+profit\s+(?:or|and|&)\s+loss(?:es)?",
        # Ch.18A biotech and other pre-revenue issuers title this "STATEMENTS OF LOSS"; the old
        # lexicon required the word "profit" and so resolved those filings to nothing.
        r"statements?\s+of\s+(?:loss|profit|income|earnings|operations)(?:es|s)?\b",
        r"income\s+statements?",
        r"profit\s+and\s+loss\s+accounts?",                 # legacy HK GAAP
        r"[損损]益[表報报賬帳账]",
        r"[利][潤润][報报]?表",
        r"[虧亏][損损][報报]?表",
        r"收益[報报]?表",
        r"[經经][營营][業业][績绩][報报]?表",
    ], [
        r"profit\s+(?:or|and|&)\s+loss(?:es)?",
    ]),
]

# 綜合收益表 vs 综合收益表: Traditional HK usage is 綜合 = consolidated and 全面收益 = comprehensive
# income, so 綜合收益表 is the income statement. PRC Simplified usage is 合并 = consolidated and
# 综合收益 = comprehensive income. Both land on profit_and_loss here because the bare 收益表 pattern
# is SHORTER than 全面收益表, so a genuine comprehensive-income page still wins on match length. A
# Simplified CAS filing with a separate 综合收益表 page is therefore tagged profit_and_loss, and is
# flagged in page.evidence["title_ambig"] rather than silently decided.
_ZH_CI_AMBIG = re.compile(r"[綜综]合收益[報报]?表")

# Single-statement presentation: one page carrying P&L and OCI together, including the all-loss form.
_OCI_COMBINED = re.compile(
    r"\b(?:profit\s+(?:or|and|&)\s+loss(?:es)?|loss(?:es)?|income|operations)\b"
    r"[^\n]{0,24}\band\s+other\s+comprehensive\s+(?:income|loss)\b"
    r"|[損损]益及其他(?:全面|[綜综]合)(?:收益|[虧亏][損损])"
    r"|[虧亏][損损]及其他(?:全面|[綜综]合)(?:收益|[虧亏][損损])"
    r"|收益及其他全面收益", re.I)

# A title candidate is disqualified outright by any of these. Positive matching alone cannot separate
# a face title from a note heading, a contents line or auditor prose quoting the same words — and the
# contents page is the expensive case, because its titles used to start the face region early and
# hand pages of front matter to the extractor as statements.
_TITLE_NEGATIVE = re.compile(
    r"notes?\s+to\s+the\b|notes?\s+to\s+(?:consolidated|financial)\b"
    r"|notes\s+forming\s+part\s+of"
    r"|附\s*[註注]"
    r"|\.{3,}|·{3,}|…"                                                   # contents leader dots
    r"|\bpages?\s+\d+|\bset\s+out\s+(?:on|in)\b|\brefer(?:\s+to)?\b"
    r"|\bin\s+our\s+opinion\b|\bwe\s+have\s+audited\b|\bas\s+described\s+in\b"
    r"|\breconciliation\s+of\b|\bextract(?:ed|s)?\s+from\b|\bsummar(?:y|ised)\b"
    r"|\bpro\s+forma\b|[備备]考"
    r"|\bunaudited\s+supplementary\b"
    r"|[載载][於于]|[見见]\s*[附第]|[第]\s*\d+\s*[頁页]"
    r"|^\s*(?:contents|index|目[錄录])\s*$", re.I)

# Statement-ish but unresolved. Logged onto the document so lexicon coverage is measurable instead of
# guessed: a filing whose titles are all recognised and one whose titles are all missed both look
# like silence otherwise.
_TITLE_HINT = re.compile(
    r"statements?\b|balance\s+sheet|account\b|[報报]?表\s*[（(]?\s*$|[報报]表", re.I)

_NOTES_BANNER = [
    r"notes?\s+to\s+(?:the\s+)?(?:consolidated\s+|unconsolidated\s+|standalone\s+)?"
    r"financial\s+statements",
    r"notes?\s+to\s+(?:the\s+)?accounts",            # legacy HK GAAP
    r"notes\s+forming\s+part\s+of\s+the",
    r"(?:material|significant)\s+accounting\s+polic",
    r"[財财][務务][報报]?表?附\s*[註注]",
    r"[合併合并綜综]+[財财][務务][報报]表附\s*[註注]",
    r"(?:重要|主要)[會会][計计]政策",
]

_BACKMATTER = re.compile(
    r"five[\s-]?year\s+(?:financial\s+)?summary|financial\s+(?:summary|highlights)"
    r"|[五][年][財财][務务][摘概][要要]|[財财][務务][摘概][要要]", re.I)

# Scope. Consolidated is tested FIRST, because "…of the Company and its subsidiaries" contains the
# word Company and must not be read as the company-only statement.
_SCOPE_CONSOL = re.compile(r"\bconsolidated\b|\bgroup\b|[合][併并]", re.I)
_SCOPE_COMPANY = re.compile(
    r"\bcompany\b|\bthe\s+bank\b|\bparent\b|\bstandalone\b|\bunconsolidated\b"
    r"|[母][公][司]|[本][公][司]", re.I)
# Traditional 綜合 means BOTH "consolidated" and "comprehensive"; a scope marker only when it is not
# immediately preceding a comprehensive-income token.
_ZH_CONSOL_AMBIG = re.compile(r"[綜综]合(?!收益|[損损]益|全面|[虧亏][損损])")

# Prose pages that discuss the statements without being one.
_NARRATIVE = re.compile(
    r"\bin\s+our\s+opinion\b|\bwe\s+have\s+audited\b|\bindependent\s+auditor"
    r"|\bdirectors'?\s+report\b|\bkey\s+audit\s+matters?\b|\bbasis\s+for\s+opinion\b"
    r"|[核][數数][師师][報报][告告]|[董][事][會会][報报][告告]|[獨独][立][核][數数][師师]", re.I)

# The notes section opens at note 1 even when a filing omits the banner.
_NOTE_ONE = re.compile(r"(?m)^\s*(?:note\s*)?1[.)、]?\s+[A-Za-z一-鿿]{2,}")
# A numbered note heading, e.g. "14. Cash and cash equivalents" / "14 现金及现金等价物".
_NUMBERED_HEADING = re.compile(r"(?m)^\s*(?:note\s*)?\d{1,2}[.)、]?\s+[A-Za-z一-鿿]{2,}")
# A page's repeating running header, skipped so a title is read from the page's own heading.
_RUNNING_HEADER = re.compile(r"annual report|interim report|年報|年度報告|中期報告", re.I)
_NUM_TOKEN = re.compile(r"\(?-?[\d,]+\.?\d*\)?")
# A split title runs to two or three short lines; beyond that a "run" of heading-shaped lines is
# the title plus the top of the table.
_JOIN_MAX_LINES = 3
_JOIN_MAX_CHARS = 90

# A statement face never titles three different statements; a contents page does.
_MAX_DISTINCT_TITLES = 3
# Which of the resolved names the rest of the pipeline understands. `comprehensive_income` is folded
# into profit_and_loss for page.statement because ontology mapping scopes candidate concepts by this
# value and no template declares a separate OCI statement — a page tagged with a statement the
# template does not carry would reject every concept on it. The finer label is kept in evidence.
_STATEMENT_ALIAS = {"comprehensive_income": "profit_and_loss"}


@dataclass
class PageFeat:
    """One page's local evidence, before the decode weighs it against its neighbours."""

    index: int = 0
    title_lines: list[str] = field(default_factory=list)
    statement: str | None = None
    oci_combined: bool = False
    title_ambig: bool = False
    matched_title: str | None = None
    unmapped: list[str] = field(default_factory=list)
    strong_title: bool = False
    narrative: bool = False
    notes_banner: bool = False
    note_heading: bool = False
    note_one: bool = False
    backmatter: bool = False
    numeric_density: float = 0.0
    scope: str | None = None
    scope_columns: list[str] = field(default_factory=list)


def _looks_like_heading(line: str) -> bool:
    """A statement title is a short heading line, not a sentence. Auditor prose such as "We audited
    the statement of profit or loss …" mentions face phrases but is long and ends like a sentence."""
    s = line.strip()
    if not s or len(s) > 110:
        return False
    if s.endswith((".", ";", ":", ",", "。", "，", "、", "；", "：")):
        return False
    # At most ONE figure. A title may carry a year; a data row carries a caption and its amounts
    # ("Revenue   Note 5   45,230"), and those rows were being joined onto the title above them —
    # the classification stayed right but `matched_title` came out as the title plus half the table,
    # which is the evidence a reader checks and the text scope resolution reads.
    if len(_NUM_TOKEN.findall(s)) > 1:
        return False
    return sum(ch.isdigit() for ch in s) <= 8


def _title_zone(lines: list[dict], limit: int = 8) -> list[dict]:
    """The page's heading lines: the first few non-empty ones, minus the leading page number and the
    repeating running header, which sit above a real title."""
    out: list[dict] = []
    for line in lines:
        s = line["text"].strip()
        if not s or re.fullmatch(r"\d{1,4}", s) or _RUNNING_HEADER.search(s):
            continue
        out.append(line)
        if len(out) >= limit:
            break
    return out


def _title_candidates(lines: list[dict]) -> list[dict]:
    """Heading-like lines, plus joins of consecutive heading-like lines so a title split across two
    lines ("CONSOLIDATED STATEMENT OF" / "CASH FLOWS") is matched as one."""
    cands = [dict(c) for c in lines if _looks_like_heading(c["text"])]

    def joined(run: list[dict]) -> None:
        """Join a run, but only as far as a TITLE plausibly runs.

        A genuinely split title is short ("CONSOLIDATED STATEMENT OF" / "CASH FLOWS"). Joining a
        whole run unbounded swept the first data rows onto the end of the title, because a row
        carrying one figure is heading-shaped by every other measure — so the cap is on the join,
        not on the line.
        """
        if len(run) < 2:
            return
        text = run[0]["text"]
        for nxt in run[1:_JOIN_MAX_LINES]:
            candidate = f"{text} {nxt['text']}"
            if len(candidate) > _JOIN_MAX_CHARS:
                break
            text = candidate
        if text != run[0]["text"]:
            cands.append({**run[0], "text": text})

    run: list[dict] = []
    for line in lines:
        if _looks_like_heading(line["text"]):
            run.append(line)
        else:
            joined(run)
            run = []
    joined(run)
    return cands


def _anchored(t: str) -> bool:
    """English structural anchor. Chinese patterns are self-anchoring — deliberately NO end-of-line
    anchor, which rejected every continuation page titled 綜合權益變動表(續) and every bilingual
    one-line title."""
    return bool(_EN_ANCHOR.search(t))


def _resolve_statement(cands: list[dict]) -> tuple[str | None, bool, str | None, bool]:
    """(statement, oci_combined, matched_title, ambiguous).

    POSITION first, then match length — never list order. Pages genuinely carry two candidates (an
    equity-statement tail above a cash-flow title; P&L above OCI), and longest-match-at-topmost-y is
    what picks the right one. Reverting to list order reintroduces the equity-tail bug.
    """
    hits: list[tuple[float, int, str, str]] = []
    names: set[str] = set()
    for c in cands:
        t = c["text"]
        if _TITLE_NEGATIVE.search(t):        # note heading / contents line / auditor prose
            continue
        low = t.lower()
        best: tuple[int, str] | None = None
        for name, strong, weak in _STATEMENTS:
            for p in strong:
                m = re.search(p, low) or re.search(p, t)
                if m and (best is None or len(m.group(0)) > best[0]):
                    best = (len(m.group(0)), name)
            # ``anchored`` on the candidate itself, for text whose CONTEXT is the anchor. The
            # English-anchor test exists because a page's prose says "cash flows" constantly, so a
            # weak pattern alone would classify an auditor's paragraph; a worksheet TAB NAME is not
            # prose but a deliberate label of what the sheet holds, and "Cash Flow" on a tab means
            # exactly one thing. Page candidates never set it, so nothing about the page path moves.
            if c.get("anchored") or _anchored(t):
                for p in weak:
                    m = re.search(p, low) or re.search(p, t)
                    if m and (best is None or len(m.group(0)) > best[0]):
                        best = (len(m.group(0)), name)
        if best:
            hits.append((c.get("y", 0.0), -best[0], best[1], t))
            names.add(best[1])

    if not hits:
        return None, False, None, False
    if len(names) >= _MAX_DISTINCT_TITLES:   # a contents/index page lists them all at once
        return None, False, None, False

    hits.sort(key=lambda h: (h[0], h[1]))
    _, _, name, title = hits[0]
    combined = bool(_OCI_COMBINED.search(title))
    if combined:
        name = "profit_and_loss"
    ambig = bool(_ZH_CI_AMBIG.search(title)) and not combined
    return name, combined, title, ambig


# How many leading rows of a worksheet are read looking for its title, and how many text cells are
# taken from them. A statement title sits at the top of the sheet, above the column headings; reading
# further only offers the decode data rows to mistake for a heading.
_SHEET_TITLE_ROWS = 15
_SHEET_TITLE_CELLS = 12


def statement_of_sheet(sheet_name: str, cell_texts: list[str]) -> tuple[str | None, str | None]:
    """``(statement, matched_title)`` for one worksheet, or ``(None, None)``.

    The title vocabulary is NOT restated here. A worksheet's title is the same phrase as a printed
    page's — "Consolidated statement of financial position", 綜合財務狀況表 — so the candidates go
    through ``_title_candidates`` and ``_resolve_statement`` exactly as a page's lines do, and every
    rule they carry comes along: the strong/weak patterns per statement, the negative filter that
    rejects note headings and contents lines, the two-line join, the OCI-combined collapse, and the
    contents-page guard that refuses a sheet listing every statement at once.

    The line dicts are synthesised the way ``_page_lines`` synthesises them for a page whose spans
    cannot be read — ``y`` is an ordinal, not a coordinate — because ordering is all
    ``_resolve_statement`` needs from it.

    The SHEET NAME is offered first, at the topmost ordinal, so it wins a position tie against a
    title inside the sheet. It is the more reliable statement of what a sheet IS: a tab called
    "Balance Sheet" says so deliberately, whereas the first rows of a sheet may carry the entity name
    or a prior statement's tail. A name that matches nothing contributes no candidate at all, so an
    unhelpful "Sheet1" costs nothing.
    """
    lines: list[dict] = [{"text": (sheet_name or "").strip(), "y": -1.0, "size": 0.0,
                          "bold": False, "anchored": True}]
    for i, text in enumerate(cell_texts[:_SHEET_TITLE_CELLS]):
        if text and text.strip():
            lines.append({"text": text.strip(), "y": float(i), "size": 0.0, "bold": False})
    lines = [line for line in lines if line["text"]]
    if not lines:
        return None, None
    statement, _combined, matched, _ambig = _resolve_statement(_title_candidates(lines))
    return (_STATEMENT_ALIAS.get(statement or "", statement), matched)


def sheet_title_cells(sheet) -> list[str]:
    """The text cells of a worksheet's leading rows, in reading order — the input above."""
    out: list[str] = []
    for row in sheet.iter_rows(min_row=1, max_row=_SHEET_TITLE_ROWS, values_only=True):
        for value in row:
            if isinstance(value, str) and value.strip():
                out.append(value.strip())
    return out


def _scope_of(title: str | None, lines: list[dict], page_h: float,
              in_notes_region: bool = False) -> tuple[str | None, list[str]]:
    """Scope from the title, plus column-header scope when a Group and a Company column sit side by
    side on one face page — routine in HK balance sheets, and the reason scope_columns exists."""
    scope = None
    if title:
        if _SCOPE_CONSOL.search(title) or _ZH_CONSOL_AMBIG.search(title):
            scope = "consolidated"
        elif _SCOPE_COMPANY.search(title):
            scope = "company"
        elif in_notes_region:
            # A face-titled page inside the notes carrying no consolidation token is the Company-only
            # statement of financial position: HK filings print it there, past note 40, untitled as
            # to scope.
            scope = "company"
    band = " ".join(l["text"] for l in lines if l.get("y", 0.0) <= 0.42 * (page_h or 1.0))
    cols: list[str] = []
    if re.search(r"\bgroup\b|[本][集][團团]", band, re.I):
        cols.append("consolidated")
    if re.search(r"\bcompany\b|\bbank\b|[本][公][司]", band, re.I):
        cols.append("company")
    if len(cols) == 2:
        scope = "mixed"
    return scope, cols


def _page_lines(page) -> tuple[list[dict], float]:
    """Lines as (text, y, size, bold), top-down. Read from the span dict rather than plain text
    because a title's position and weight are evidence the decode uses."""
    height = float(getattr(page.rect, "height", 0.0) or 0.0)
    try:
        data = page.get_text("dict")
    except Exception:  # noqa: BLE001 — a damaged page still yields plain text below
        data = None
    lines: list[dict] = []
    if data:
        for block in data.get("blocks", []):
            for line in block.get("lines", []):
                spans = line.get("spans") or []
                text = "".join(s.get("text", "") for s in spans).strip()
                if not text:
                    continue
                size = max((float(s.get("size", 0.0)) for s in spans), default=0.0)
                bold = any("bold" in str(s.get("font", "")).lower() for s in spans)
                y = float(line.get("bbox", (0, 0, 0, 0))[1])
                lines.append({"text": text, "y": y, "size": size, "bold": bold})
        lines.sort(key=lambda l: l["y"])
    if not lines:
        raw = page.get_text("text") or ""
        lines = [{"text": s.strip(), "y": float(i), "size": 0.0, "bold": False}
                 for i, s in enumerate(raw.splitlines()) if s.strip()]
    return lines, height


def _features(index: int, lines: list[dict], page_h: float, text: str) -> PageFeat:
    f = PageFeat(index=index)
    zone = _title_zone(lines)
    f.title_lines = [l["text"] for l in zone]
    cands = _title_candidates(zone)

    f.statement, f.oci_combined, title, f.title_ambig = _resolve_statement(cands)
    joined = " ".join(f.title_lines)
    f.narrative = bool(_NARRATIVE.search(joined) or _NARRATIVE.search(text[:1500]))
    if f.narrative:
        # The auditor's report names every statement it audited, in bold, in the top band.
        f.statement, title = None, None
    f.matched_title = title
    f.strong_title = f.statement is not None

    f.notes_banner = any(re.search(p, joined, re.I) for p in _NOTES_BANNER)
    f.note_heading = bool(_NUMBERED_HEADING.search(text)) and not f.strong_title
    f.note_one = bool(_NOTE_ONE.search(text)) and not f.strong_title
    f.backmatter = bool(_BACKMATTER.search(joined))

    tokens = text.split()
    f.numeric_density = (sum(1 for t in tokens if _NUM_TOKEN.fullmatch(t)) / len(tokens)
                         if tokens else 0.0)

    if not f.statement:
        for c in cands:
            t = c["text"]
            if _TITLE_NEGATIVE.search(t) or len(t) > 120:
                continue
            if _TITLE_HINT.search(t):
                f.unmapped.append(t)

    f.scope, f.scope_columns = _scope_of(title, lines, page_h)
    return f


# ------------------------------------------------------------------ decode ---
# States, in document order. PRE and POST both surface as OTHER; they are separate states because
# what may follow them differs — front matter can become a face, back-matter should not.
_PRE, _FACE, _NOTES, _POST = "pre", "face", "notes", "post"
_STATES = (_PRE, _FACE, _NOTES, _POST)

# Cost of leaving state A for state B. Zero is "expected in a filing"; a large number is "possible,
# but the evidence had better be strong". NOTES → FACE is the load-bearing one: at 3.0 a genuine
# face title after the notes can win, which is how a Company-only balance sheet printed past note 40
# is recovered — the fixed region walk this replaces made that transition impossible.
_TRANSITION: dict[tuple[str, str], float] = {
    (_PRE, _PRE): 0.0, (_PRE, _FACE): 0.0, (_PRE, _NOTES): 1.0, (_PRE, _POST): 6.0,
    (_FACE, _FACE): 0.0, (_FACE, _NOTES): 0.0, (_FACE, _POST): 4.0, (_FACE, _PRE): 8.0,
    (_NOTES, _NOTES): 0.0, (_NOTES, _POST): 0.0, (_NOTES, _FACE): 3.0, (_NOTES, _PRE): 10.0,
    (_POST, _POST): 0.0, (_POST, _NOTES): 6.0, (_POST, _FACE): 6.0, (_POST, _PRE): 10.0,
}


def _emission(f: PageFeat, state: str) -> float:
    """How well this page's own evidence fits one state. Deliberately coarse: the decode's job is to
    combine weak local signals with document order, not to be certain page by page."""
    dense = 1.0 if f.numeric_density >= 0.18 else 0.0
    if state == _FACE:
        s = 6.0 if f.strong_title else 0.0
        s += dense
        s -= 6.0 if f.narrative else 0.0
        s -= 4.0 if f.notes_banner else 0.0
        s -= 3.0 if (f.note_heading or f.note_one) else 0.0
        s -= 4.0 if f.backmatter else 0.0
        return s
    if state == _NOTES:
        s = 6.0 if f.notes_banner else 0.0
        s += 4.0 if f.note_one else 0.0
        s += 2.5 if f.note_heading else 0.0
        s += dense * 0.5
        s -= 4.0 if f.strong_title else 0.0
        s -= 5.0 if f.narrative else 0.0
        s -= 4.0 if f.backmatter else 0.0
        return s
    if state == _POST:
        return 6.0 if f.backmatter else -2.0
    # PRE: prose, or simply nothing that looks like a statement or a note.
    s = 1.0
    s += 2.0 if f.narrative else 0.0
    s -= 3.0 if f.strong_title else 0.0
    s -= 3.0 if f.notes_banner else 0.0
    s -= 1.0 if dense else 0.0
    return s


def _decode(feats: list[PageFeat]) -> tuple[list[str], list[float]]:
    """Viterbi over the page sequence. Returns the state path and each page's decode MARGIN — how
    much better the chosen state was than the runner-up, which is a measured confidence rather than
    the fixed constant per branch the previous classifier served."""
    if not feats:
        return [], []
    score = {s: _emission(feats[0], s) + (0.0 if s in (_PRE, _FACE) else -2.0) for s in _STATES}
    back: list[dict[str, str]] = []
    for f in feats[1:]:
        nxt: dict[str, float] = {}
        step: dict[str, str] = {}
        for s in _STATES:
            best_prev, best_val = None, float("-inf")
            for p in _STATES:
                cost = _TRANSITION.get((p, s))
                if cost is None:
                    continue
                v = score[p] - cost
                if v > best_val:
                    best_prev, best_val = p, v
            nxt[s] = best_val + _emission(f, s)
            step[s] = best_prev or _PRE
        score, _ = nxt, back.append(step)

    last = max(_STATES, key=lambda s: score[s])
    path = [last]
    for step in reversed(back):
        path.append(step[path[-1]])
    path.reverse()

    # The margin is per page against its own alternatives, given the state actually chosen.
    margins: list[float] = []
    for f, s in zip(feats, path):
        others = [_emission(f, o) for o in _STATES if o != s]
        margins.append(_emission(f, s) - max(others))
    return path, margins


_KIND = {_PRE: PageKind.OTHER, _FACE: PageKind.FACE,
         _NOTES: PageKind.NOTES, _POST: PageKind.OTHER}


def _confidence(margin: float) -> float:
    """A decode margin mapped into 0..1. Measured, not asserted: the classifier this replaces served
    a fixed 0.72 for every face page and 0.4 for everything else, so a page it was sure about and one
    it guessed reported the same number."""
    return round(min(0.97, max(0.30, 0.5 + margin / 20.0)), 4)


def dump_review(doc: DocumentModel) -> str:
    """A per-page decision table, plus the unmapped titles as a trailing block.

    The unmapped titles are document-level and are what tell you whether the lexicon still has holes,
    so they are written once at the end rather than repeated on every row.
    """
    rows = ["idx\tkind\tstatement\tscope\tcolumns\tconf\tmatched_title"]
    for p in doc.pages:
        ev = p.evidence or {}
        rows.append("\t".join([
            str(p.index), str(getattr(p.kind, "value", p.kind)), p.statement or "—",
            p.scope or "—", ",".join(p.scope_columns) or "—",
            "" if p.classification_confidence is None else f"{p.classification_confidence:.2f}",
            str(ev.get("matched_title") or "—"),
        ]))
    if doc.unmapped_titles:
        rows.append("")
        rows.append(f"# unmapped titles ({len(doc.unmapped_titles)}) — statement-ish, resolved to "
                    f"nothing; fold the real vocabulary back into _STATEMENTS")
        rows.extend(f"# {t}" for t in doc.unmapped_titles)
    return "\n".join(rows)


class ClassifyStage:
    name = "classify"

    @staticmethod
    def _classify_workbook(doc: DocumentModel, data: bytes,
                           ctx: PipelineContext) -> DocumentModel:
        """Name each worksheet's statement, so a spreadsheet is scoped like a page.

        THE DEFECT THIS CLOSES. This stage used to return early for anything that is not a PDF, so
        every worksheet kept ``statement=None`` from ingest — and a statement is not decoration
        downstream, it is a BOUNDARY. ``residual._section_of_row`` guards each of its structural
        signals with ``if statement and statement_of(nxt) not in (None, statement)``, which is inert
        when the statement is None: the walk then runs past the end of the sheet it started on and a
        balance-sheet row can take its section from a cash-flow subtotal on a later sheet. The v1
        router (``residual._route_by_template``) is keyed by statement type outright, so it placed no
        Excel row at all. ``map_ontology.batch_groups`` likewise had no statement to batch by, so
        every spreadsheet row was mapped with the whole ontology in front of it.

        The classifier's own machinery decides it (``statement_of_sheet``); nothing about the
        vocabulary is duplicated for spreadsheets.

        A sheet whose title resolves is a FACE page. One whose title does not is left as ingest set
        it rather than guessed at: ``kind`` gates ``doc.face_pages()``, and calling a cover sheet or a
        list of assumptions a face would put its rows into the statement.
        """
        try:
            import openpyxl
        except ImportError:                      # pragma: no cover - openpyxl is a hard dependency
            return doc
        try:
            wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        except Exception as exc:  # noqa: BLE001 — a workbook we cannot reopen is not a failure here
            ctx.log(f"classify:xlsx_open_failed:{exc}")
            return doc
        try:
            names = wb.sheetnames
            for page in doc.pages:
                if page.index >= len(names):
                    continue
                name = names[page.index]
                statement, matched = statement_of_sheet(name, sheet_title_cells(wb[name]))
                page.statement = statement
                if statement:
                    page.kind = PageKind.FACE
                page.evidence = {"sheet": name, "matched_title": matched}
                ctx.log(f"classify:sheet={name}:statement={statement or 'unresolved'}")
        finally:
            wb.close()
        return doc

    def run(self, doc: DocumentModel, ctx: PipelineContext) -> DocumentModel:
        data = ctx.raw_bytes
        if not data:
            return doc
        if doc.fmt in (DocFormat.XLSX, DocFormat.XLS):
            return self._classify_workbook(doc, data, ctx)
        if doc.fmt.value != "pdf":
            return doc
        try:
            import fitz
        except ImportError:
            return doc
        try:
            pdf = fitz.open(stream=data, filetype="pdf")
        except Exception:  # noqa: BLE001
            return doc

        pages = [p for p in doc.pages if p.index < len(pdf)]
        feats: list[PageFeat] = []
        cache: list[tuple[list[dict], float]] = []
        for page_src in pages:
            lines, height = _page_lines(pdf[page_src.index])
            text = "\n".join(l["text"] for l in lines)
            cache.append((lines, height))
            feats.append(_features(page_src.index, lines, height, text))

        path, margins = _decode(feats)

        # A statement runs across several pages and only the first is titled, so a face page with no
        # resolvable title inherits the last one named. Reset when the face run ends.
        current: str | None = None
        seen_notes = False
        for page_src, f, state, margin, (lines, height) in zip(
                pages, feats, path, margins, cache):
            page_src.kind = _KIND[state]
            page_src.classification_confidence = _confidence(margin)
            if state == _FACE:
                named = _STATEMENT_ALIAS.get(f.statement or "", f.statement)
                if named:
                    current = named
                page_src.statement = current
                # Scope is resolved again here because only the decode knows whether this face page
                # sits after the notes — which is what makes an untitled one the Company statement.
                scope, cols = _scope_of(f.matched_title, lines, height,
                                        in_notes_region=seen_notes)
                page_src.scope = f.scope or scope
                page_src.scope_columns = f.scope_columns or cols
            else:
                page_src.statement = None
                current = None if state == _NOTES else current
            if state == _NOTES:
                seen_notes = True
            page_src.evidence = {"state": state, "matched_title": f.matched_title,
                                 "title_ambig": f.title_ambig, "margin": round(margin, 2),
                                 "oci_combined": f.oci_combined}
            if f.unmapped:
                doc.unmapped_titles.extend(f.unmapped[:3])

        pdf.close()
        doc.unmapped_titles = sorted(set(doc.unmapped_titles))[:60]
        stmts = {p.statement for p in doc.face_pages() if p.statement}
        ambig = sum(1 for p in doc.pages if (p.evidence or {}).get("title_ambig"))
        ctx.log(f"classify:face={len(doc.face_pages())} notes={len(doc.notes_pages())} "
                f"statements={sorted(stmts)}")
        ctx.log(f"classify:unmapped_titles={len(doc.unmapped_titles)} title_ambig={ambig}")
        return doc
