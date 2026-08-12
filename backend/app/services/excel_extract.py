"""Deterministic Excel → line items with exact cell-level provenance.

Native spreadsheets need no OCR and no LLM to extract: the value and its origin are read
straight from the workbook. Each numeric row becomes a ``LineItem`` whose ``ExtractedValue``
carries a ``Provenance`` pointing at the exact sheet + cell (e.g. "P&L!C14"). Column headers
are used to key values by period. Semantic mapping (which canonical concept each row is) is
a separate, LLM-driven step — this stage only produces trustworthy, source-anchored facts.
"""
from __future__ import annotations

import io
import re
from decimal import Decimal, InvalidOperation

from app.core.models.enums import Basis, LineRole, ValueSource
from app.core.models.geometry import Provenance
from app.core.models.line_item import ExtractedValue, LineItem, UnitContext

_PRODUCER = "extract:xlsx@0.1.0"
_CELL_RE = re.compile(r"^([A-Za-z]+)(\d+)$")


def _col_letter(idx1: int) -> str:
    """1-based column index → Excel column letters (1→A, 27→AA)."""
    s = ""
    while idx1 > 0:
        idx1, r = divmod(idx1 - 1, 26)
        s = chr(65 + r) + s
    return s


def _col_index0(letters: str) -> int:
    """Excel column letters → 0-based column index (A→0, AA→26)."""
    n = 0
    for ch in letters.upper():
        n = n * 26 + (ord(ch) - 64)
    return n - 1


def _parse_cell(ref: str) -> tuple[int, int] | None:
    """"C14" → (col0=2, row0=13); None if malformed."""
    m = _CELL_RE.match(ref.strip())
    if not m:
        return None
    return _col_index0(m.group(1)), int(m.group(2)) - 1


def _to_decimal(v) -> Decimal | None:
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        try:
            return Decimal(str(v))
        except InvalidOperation:
            return None
    if isinstance(v, str):
        t = v.strip().replace(",", "")
        neg = t.startswith("(") and t.endswith(")")
        t = t.strip("()").replace("%", "").strip()
        if not t:
            return None
        try:
            d = Decimal(t)
            return -d if neg else d
        except InvalidOperation:
            return None
    return None


_NOTE_HEADER = re.compile(r"^\s*note", re.IGNORECASE)
_NOTE_NUM = re.compile(r"(\d{1,3})")


def extract_workbook(data: bytes, *, document_id: str | None = None, log=None) -> list[LineItem]:
    """Read every sheet; emit a LineItem per labelled, numeric row with cell provenance.

    Reaches parity with the PDF path, and reads the same rulebook it does
    (``scope_selection``): a dedicated "Note"/"Note No." column becomes the row's note reference
    rather than a value; a two-basis header band (Group | Company as well as Consolidated |
    Standalone) maps each value column to its basis; the current period is the column whose
    heading is the latest DATE; the sheet's declared unit is persisted on every fact; and two
    facts differing on a declared dimension are kept as two facts (``column_guard``)."""
    import openpyxl

    from app.services.row_reconstruct import (
        _entity_signals,
        _restated_markers,
        _unit_signals,
        guard_dimensions,
        in_force_rules,
        store_fact,
    )

    scope, _norm = in_force_rules()
    signals, restated = _entity_signals(scope), _restated_markers(scope)
    dims = guard_dimensions(scope)

    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    items: list[LineItem] = []
    ordinal = 0
    try:
        for sheet_index, name in enumerate(wb.sheetnames):
            ws = wb[name]
            rows = [list(r) for r in ws.iter_rows(values_only=True)]
            if not rows:
                continue
            label_col, value_cols = _detect_columns(rows)
            if label_col is None or not value_cols:
                continue
            note_col = _detect_note_col(rows, label_col, value_cols)
            if note_col is not None and note_col in value_cols:
                value_cols = [c for c in value_cols if c != note_col]
            basis_map = _detect_basis_bands(rows, value_cols, label_col, signals)
            unit_signals = _unit_signals(scope)
            headers = _period_headers(rows, value_cols, basis_map, signals=signals,
                                      unit_signals=unit_signals, restated=restated)
            unit_ctx = _sheet_unit(rows, unit_signals, sheet_index)
            if log and unit_ctx.units_label:
                log(f"extract:sheet={name}:units="
                    f"{unit_ctx.currency or 'unknown_ccy'}/{unit_ctx.units_label}")
            for r_idx, row in enumerate(rows):
                if label_col >= len(row):
                    continue
                label = row[label_col]
                if not isinstance(label, str) or not label.strip():
                    continue
                li = _row_item(name, sheet_index, r_idx, label.strip(), row, label_col,
                               value_cols, headers, document_id, ordinal, note_col, basis_map,
                               unit_ctx=unit_ctx, dims=dims, log=log, store=store_fact)
                if li is not None:
                    items.append(li)
                    ordinal += 1
    finally:
        wb.close()
    return items


def _sheet_unit(rows: list[list], signals, sheet_index: int) -> UnitContext:
    """The unit declared in this SHEET's header, persisted on every fact it produces.

    ``units_and_currency`` says to re-resolve per statement and never to normalise scale
    silently: a workbook whose sheets are in different scales gets one unit per sheet, and a
    sheet that declares none gets an empty context rather than an assumed factor of one.
    """
    from app.services.row_reconstruct import _fold_for_match, _unit_context

    for row in rows[:8]:
        blob = _fold_for_match(" ".join(str(v) for v in row if isinstance(v, str)))
        for folded, ccy, scale, word in signals:
            if folded and folded in blob:
                return _unit_context((ccy, scale, word or folded), sheet_index)
    return UnitContext()


def _detect_note_col(rows: list[list], label_col: int, value_cols: list[int]) -> int | None:
    """A column whose header cell reads "Note"/"Note No." — its numbers are note references,
    not values. Checked against the text header rows (before the numbers start)."""
    for row in rows[:8]:
        for c in range(len(row)):
            if c == label_col:
                continue
            v = row[c]
            if isinstance(v, str) and _NOTE_HEADER.match(v) and len(v.strip()) <= 12:
                return c
    return None


def _detect_basis_bands(rows: list[list], value_cols: list[int],
                        label_col: int | None = None,
                        signals: tuple[tuple[Basis, str], ...] = ()) -> dict[int, Basis]:
    """Map each value column to a basis from a header band that names them (a caption applies to
    value columns at/after its own column, until the next caption).

    Driven by ``scope_selection.entity_scope.signals``, so a sheet headed "Group | Company" —
    which the previous token list did not know — bands as one, and three defects it had are
    closed with it:

    * it scanned the LABEL column, where "Consolidated statement of financial position" and a
      company name both live, so the label cell defined the band for the whole sheet;
    * it stopped at the FIRST row carrying any caption, so a two-line header whose basis captions
      sit on the second line was never reached;
    * it accepted ONE caption, and one caption relabels every column — the same both-or-nothing
      rule the PDF path applies (see ``row_reconstruct._basis_bands``).
    """
    from app.services.row_reconstruct import _signal_side

    best: list[tuple[int, Basis]] = []
    for row in rows[:8]:
        markers: list[tuple[int, Basis]] = []
        for c in value_cols:                     # value columns only: never the label column
            if c == label_col or c >= len(row):
                continue
            side = _signal_side(row[c], signals) if isinstance(row[c], str) else None
            if side is not None:
                markers.append((c, side))
        if len({b for _, b in markers}) >= 2:    # both-or-nothing
            best = markers
            break
    if not best:
        return {}
    best.sort()
    out: dict[int, Basis] = {}
    for c in value_cols:
        basis = best[0][1]
        for mc, mb in best:
            if mc <= c:
                basis = mb
            else:
                break
        out[c] = basis
    return out


def _detect_columns(rows: list[list]) -> tuple[int | None, list[int]]:
    """Label column = the text-heaviest column; value columns = the numeric ones."""
    width = max((len(r) for r in rows), default=0)
    text_score = [0] * width
    num_score = [0] * width
    for row in rows:
        for c in range(len(row)):
            v = row[c]
            if isinstance(v, str) and v.strip() and _to_decimal(v) is None:
                text_score[c] += 1
            elif _to_decimal(v) is not None:
                num_score[c] += 1
    if not any(num_score):
        return None, []
    label_col = max(range(width), key=lambda c: text_score[c]) if any(text_score) else 0
    value_cols = [c for c in range(width) if c != label_col and num_score[c] >= 1]
    return label_col, value_cols


def _positional(i: int, c: int) -> str:
    return "current" if i == 0 else "prior" if i == 1 else f"col{c}"


def _period_headers(rows: list[list], value_cols: list[int], basis_map=None, *,
                    signals: tuple[tuple[Basis, str], ...] = (),
                    unit_signals: tuple[tuple[str, str, object, str], ...] = (),
                    restated: tuple[str, ...] = ()) -> dict[int, tuple[str, str | None]]:
    """(period_label, period_display) per value column, from the first text header row.

    A header that NAMES A PERIOD ("31 December 2024", "FY2025", "2023") is labelled positionally —
    current, prior, colN — with the printed text kept as the display, exactly as the native-PDF
    path does. Storing the text as the label instead looked harmless and silently defeated
    current/prior resolution for every spreadsheet: nothing downstream recognised the column, the
    positional fallback fired for every row, and a line with a figure in the prior column only had
    that figure reported as the current year.

    A header that names something other than a period ("Retained profits") is an equity component,
    and its identity IS its name — so that is kept as the label.

    Rows that name the basis (a Consolidated/Standalone band) are skipped; those aren't periods.

    Position is counted WITHIN a basis: a Consolidated | Standalone sheet has four value columns
    and two of each period, so "current" is the first column of its own band. Counting across the
    whole row would label the standalone pair col2/col3 and leave the standalone statement with no
    current or prior column at all.

    Which of the period columns is the CURRENT one comes from the heading date when every column
    of a band carries one (``scope_selection.period_selection``: "not from column position"), and a
    column headed as restated keeps the comparative slot without displacing an original printed
    beside it (``restatement_rule``).
    """
    from app.services.row_reconstruct import _is_date_ish, _signal_side

    for row in rows:
        if any(c < len(row) and isinstance(row[c], str) and row[c].strip() for c in value_cols):
            # A YEAR is not an amount. A sheet whose columns are headed 2024 / 2023 parses those
            # cells as numbers, so the header row looked like data and the columns lost their
            # headings altogether — no date to order the periods by, and no display text.
            has_numbers = any(d is not None and not _is_date_ish(d)
                              for d in (_to_decimal(row[c]) for c in value_cols
                                        if c < len(row)))
            # A basis band names entities, not periods, and a units row names the scale. Both are
            # recognised from the declared signals: with "group"/"company" missing from the old
            # token list a Group | Company band row was read as the period header and every column
            # took an entity's name as its period, and a "RMB'000" row did the same with the scale.
            is_basis_band = any(c < len(row) and isinstance(row[c], str)
                                and _signal_side(row[c], signals) is not None
                                for c in value_cols)
            if has_numbers or is_basis_band or _is_units_row(row, value_cols, unit_signals):
                continue
            texts = {c: (str(row[c]).strip() if c < len(row) and row[c] else "")
                     for c in value_cols}
            return _column_labels(texts, value_cols, basis_map, restated)
    # No text header row at all: positional, first value column of each basis = current period.
    return _column_labels({c: "" for c in value_cols}, value_cols, basis_map, restated)


def _is_units_row(row: list, value_cols: list[int],
                  unit_signals: tuple[tuple[str, str, object, str], ...]) -> bool:
    """Whether every captioned value cell of this row is a declared unit annotation."""
    from app.services.row_reconstruct import _fold_for_match

    captions = [str(row[c]).strip() for c in value_cols
                if c < len(row) and isinstance(row[c], str) and row[c].strip()]
    if not captions or not unit_signals:
        return False
    return all(any(folded and folded in _fold_for_match(t) for folded, *_ in unit_signals)
               for t in captions)


def _column_labels(texts: dict[int, str], value_cols: list[int], basis_map,
                   restated: tuple[str, ...]) -> dict[int, tuple[str, str | None]]:
    """(period_label, period_display) per value column, ordered by heading date within each basis."""
    from app.services.periods import looks_like_period
    from app.services.row_reconstruct import _is_restated, _period_date

    out: dict[int, tuple[str, str | None]] = {}
    bands = {(basis_map or {}).get(c, Basis.CONSOLIDATED) for c in value_cols}
    for band in bands:
        cols = [c for c in value_cols if (basis_map or {}).get(c, Basis.CONSOLIDATED) == band]
        # A header that names something other than a period ("Retained profits") is an equity
        # component, and its identity IS its name.
        periods = [c for c in cols if not texts[c] or looks_like_period(texts[c])]
        dates = {c: _period_date(texts[c]) for c in periods}
        flags = {c: _is_restated(texts[c], restated) for c in periods}
        groups = sorted({d for d in dates.values() if d}, reverse=True)
        if len(groups) > 1 and all(dates[c] for c in periods):
            slot_of = {c: groups.index(dates[c]) for c in periods}
        else:
            slot_of = {c: i for i, c in enumerate(periods)}
        used: set[str] = set()
        for c in sorted(periods, key=lambda c: (slot_of[c], flags[c])):
            base = _positional(slot_of[c], c)
            label = base if base not in used else (
                f"{base}_restated" if flags[c] else f"{base}_col{c}")
            used.add(label)
            out[c] = (label, texts[c] or None)
        for c in cols:
            if c not in out:
                out[c] = (texts[c], texts[c])
    return out


def _note_from_cell(v) -> str | None:
    """Parse a note reference from a note-column cell: 14, 14.0, "Note 14", or "14"."""
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return str(int(v))
    if isinstance(v, str):
        m = _NOTE_NUM.search(v)
        return m.group(1) if m else None
    return None


def _row_item(sheet, sheet_index, r_idx, label, row, label_col, value_cols, headers,
              document_id, ordinal, note_col=None, basis_map=None, *,
              unit_ctx: UnitContext | None = None, dims: tuple[str, ...] = (),
              log=None, store=None) -> LineItem | None:
    from app.core.models.line_item import NoteRef

    li = LineItem(source_label=label, ordinal=ordinal, role=LineRole.LINE,
                  source=ValueSource.MACHINE)
    if note_col is not None and note_col < len(row):
        note = _note_from_cell(row[note_col])
        if note:
            li.note_number = note
            li.note_refs.append(NoteRef(raw=note, numbers=[note]))
    got = False
    for c in value_cols:
        if c >= len(row):
            continue
        dec = _to_decimal(row[c])
        if dec is None:
            continue
        basis = (basis_map or {}).get(c, Basis.CONSOLIDATED)
        prov = Provenance(
            document_id=document_id, page_index=sheet_index, sheet=sheet,
            cell=f"{_col_letter(c + 1)}{r_idx + 1}",
            label_cell=f"{_col_letter(label_col + 1)}{r_idx + 1}",
            text_snippet=label, source_kind="spreadsheet", producer=_PRODUCER,
        )
        label_, display_ = headers.get(c, (f"col{c}", None))
        ev = ExtractedValue(value_raw=dec, value=dec, basis=basis,
                            period_label=label_, period_display=display_,
                            unit_ctx=unit_ctx or UnitContext(), provenance=prov)
        if store is not None:
            # Two columns can carry the same header text ("Total" twice, a period repeated under
            # two units): column_guard says those are distinct facts, and set_value alone would
            # keep whichever cell was read last.
            store(li, ev, dims, log=log, where=f"sheet={sheet}:{r_idx + 1}:")
        else:
            li.set_value(ev)
        got = True
    return li if got else None


def cell_context(data: bytes, *, sheet: str, cell: str, radius: int = 4) -> dict:
    """A small window of the sheet around ``cell`` — the spreadsheet equivalent of the PDF
    page image, so the frontend can render the value *in situ* and highlight its cell.

    Raises ``ValueError`` for a bad cell ref and ``KeyError`` for an unknown sheet.
    """
    import openpyxl

    parsed = _parse_cell(cell)
    if parsed is None:
        raise ValueError(f"Bad cell reference: {cell!r}")
    tcol, trow = parsed

    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    try:
        if sheet not in wb.sheetnames:
            raise KeyError(sheet)
        rows = [list(r) for r in wb[sheet].iter_rows(values_only=True)]
    finally:
        wb.close()

    width = max((len(r) for r in rows), default=tcol + 1)
    r0, r1 = max(0, trow - radius), min(max(len(rows), trow + 1), trow + radius + 1)
    c0, c1 = max(0, tcol - radius), min(max(width, tcol + 1), tcol + radius + 1)

    grid: list[list[dict]] = []
    for r in range(r0, r1):
        row = rows[r] if r < len(rows) else []
        line: list[dict] = []
        for c in range(c0, c1):
            v = row[c] if c < len(row) else None
            line.append({
                "ref": f"{_col_letter(c + 1)}{r + 1}",
                "value": "" if v is None else str(v),
                "is_target": c == tcol and r == trow,
                "numeric": _to_decimal(v) is not None,
            })
        grid.append(line)

    return {
        "sheet": sheet,
        "target": cell,
        "col_letters": [_col_letter(c + 1) for c in range(c0, c1)],
        "row_numbers": list(range(r0 + 1, r1 + 1)),
        "grid": grid,
    }
