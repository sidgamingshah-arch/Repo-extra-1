"""Derived analysis from a real extraction — all computed from the extracted values (never
fabricated):

1. Ratios     — standard financial ratios computed from canonical line-item values.
2. Disclosures— presence scan for qualitative items (auditor qualification, contingent
                liabilities, guarantees, …) with the page + snippet where found.
3. Free notes — plain-language notes generated from the extracted movements/ratios.

Ratios and notes are recomputed from the current rows (so edits/reverts are reflected);
disclosures come from a one-time text scan stored on the run.
"""
from __future__ import annotations

import re

from app.services.periods import concept_value

# --- 1. Ratios -------------------------------------------------------------
# Each ratio references canonical keys; num/den are (key, sign) so we can net inventories
# out of the quick ratio etc. A ratio is computed only when ALL its inputs are present.
# Reusable component term-lists. Debt/cash components are "opt" — a company reports only some
# of them, and the aggregate should still compute from whatever is present.
_DEBT = [
    ("bs_non_current_liabilities__non_current_borrowings", 1, "opt"),
    ("bs_non_current_liabilities__non_current_bonds_payable", 1, "opt"),
    ("bs_non_current_liabilities__non_current_notes_payable", 1, "opt"),
    ("bs_non_current_liabilities__non_current_lease_liabilities", 1, "opt"),
    ("bs_current_liabilities__current_borrowings", 1, "opt"),
    ("bs_current_liabilities__current_potion_of_long_term_debt", 1, "opt"),
    ("bs_current_liabilities__cuurent_notes_payable", 1, "opt"),
    ("bs_current_liabilities__current_lease_liabilities", 1, "opt"),
]
_CASH = [
    ("bs_current_assets__cash_and_cash_equivalents", 1, "opt"),
    ("bs_current_assets__bank_balances_other_than_cash_and_cash_equivalents", 1, "opt"),
]
# EBITDA = EBIT (required anchor) + depreciation & amortisation (optional add-back).
_EBITDA = [("pl_operating_profit_ebit", 1), ("pl_expenses__depreciation_and_amortisation_expense", 1, "opt")]
_CFO = "cf_cash_flow_from_operating_activities__net_cash_from_operating_activities"
_CAPEX = "cf_cash_flow_from_investing_activities__purchase_of_property_plant_and_equipment"
_INTEREST_PAID = "cf_cash_flow_from_financing_activities__interest_paid"
_REPAYMENT = "cf_cash_flow_from_financing_activities__repayment_of_borrowings"
_EQUITY = "bs_equity__total_equity"
_TCL = "bs_current_liabilities__total_current_liabilities"
_INTEREST = "pl_non_operating_expenses__interest_expense"
# Cost base for the inventory/payables day-cycle: prefer the explicit cost-of-goods-sold line,
# but many statement formats don't break it out — fall back to the total operating cost
# subtotal (almost always present) so DIO/DPO/CCC still compute.
_COST_BASE = ("pl_expenses__cost_of_goods_sold", "pl_expenses__total_operating_cost")

_RATIOS = [
    # ---- Liquidity -------------------------------------------------------------------
    {"key": "current_ratio", "label": "Current ratio", "unit": "x", "category": "Liquidity",
     "label_i18n": {"zh": "流动比率", "ar": "نسبة التداول", "fr": "Ratio de liquidité générale"},
     "num": [("bs_current_assets__total_current_assets", 1)],
     "den": [(_TCL, 1)],
     "formula": "Total current assets / Total current liabilities"},
    {"key": "quick_ratio", "label": "Quick ratio", "unit": "x", "category": "Liquidity",
     "label_i18n": {"zh": "速动比率", "ar": "نسبة السيولة السريعة", "fr": "Ratio de liquidité réduite"},
     "num": [("bs_current_assets__total_current_assets", 1), ("bs_current_assets__inventories", -1)],
     "den": [(_TCL, 1)],
     "formula": "(Total current assets − Inventories) / Total current liabilities"},
    {"key": "cash_ratio", "label": "Cash ratio", "unit": "x", "category": "Liquidity",
     "label_i18n": {"zh": "现金比率", "ar": "نسبة النقد", "fr": "Ratio de liquidité immédiate"},
     "num": list(_CASH), "den": [(_TCL, 1)],
     "formula": "(Cash + bank balances) / Total current liabilities"},
    {"key": "operating_cash_flow_ratio", "label": "Operating cash flow ratio", "unit": "x",
     "category": "Liquidity",
     "label_i18n": {"zh": "经营现金流比率", "ar": "نسبة التدفق النقدي التشغيلي",
                    "fr": "Ratio de flux de trésorerie d'exploitation"},
     "num": [(_CFO, 1)], "den": [(_TCL, 1)],
     "formula": "Net cash from operating activities / Total current liabilities"},

    # ---- Leverage --------------------------------------------------------------------
    {"key": "debt_to_equity", "label": "Liabilities to equity", "unit": "x", "category": "Leverage",
     "label_i18n": {"zh": "负债权益比", "ar": "نسبة الالتزامات إلى حقوق الملكية",
                    "fr": "Passif / Capitaux propres"},
     "num": [("bs_non_current_liabilities__total_non_current_liabilities", 1),
             (_TCL, 1)],
     "den": [(_EQUITY, 1)],
     "formula": "(Non-current + current liabilities) / Total equity"},
    {"key": "gross_gearing", "label": "Gross gearing (debt/equity)", "unit": "x", "category": "Leverage",
     "label_i18n": {"zh": "总杠杆（债务/权益）", "ar": "الرافعة الإجمالية (الدين/حقوق الملكية)",
                    "fr": "Endettement brut (dette/capitaux propres)"},
     "num": list(_DEBT), "den": [(_EQUITY, 1)],
     "formula": "Total debt / Total equity"},
    {"key": "net_gearing", "label": "Net gearing (net debt/equity)", "unit": "x", "category": "Leverage",
     "label_i18n": {"zh": "净杠杆（净债务/权益）", "ar": "صافي الرافعة (صافي الدين/حقوق الملكية)",
                    "fr": "Endettement net (dette nette/capitaux propres)"},
     "num": _DEBT + [(k, -1, "opt") for (k, _s, _m) in _CASH], "den": [(_EQUITY, 1)],
     "formula": "(Total debt − Cash) / Total equity"},
    {"key": "debt_to_capital", "label": "Debt to capital", "unit": "%", "category": "Leverage",
     "label_i18n": {"zh": "债务资本比", "ar": "الدين إلى رأس المال", "fr": "Dette / Capital"},
     "num": list(_DEBT), "den": _DEBT + [(_EQUITY, 1)],
     "formula": "Total debt / (Total debt + Total equity)"},
    {"key": "debt_ratio", "label": "Debt ratio (liabilities/assets)", "unit": "%", "category": "Leverage",
     "label_i18n": {"zh": "资产负债率", "ar": "نسبة الدين (الالتزامات/الأصول)",
                    "fr": "Ratio d'endettement (passif/actif)"},
     "num": [("bs_non_current_liabilities__total_non_current_liabilities", 1, "opt"),
             (_TCL, 1, "opt")],
     "den": [("bs_total_assets", 1)],
     "formula": "Total liabilities / Total assets"},
    {"key": "debt_to_ebitda", "label": "Total debt / EBITDA", "unit": "x", "category": "Leverage",
     "label_i18n": {"zh": "总债务/EBITDA", "ar": "إجمالي الدين / EBITDA", "fr": "Dette totale / EBITDA"},
     "num": list(_DEBT), "den": list(_EBITDA),
     "formula": "Total debt / EBITDA (EBIT + depreciation & amortisation)"},
    {"key": "net_debt_to_ebitda", "label": "Net debt / EBITDA", "unit": "x", "category": "Leverage",
     "label_i18n": {"zh": "净债务/EBITDA", "ar": "صافي الدين / EBITDA", "fr": "Dette nette / EBITDA"},
     "num": _DEBT + [(k, -1, "opt") for (k, _s, _m) in _CASH], "den": list(_EBITDA),
     "formula": "(Total debt − Cash) / EBITDA"},
    {"key": "equity_multiplier", "label": "Equity multiplier (assets/equity)", "unit": "x",
     "category": "Leverage",
     "label_i18n": {"zh": "权益乘数（资产/权益）", "ar": "مضاعف حقوق الملكية (الأصول/حقوق الملكية)",
                    "fr": "Multiplicateur de capitaux propres (actif/capitaux propres)"},
     "num": [("bs_total_assets", 1)], "den": [(_EQUITY, 1)],
     "formula": "Total assets / Total equity"},
    {"key": "equity_ratio", "label": "Equity ratio", "unit": "%", "category": "Leverage",
     "label_i18n": {"zh": "权益比率", "ar": "نسبة حقوق الملكية", "fr": "Ratio de capitaux propres"},
     "num": [(_EQUITY, 1)], "den": [("bs_total_assets", 1)],
     "formula": "Total equity / Total assets"},

    # ---- Coverage --------------------------------------------------------------------
    {"key": "interest_coverage", "label": "EBIT interest coverage", "unit": "x", "category": "Coverage",
     "label_i18n": {"zh": "利息保障倍数（EBIT）", "ar": "تغطية الفائدة (EBIT)",
                    "fr": "Couverture des intérêts (EBIT)"},
     "num": [("pl_operating_profit_ebit", 1)], "den": [(_INTEREST, 1)],
     "formula": "Operating profit (EBIT) / Interest expense"},
    {"key": "ebitda_interest_coverage", "label": "EBITDA interest coverage", "unit": "x",
     "category": "Coverage",
     "label_i18n": {"zh": "利息保障倍数（EBITDA）", "ar": "تغطية الفائدة (EBITDA)",
                    "fr": "Couverture des intérêts (EBITDA)"},
     "num": list(_EBITDA), "den": [(_INTEREST, 1)],
     "formula": "EBITDA / Interest expense"},
    {"key": "debt_service_coverage", "label": "Debt-service coverage (DSCR)", "unit": "x",
     "category": "Coverage",
     "label_i18n": {"zh": "偿债保障倍数", "ar": "تغطية خدمة الدين", "fr": "Couverture du service de la dette"},
     "num": list(_EBITDA),
     "den": [(_INTEREST, 1),
             ("bs_current_liabilities__current_potion_of_long_term_debt", 1, "opt"),
             ("bs_current_liabilities__current_borrowings", 1, "opt")],
     "formula": "EBITDA / (Interest expense + current portion of long-term debt + current borrowings)"},
    {"key": "cfo_to_total_debt", "label": "Cash flow to total debt", "unit": "%", "category": "Coverage",
     "label_i18n": {"zh": "经营现金流/总债务", "ar": "التدفق النقدي إلى إجمالي الدين",
                    "fr": "Flux de trésorerie / Dette totale"},
     "num": [(_CFO, 1)], "den": list(_DEBT),
     "formula": "Net cash from operating activities / Total debt"},
    {"key": "cfo_interest_coverage", "label": "Cash-flow interest coverage", "unit": "x",
     "category": "Coverage",
     "label_i18n": {"zh": "现金流利息保障倍数", "ar": "تغطية الفائدة من التدفق النقدي",
                    "fr": "Couverture des intérêts par les flux de trésorerie"},
     "num": [(_CFO, 1)], "den": [(_INTEREST, 1)],
     "formula": "Net cash from operating activities / Interest expense"},
    {"key": "fcf_to_total_debt", "label": "Free cash flow to total debt", "unit": "%",
     "category": "Coverage",
     "label_i18n": {"zh": "自由现金流/总债务", "ar": "التدفق النقدي الحر إلى إجمالي الدين",
                    "fr": "Flux de trésorerie disponible / Dette totale"},
     "num": [(_CFO, 1), (_CAPEX, -1, "opt")], "den": list(_DEBT),
     "formula": "(Net cash from operations − Capex) / Total debt"},
    {"key": "ffo_to_total_debt", "label": "Funds from operations to debt (FFO/debt)", "unit": "%",
     "category": "Coverage",
     "label_i18n": {"zh": "经营资金/总债务", "ar": "الأموال من العمليات إلى الدين",
                    "fr": "Fonds provenant de l'exploitation / Dette totale"},
     "num": [("pl_profit_for_the_year", 1),
             ("pl_expenses__depreciation_and_amortisation_expense", 1, "opt")],
     "den": list(_DEBT),
     "formula": "(Profit for the year + depreciation & amortisation) / Total debt"},
    {"key": "cash_debt_service_coverage", "label": "Cash debt-service coverage", "unit": "x",
     "category": "Coverage",
     "label_i18n": {"zh": "现金偿债保障倍数", "ar": "تغطية خدمة الدين النقدية",
                    "fr": "Couverture du service de la dette par la trésorerie"},
     "num": [(_CFO, 1)], "den": [(_INTEREST_PAID, 1), (_REPAYMENT, 1, "opt")],
     "formula": "Net cash from operating activities / (Interest paid + repayment of borrowings)"},

    # ---- Efficiency (working-capital cycle) ------------------------------------------
    {"key": "working_capital_to_assets", "label": "Working capital / total assets", "unit": "%",
     "category": "Efficiency",
     "label_i18n": {"zh": "营运资本/总资产", "ar": "رأس المال العامل / إجمالي الأصول",
                    "fr": "Fonds de roulement / Actif total"},
     "num": [("bs_current_assets__total_current_assets", 1), (_TCL, -1)],
     "den": [("bs_total_assets", 1)],
     "formula": "(Total current assets − Total current liabilities) / Total assets"},
    {"key": "dso", "label": "Days sales outstanding (DSO)", "unit": "days", "category": "Efficiency",
     "label_i18n": {"zh": "应收账款周转天数", "ar": "أيام تحصيل الذمم المدينة",
                    "fr": "Délai de recouvrement (jours)"},
     "num": [("bs_current_assets__trade_receivables", 1)],
     "den": [("pl_income__revenue_from_operations", 1)],
     "formula": "Trade receivables / Revenue × 365"},
    {"key": "dio", "label": "Days inventory outstanding (DIO)", "unit": "days", "category": "Efficiency",
     "label_i18n": {"zh": "存货周转天数", "ar": "أيام بقاء المخزون", "fr": "Délai d'écoulement des stocks (jours)"},
     "num": [("bs_current_assets__inventories", 1)],
     "den": [(_COST_BASE, 1)],
     "formula": "Inventories / Cost of goods sold (or total operating cost) × 365"},
    {"key": "dpo", "label": "Days payables outstanding (DPO)", "unit": "days", "category": "Efficiency",
     "label_i18n": {"zh": "应付账款周转天数", "ar": "أيام سداد الذمم الدائنة",
                    "fr": "Délai de paiement fournisseurs (jours)"},
     "num": [("bs_current_liabilities__current_trade_payables", 1)],
     "den": [(_COST_BASE, 1)],
     "formula": "Trade payables / Cost of goods sold (or total operating cost) × 365"},

    # ---- Profitability ---------------------------------------------------------------
    {"key": "net_margin", "label": "Net profit margin", "unit": "%", "category": "Profitability",
     "label_i18n": {"zh": "净利率", "ar": "هامش صافي الربح", "fr": "Marge nette"},
     "num": [("pl_profit_for_the_year", 1)], "den": [("pl_income__revenue_from_operations", 1)],
     "formula": "Profit for the year / Revenue"},
    {"key": "operating_margin", "label": "Operating margin", "unit": "%", "category": "Profitability",
     "label_i18n": {"zh": "营业利润率", "ar": "هامش التشغيل", "fr": "Marge opérationnelle"},
     "num": [("pl_operating_profit_ebit", 1)], "den": [("pl_income__revenue_from_operations", 1)],
     "formula": "Operating profit (EBIT) / Revenue"},
    {"key": "ebitda_margin", "label": "EBITDA margin", "unit": "%", "category": "Profitability",
     "label_i18n": {"zh": "EBITDA利润率", "ar": "هامش EBITDA", "fr": "Marge EBITDA"},
     "num": list(_EBITDA), "den": [("pl_income__revenue_from_operations", 1)],
     "formula": "EBITDA (EBIT + depreciation & amortisation) / Revenue"},
    {"key": "return_on_capital_employed", "label": "Return on capital employed (ROCE)", "unit": "%",
     "category": "Profitability",
     "label_i18n": {"zh": "已运用资本回报率", "ar": "العائد على رأس المال المستخدم",
                    "fr": "Rentabilité des capitaux engagés"},
     "num": [("pl_operating_profit_ebit", 1)],
     "den": [("bs_total_assets", 1), (_TCL, -1, "opt")],
     "formula": "Operating profit (EBIT) / (Total assets − Total current liabilities)"},
    {"key": "return_on_equity", "label": "Return on equity", "unit": "%", "category": "Profitability",
     "label_i18n": {"zh": "净资产收益率", "ar": "العائد على حقوق الملكية", "fr": "Rentabilité des capitaux propres"},
     "num": [("pl_profit_for_the_year", 1)], "den": [(_EQUITY, 1)],
     "formula": "Profit for the year / Total equity"},
    {"key": "return_on_assets", "label": "Return on assets", "unit": "%", "category": "Profitability",
     "label_i18n": {"zh": "总资产收益率", "ar": "العائد على الأصول", "fr": "Rentabilité des actifs"},
     "num": [("pl_profit_for_the_year", 1)], "den": [("bs_total_assets", 1)],
     "formula": "Profit for the year / Total assets"},
]


def _num(v) -> float | None:
    if v is None:
        return None
    try:
        return float(str(v).replace(",", ""))
    except ValueError:
        return None


def _group_by_key(rows: list[dict]) -> dict[str, list[dict]]:
    """Extracted rows grouped by the concept they map to — several printed lines can share one."""
    out: dict[str, list[dict]] = {}
    for r in rows:
        k = r.get("canonical_key")
        if k:
            out.setdefault(k, []).append(r)
    return out


def _value(by_key: dict, key: str, basis: str, period: str) -> float | None:
    """One concept's figure, read the same way the statement grid reads it — summed across every
    printed line that maps to the concept, or the analyst's manual value where one replaced it.
    A ratio computed off a different number than the statement shows is not a ratio of that
    statement."""
    group = by_key.get(key)
    if not group:
        return None
    return concept_value(group if isinstance(group, list) else [group], basis, period)


def _term_value(by_key, keyspec, basis, period) -> float | None:
    """Resolve a term's value. ``keyspec`` is a single canonical key, or a tuple/list of
    candidate keys tried in order (first one present wins) — used for a graceful fallback,
    e.g. cost of goods sold → total operating cost when a filing doesn't break out COGS."""
    if isinstance(keyspec, (tuple, list)):
        for k in keyspec:
            v = _value(by_key, k, basis, period)
            if v is not None:
                return v
        return None
    return _value(by_key, keyspec, basis, period)


def _side(by_key, terms, basis, period) -> float | None:
    """Sum one side of a ratio. Each term is ``(key, sign)`` — required, missing => whole side
    None — or ``(key, sign, "opt")`` — optional component, missing => treated as 0 (so an
    aggregate like 'total debt' still computes when a company reports only some components).
    A term's key may itself be a tuple of fallback candidates (see ``_term_value``). The side
    is None if no term contributed a value at all (nothing to measure)."""
    total = 0.0
    present = False
    for term in terms:
        key, sign = term[0], term[1]
        mode = term[2] if len(term) > 2 else "req"
        val = _term_value(by_key, key, basis, period)
        if val is None:
            if mode == "opt":
                continue                     # optional component absent → contributes 0
            return None                      # a required input is missing → side unavailable
        total += sign * val
        present = True
    return total if present else None


def _term_label(by_key, key) -> str:
    """A readable name for a ratio input: the caption the document printed for it when the line
    was extracted, else the canonical key made readable."""
    keys = list(key) if isinstance(key, (tuple, list)) else [key]
    for k in keys:
        group = by_key.get(k)
        if group:
            rows = group if isinstance(group, list) else [group]
            lbl = next((r.get("source_label") for r in rows if r.get("source_label")), None)
            if lbl:
                return lbl
    tail = str(keys[0]).split("__")[-1]
    return tail.replace("_", " ").capitalize()


def _resolved_key(by_key, keyspec, basis, period) -> str:
    """Which canonical key a term actually resolved to.

    A term may name several candidates tried in order (cost of goods sold → total operating cost),
    so the key that carried the figure is the one the analyst has to be able to click through to.
    Falls back to the first candidate when none had a value, so the row still names what it wanted.
    """
    keys = list(keyspec) if isinstance(keyspec, (tuple, list)) else [keyspec]
    for k in keys:
        if _value(by_key, k, basis, period) is not None:
            return str(k)
    return str(keys[0]) if keys else ""


def _side_inputs(by_key, terms, basis, period) -> list[dict]:
    """Every input that went into one side of a ratio, with the value actually used.

    A ratio the analyst cannot take apart is a number to be taken on trust. Listing the inputs —
    each with its own canonical key, its sign and its figure — makes the arithmetic checkable
    against the statement, and an absent input visibly absent rather than silently zero.
    """
    out = []
    for term in terms:
        key, sign = term[0], term[1]
        mode = term[2] if len(term) > 2 else "req"
        val = _term_value(by_key, key, basis, period)
        out.append({
            "canonical_key": _resolved_key(by_key, key, basis, period),
            "label": _term_label(by_key, key), "sign": sign, "value": val,
            "optional": mode == "opt",
        })
    return out


# unit → (multiplier applied to num/den, suffix formatter)
_UNIT_SCALE = {"x": 1, "%": 100, "days": 365}


def _display(value: float | None, unit: str) -> str:
    if value is None:
        return "—"
    if unit == "%":
        return f"{value}%"
    if unit == "days":
        return f"{value} days"
    return f"{value}×"


# Order categories are shown in (credit-analyst first): liquidity, leverage, coverage, then
# efficiency and profitability.
_CATEGORY_ORDER = ["Liquidity", "Leverage", "Coverage", "Efficiency", "Profitability"]


def compute_ratios(rows: list[dict], *, basis: str = "consolidated", period: str = "current",
                   locale: str = "en") -> list[dict]:
    """Compute the ratio catalog from the extracted values. Ratios missing an input are
    returned as unavailable (never fabricated), so the UI/export can show the full set,
    grouped by category (liquidity / leverage / coverage / efficiency / profitability)."""
    by_key = _group_by_key(rows)
    computed: dict[str, float | None] = {}
    out: list[dict] = []
    for d in _RATIOS:
        label = (d.get("label_i18n", {}).get(locale) if locale != "en" else None) or d["label"]
        num = _side(by_key, d["num"], basis, period)
        den = _side(by_key, d["den"], basis, period)
        unit = d["unit"]
        available = num is not None and den not in (None, 0)
        value = round((num / den) * _UNIT_SCALE[unit], 2) if available else None
        computed[d["key"]] = value
        out.append({
            "key": d["key"], "label": label, "category": d.get("category", "Profitability"),
            "unit": unit, "formula": d["formula"], "value": value,
            "display": _display(value, unit), "available": available,
            # The arithmetic, openable: which extracted figures were used, and with what sign.
            "inputs": {"numerator": _side_inputs(by_key, d["num"], basis, period),
                       "denominator": _side_inputs(by_key, d["den"], basis, period)},
        })

    # Cash conversion cycle is a combination of the day-based ratios (DSO + DIO − DPO), so it is
    # derived after the loop from their computed values rather than from raw keys.
    dso, dio, dpo = computed.get("dso"), computed.get("dio"), computed.get("dpo")
    if dso is not None and dio is not None and dpo is not None:
        ccc = round(dso + dio - dpo, 2)
        out.append({
            "key": "cash_conversion_cycle",
            "label": ({"zh": "现金转换周期", "ar": "دورة التحويل النقدي",
                       "fr": "Cycle de conversion de trésorerie"}.get(locale) if locale != "en" else None)
                     or "Cash conversion cycle",
            "category": "Efficiency", "unit": "days",
            "formula": "DSO + DIO − DPO", "value": ccc, "display": _display(ccc, "days"),
            "available": True,
        })

    order = {c: i for i, c in enumerate(_CATEGORY_ORDER)}
    out.sort(key=lambda r: order.get(r["category"], len(order)))
    return out


# --- 2. Disclosures --------------------------------------------------------
_DISCLOSURES = [
    {"key": "auditor_qualification", "label": "Auditor qualification / opinion",
     "label_i18n": {"zh": "审计意见/保留意见", "ar": "تحفّظ المدقق", "fr": "Réserve de l'auditeur"},
     "patterns": [r"qualified opinion", r"adverse opinion", r"disclaimer of opinion",
                  r"emphasis of matter", r"basis for qualified"]},
    {"key": "going_concern", "label": "Going concern",
     "label_i18n": {"zh": "持续经营", "ar": "الاستمرارية", "fr": "Continuité d'exploitation"},
     "patterns": [r"going concern"]},
    {"key": "contingent_liabilities", "label": "Contingent liabilities",
     "label_i18n": {"zh": "或有负债", "ar": "الالتزامات المحتملة", "fr": "Passifs éventuels"},
     "patterns": [r"contingent liabilit", r"contingenc(y|ies)"]},
    {"key": "guarantees", "label": "Guarantees",
     "label_i18n": {"zh": "担保", "ar": "الضمانات", "fr": "Garanties"},
     "patterns": [r"\bguarantee", r"financial guarantee"]},
    {"key": "commitments", "label": "Commitments",
     "label_i18n": {"zh": "承诺事项", "ar": "الالتزامات", "fr": "Engagements"},
     "patterns": [r"capital commitment", r"\bcommitments?\b"]},
    {"key": "related_party", "label": "Related-party transactions",
     "label_i18n": {"zh": "关联方交易", "ar": "معاملات الأطراف ذات العلاقة", "fr": "Parties liées"},
     "patterns": [r"related part(y|ies)"]},
    {"key": "subsequent_events", "label": "Subsequent events",
     "label_i18n": {"zh": "期后事项", "ar": "الأحداث اللاحقة", "fr": "Événements postérieurs"},
     "patterns": [r"subsequent event", r"events after the reporting"]},
    {"key": "litigation", "label": "Litigation / legal proceedings",
     "label_i18n": {"zh": "诉讼", "ar": "التقاضي", "fr": "Litiges"},
     "patterns": [r"litigation", r"legal proceeding", r"lawsuit"]},
]


def _snippet(text: str, match: re.Match, width: int = 90) -> str:
    start = max(0, match.start() - width // 2)
    end = min(len(text), match.end() + width // 2)
    return re.sub(r"\s+", " ", text[start:end]).strip()


def scan_disclosures(pages: list[tuple[int, str]], locale: str = "en") -> list[dict]:
    """Scan page texts for the disclosure catalog. Returns one entry per catalog item with
    whether it was found, the page, and a surrounding snippet — a presence check, not a full
    parse (honest about what a generic scan can claim)."""
    out: list[dict] = []
    for d in _DISCLOSURES:
        label = (d["label_i18n"].get(locale) if locale != "en" else None) or d["label"]
        hit = None
        for page_index, text in pages:
            low = text.lower()
            for pat in d["patterns"]:
                m = re.search(pat, low)
                if m:
                    hit = {"page": page_index + 1, "snippet": _snippet(text, m)}
                    break
            if hit:
                break
        out.append({"key": d["key"], "label": label, "present": hit is not None,
                    "page": hit["page"] if hit else None, "snippet": hit["snippet"] if hit else ""})
    return out


def localize_disclosures(disclosures: list[dict], locale: str = "en") -> list[dict]:
    """Re-label stored disclosure entries in the output locale (they are scanned/stored in
    English). Keeps present/page/snippet; only the label is localized, by key."""
    if locale == "en":
        return disclosures
    labels = {d["key"]: d.get("label_i18n", {}) for d in _DISCLOSURES}
    out = []
    for d in disclosures:
        loc = labels.get(d.get("key"), {}).get(locale)
        out.append({**d, "label": loc or d.get("label")})
    return out


# --- 3. Credit analysis (numeric factors + report-narrative signals) --------
# A detailed credit view built from BOTH the extracted values (ratios) and the annual
# report's narrative (the stored disclosure scan) — not the numbers alone. Deterministic and
# transparent: every factor cites its ratio and threshold; every flag cites its report page.

_STANCE_LABEL = {
    "strong": {"en": "Strong", "zh": "强", "ar": "قوي", "fr": "Solide"},
    "adequate": {"en": "Adequate", "zh": "适中", "ar": "ملائم", "fr": "Adéquat"},
    "weak": {"en": "Weak / speculative", "zh": "偏弱 / 投机级", "ar": "ضعيف / مضاربي",
             "fr": "Faible / spéculatif"},
    "insufficient": {"en": "Insufficient data", "zh": "数据不足", "ar": "بيانات غير كافية",
                     "fr": "Données insuffisantes"},
}
_TONE_LABEL = {
    "strong": {"en": "Strong", "zh": "强", "ar": "قوي", "fr": "Solide"},
    "adequate": {"en": "Adequate", "zh": "适中", "ar": "ملائم", "fr": "Adéquat"},
    "weak": {"en": "Weak", "zh": "弱", "ar": "ضعيف", "fr": "Faible"},
}
_CREDIT_CAT_LABEL = {
    "Leverage": {"en": "Leverage", "zh": "杠杆", "ar": "الرافعة المالية", "fr": "Levier"},
    "Coverage": {"en": "Coverage", "zh": "偿债保障", "ar": "التغطية", "fr": "Couverture"},
    "Liquidity": {"en": "Liquidity", "zh": "流动性", "ar": "السيولة", "fr": "Liquidité"},
    "Profitability": {"en": "Profitability", "zh": "盈利能力", "ar": "الربحية", "fr": "Rentabilité"},
}

# Per category, an ordered preference of ratio keys with (higher_is_better, strong, weak)
# thresholds. The first AVAILABLE ratio becomes that category's factor.
_CREDIT_FACTORS: dict[str, list[tuple]] = {
    "Leverage": [("debt_to_equity", False, 1.0, 2.0), ("net_debt_to_ebitda", False, 2.0, 4.0),
                 ("debt_ratio", False, 45.0, 65.0)],
    "Coverage": [("interest_coverage", True, 4.0, 1.5), ("ebitda_interest_coverage", True, 5.0, 2.0),
                 ("ffo_to_total_debt", True, 30.0, 12.0), ("cfo_to_total_debt", True, 25.0, 10.0)],
    "Liquidity": [("current_ratio", True, 1.5, 1.0), ("quick_ratio", True, 1.0, 0.7)],
    "Profitability": [("net_margin", True, 8.0, 2.0), ("operating_margin", True, 10.0, 3.0),
                      ("return_on_capital_employed", True, 12.0, 4.0)],
}
_CREDIT_CAT_SEQUENCE = ["Leverage", "Coverage", "Liquidity", "Profitability"]

# Narrative signals (keyed to the disclosure scan) that bear on credit: (severity, implication).
_CREDIT_NARRATIVE: dict[str, tuple[str, dict]] = {
    "going_concern": ("severe", {
        "en": "Going-concern language present — a material threat to creditworthiness.",
        "zh": "存在持续经营相关表述——对信用状况构成重大威胁。",
        "ar": "توجد إشارة إلى الاستمرارية — تهديد جوهري للجدارة الائتمانية.",
        "fr": "Mention de continuité d'exploitation — menace importante pour la solvabilité."}),
    "auditor_qualification": ("high", {
        "en": "Modified / qualified audit opinion — reduces reliance on the reported figures.",
        "zh": "审计意见被修正/保留——降低对所报告数字的可依赖程度。",
        "ar": "رأي تدقيق معدَّل/متحفَّظ — يقلل الاعتماد على الأرقام المُبلَّغ عنها.",
        "fr": "Opinion d'audit modifiée / avec réserve — réduit la fiabilité des chiffres publiés."}),
    "contingent_liabilities": ("watch", {
        "en": "Contingent liabilities disclosed — potential off-balance-sheet claims on cash.",
        "zh": "披露或有负债——可能存在表外的现金索求。",
        "ar": "الإفصاح عن التزامات محتملة — مطالبات محتملة خارج الميزانية على النقد.",
        "fr": "Passifs éventuels divulgués — créances potentielles hors bilan sur la trésorerie."}),
    "guarantees": ("watch", {
        "en": "Guarantees given — contingent exposure beyond recorded debt.",
        "zh": "提供担保——超出已入账债务的或有敞口。",
        "ar": "ضمانات ممنوحة — تعرّض محتمل يتجاوز الدين المسجَّل.",
        "fr": "Garanties accordées — exposition éventuelle au-delà de la dette comptabilisée."}),
    "litigation": ("watch", {
        "en": "Litigation / legal proceedings disclosed — potential financial impact.",
        "zh": "披露诉讼/法律程序——可能带来财务影响。",
        "ar": "الإفصاح عن تقاضٍ / إجراءات قانونية — أثر مالي محتمل.",
        "fr": "Litiges / procédures judiciaires divulgués — impact financier potentiel."}),
}
_SEVERITY_WEIGHT = {"severe": 3, "high": 2, "watch": 1}

_CREDIT_SUMMARY = {
    "en": "{stance} credit profile from {n} computed factor(s): {pos} supportive, {neg} constraining.",
    "zh": "基于 {n} 项计算因子的信用状况为「{stance}」：{pos} 项支撑，{neg} 项制约。",
    "ar": "ملف ائتماني «{stance}» بناءً على {n} عامل محسوب: {pos} داعمة، {neg} مُقيِّدة.",
    "fr": "Profil de crédit « {stance} » à partir de {n} facteur(s) calculé(s) : {pos} favorables, {neg} contraignants.",
}
_CREDIT_SUMMARY_FLAGS = {
    "en": " Report signals to review: {list}.",
    "zh": " 需关注的年报信号：{list}。",
    "ar": " إشارات التقرير التي يجب مراجعتها: {list}.",
    "fr": " Signaux du rapport à examiner : {list}.",
}
_CREDIT_SUMMARY_NODATA = {
    "en": "Insufficient extracted values to compute a credit view.",
    "zh": "提取的数值不足，无法计算信用视图。",
    "ar": "القيم المستخرجة غير كافية لحساب رؤية ائتمانية.",
    "fr": "Valeurs extraites insuffisantes pour calculer une vue de crédit.",
}


def _loc_map(table: dict, key: str, locale: str) -> str:
    entry = table.get(key, {})
    return entry.get(locale) or entry.get("en") or key


def build_credit_analysis(rows: list[dict], disclosures: list[dict] | None = None, *,
                          basis: str = "consolidated", locale: str = "en") -> dict:
    """Detailed credit assessment from the extracted values PLUS the report narrative.

    Numeric factors (leverage / coverage / liquidity / profitability) are bucketed against
    credit thresholds; narrative flags come from the annual report's disclosure scan (going
    concern, qualified opinion, contingents, guarantees, litigation). The overall stance
    blends both — a going-concern or qualified-opinion signal caps an otherwise-strong read.
    """
    ratios = {r["key"]: r for r in compute_ratios(rows, basis=basis)}
    factors: list[dict] = []
    tone_score = 0
    pos = neg = 0
    for cat in _CREDIT_CAT_SEQUENCE:
        for key, hib, strong, weak in _CREDIT_FACTORS[cat]:
            r = ratios.get(key)
            if not (r and r.get("available") and r.get("value") is not None):
                continue
            v = r["value"]
            if hib:
                tone = "strong" if v >= strong else "weak" if v < weak else "adequate"
            else:
                tone = "strong" if v <= strong else "weak" if v > weak else "adequate"
            tone_score += 1 if tone == "strong" else -1 if tone == "weak" else 0
            pos += tone == "strong"
            neg += tone == "weak"
            factors.append({
                "category": _loc_map(_CREDIT_CAT_LABEL, cat, locale),
                "category_key": cat, "key": key, "label": r["label"],
                "value": r["value"], "display": r["display"], "unit": r["unit"],
                "tone": tone, "tone_label": _loc_map(_TONE_LABEL, tone, locale),
            })
            break  # first available ratio per category

    present = {d.get("key") for d in (disclosures or []) if d.get("present")}
    dmap = {d.get("key"): d for d in (disclosures or [])}
    flags: list[dict] = []
    penalty = 0
    for key, (sev, impl) in _CREDIT_NARRATIVE.items():
        if key not in present:
            continue
        d = dmap.get(key, {})
        penalty += _SEVERITY_WEIGHT.get(sev, 0)
        flags.append({
            "key": key, "label": d.get("label") or key, "severity": sev,
            "implication": impl.get(locale) or impl["en"],
            "page": d.get("page"), "snippet": d.get("snippet", ""),
        })

    n = len(factors)
    if n == 0:
        stance = "insufficient"
    else:
        avg = tone_score / n
        stance = "strong" if avg >= 0.5 else "adequate" if avg >= -0.25 else "weak"
        if "going_concern" in present:
            stance = "weak"
        elif "auditor_qualification" in present and stance == "strong":
            stance = "adequate"
        elif penalty >= 3 and stance == "strong":
            stance = "adequate"

    stance_label = _loc_map(_STANCE_LABEL, stance, locale)
    if stance == "insufficient":
        summary = _CREDIT_SUMMARY_NODATA.get(locale) or _CREDIT_SUMMARY_NODATA["en"]
    else:
        summary = (_CREDIT_SUMMARY.get(locale) or _CREDIT_SUMMARY["en"]).format(
            stance=stance_label, n=n, pos=pos, neg=neg)
        if flags:
            names = ", ".join(f["label"] for f in flags)
            summary += (_CREDIT_SUMMARY_FLAGS.get(locale) or _CREDIT_SUMMARY_FLAGS["en"]).format(list=names)

    return {"stance": stance, "stance_label": stance_label, "factors": factors,
            "flags": flags, "summary": summary, "basis": basis}


# Company-name suffixes across the seed jurisdictions (HK/China, India, UK, US, EU, …).
_ENTITY_SUFFIX = re.compile(
    r"\b("
    r"limited|ltd\.?|public limited company|plc|p\.l\.c\.|incorporated|inc\.?|"
    r"corporation|corp\.?|company|co\.?|l\.?l\.?c\.?|l\.?l\.?p\.?|holdings?|group|"
    r"berhad|bhd\.?|n\.?v\.?|s\.?a\.?|s\.?p\.?a\.?|gmbh|a\.?g\.?|pte\.?|sdn"
    r")\b", re.I)
_STATEMENTish = re.compile(
    r"balance sheet|statement of|profit (and|or) loss|cash flow|comprehensive income|"
    r"annual report|financial statements|notes to|independent auditor",
    re.I)


def _entity_from_segment(seg: str) -> str | None:
    """A single candidate segment → the entity name if it looks like one, else None."""
    line = re.sub(r"\s+", " ", seg).strip(" .-—·|")
    if not (3 <= len(line) <= 90):
        return None
    if _STATEMENTish.search(line):
        return None                               # a statement/section/running-header phrase
    if _ENTITY_SUFFIX.search(line) and re.search(r"[A-Za-z]", line) \
            and sum(c.isdigit() for c in line) <= 4:
        return line
    return None


def detect_entity_name(pages: list[tuple[int, str]]) -> str | None:
    """Best-effort entity name from the document's opening pages: the first prominent line (or
    slash/pipe-separated segment of one) that carries a company-name suffix (Ltd / PLC / Inc / …)
    and isn't itself a statement title or running-header phrase. Real HK/PRC filings print a
    running header like 'ACME Holdings Limited / Annual Report 2024' — the entity is the segment
    before the slash. Deterministic and honest: returns None when nothing convincing is found."""
    for _idx, text in pages[:5]:
        for raw in text.splitlines():
            # Try the whole line first, then each '/'- or '|'-separated segment (running headers
            # glue the company name to 'Annual Report YYYY', which we must split off).
            candidates = [raw, *re.split(r"[/|]", raw)]
            for cand in candidates:
                got = _entity_from_segment(cand)
                if got:
                    return got
    return None


def document_text(data: bytes, fmt: str) -> list[tuple[int, str]]:
    """Per-page (or per-sheet) plain text for the disclosure scan."""
    if fmt == "pdf":
        try:
            import fitz
        except ImportError:  # pragma: no cover
            return []
        pdf = fitz.open(stream=data, filetype="pdf")
        return [(i, pdf[i].get_text("text") or "") for i in range(pdf.page_count)]
    if fmt in ("xlsx", "xls"):
        try:
            import io
            import openpyxl
        except ImportError:  # pragma: no cover
            return []
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        pages = []
        for i, name in enumerate(wb.sheetnames):
            cells = [str(c) for row in wb[name].iter_rows(values_only=True) for c in row if c is not None]
            pages.append((i, " ".join(cells)))
        wb.close()
        return pages
    return []


# --- 3. Free-format notes --------------------------------------------------
_NOTE_LINES = [
    ("bs_current_assets__trade_receivables", "Trade receivables",
     {"zh": "应收账款", "ar": "الذمم المدينة", "fr": "Créances clients"}),
    ("bs_current_assets__cash_and_cash_equivalents", "Cash and cash equivalents",
     {"zh": "现金及现金等价物", "ar": "النقد وما في حكمه", "fr": "Trésorerie et équivalents"}),
    ("bs_current_assets__inventories", "Inventories",
     {"zh": "存货", "ar": "المخزون", "fr": "Stocks"}),
    ("pl_income__revenue_from_operations", "Revenue",
     {"zh": "收入", "ar": "الإيرادات", "fr": "Produits"}),
    ("pl_profit_for_the_year", "Profit for the year",
     {"zh": "年度利润", "ar": "ربح السنة", "fr": "Résultat de l'exercice"}),
]

# Localized sentence templates for the generated highlights. {label},{dir},{d},{cur},{prior}.
_NOTE_TR = {
    "move": {
        "en": "{label} {dir} {d:.1f}% to {cur} (prior period {prior}).",
        "zh": "{label}{dir} {d:.1f}% 至 {cur}（上期 {prior}）。",
        "ar": "{label} {dir} بنسبة {d:.1f}% إلى {cur} (الفترة السابقة {prior}).",
        "fr": "{label} {dir} de {d:.1f} % à {cur} (période précédente {prior}).",
    },
    "up": {"en": "increased", "zh": "增长", "ar": "ارتفع", "fr": "en hausse"},
    "down": {"en": "decreased", "zh": "下降", "ar": "انخفض", "fr": "en baisse"},
    "flat": {
        "en": "{label} was {cur} for the current period.",
        "zh": "{label}本期为 {cur}。",
        "ar": "{label} بلغ {cur} للفترة الحالية.",
        "fr": "{label} s'élève à {cur} pour la période en cours.",
    },
    "liquidity": {"en": "Liquidity", "zh": "流动性", "ar": "السيولة", "fr": "Liquidité"},
    "liquidity_txt": {
        "en": "Current ratio of {v} indicates {stance} short-term liquidity.",
        "zh": "流动比率为 {v}，表明短期流动性{stance}。",
        "ar": "نسبة التداول {v} تشير إلى سيولة قصيرة الأجل {stance}.",
        "fr": "Un ratio de liquidité de {v} indique une liquidité à court terme {stance}.",
    },
    "profitability": {"en": "Profitability", "zh": "盈利能力", "ar": "الربحية", "fr": "Rentabilité"},
    "profitability_txt": {
        "en": "Net profit margin was {v} of revenue.",
        "zh": "净利率为收入的 {v}。",
        "ar": "بلغ هامش صافي الربح {v} من الإيرادات.",
        "fr": "La marge nette représente {v} du chiffre d'affaires.",
    },
    "stance_comfortable": {"en": "comfortable", "zh": "充裕", "ar": "مريحة", "fr": "confortable"},
    "stance_adequate": {"en": "adequate", "zh": "尚可", "ar": "كافية", "fr": "adéquate"},
    "stance_tight": {"en": "tight", "zh": "紧张", "ar": "متوترة", "fr": "tendue"},
}


def _tr(key: str, locale: str) -> str:
    entry = _NOTE_TR[key]
    return entry.get(locale, entry["en"])


def _fmt(n: float | None) -> str:
    return "—" if n is None else f"{n:,.0f}"


def build_free_notes(rows: list[dict], *, basis: str = "consolidated", locale: str = "en") -> list[dict]:
    """Plain-language notes generated strictly from the extracted numbers: period movements
    for headline lines, and a one-line read on liquidity/profitability from the ratios."""
    by_key = _group_by_key(rows)
    notes: list[dict] = []

    for key, label_en, label_i18n in _NOTE_LINES:
        cur = _value(by_key, key, basis, "current")
        if cur is None:
            continue
        label = (label_i18n.get(locale) if locale != "en" else None) or label_en
        prior = _value(by_key, key, basis, "prior")
        if prior not in (None, 0):
            delta = (cur - prior) / abs(prior) * 100
            direction = _tr("up" if delta >= 0 else "down", locale)
            notes.append({
                "title": label,
                "text": _tr("move", locale).format(label=label, dir=direction, d=abs(delta),
                                                    cur=_fmt(cur), prior=_fmt(prior)),
            })
        else:
            notes.append({"title": label,
                          "text": _tr("flat", locale).format(label=label, cur=_fmt(cur))})

    ratios = {r["key"]: r for r in compute_ratios(rows, basis=basis)}
    cr = ratios.get("current_ratio")
    if cr and cr["available"]:
        stance = _tr("stance_comfortable" if cr["value"] >= 1.5
                     else "stance_adequate" if cr["value"] >= 1 else "stance_tight", locale)
        notes.append({"title": _tr("liquidity", locale),
                      "text": _tr("liquidity_txt", locale).format(v=cr["display"], stance=stance)})
    nm = ratios.get("net_margin")
    if nm and nm["available"]:
        notes.append({"title": _tr("profitability", locale),
                      "text": _tr("profitability_txt", locale).format(v=nm["display"])})

    if not notes:
        notes.append({"title": "Summary",
                      "text": "Not enough headline values were extracted to generate movement notes."})
    return notes
