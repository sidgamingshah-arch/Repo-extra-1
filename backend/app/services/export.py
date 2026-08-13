"""Export renderers — formatted Excel (.xlsx) and JSON.

Excel is produced with openpyxl: a statement-shaped workbook (one sheet per statement in
the template, with sections / subtotals / totals, consolidated + standalone columns, a
confidence column and note references) plus Note details / Ratios / Disclosures sheets.
Values are written as literals with their source cell/page recorded; subtotal and total
rows are styled and, for the demo workbook, carry the template's figures.
"""
from __future__ import annotations

import io
import json

from app.sample.demo import CONF_PCT
from app.services.periods import concept_value, split_current_prior

# Statement titles for the formatted export, localized like the rest of the app. Line-item
# labels come from the TEMPLATE's label_i18n, so the export is fully template-driven.
_STMT_TITLE = {
    "balance_sheet": {"en": "Balance Sheet", "zh": "资产负债表", "ar": "الميزانية العمومية",
                      "fr": "Bilan"},
    "profit_and_loss": {"en": "Profit & Loss", "zh": "利润表", "ar": "الأرباح والخسائر",
                        "fr": "Compte de résultat"},
    "cash_flow": {"en": "Cash Flow", "zh": "现金流量表", "ar": "التدفقات النقدية",
                  "fr": "Flux de trésorerie"},
}
_STMT_SHEET = {"balance_sheet": "Balance Sheet", "profit_and_loss": "Profit & Loss",
               "cash_flow": "Cash Flow"}
_BASIS_LABEL = {
    "consolidated": {"en": "Consolidated", "zh": "合并", "ar": "موحّد", "fr": "Consolidé"},
    "standalone": {"en": "Standalone", "zh": "单独", "ar": "مستقل", "fr": "Individuel"},
}
_COL = {
    "Line item": {"zh": "项目", "ar": "البند", "fr": "Poste"},
    "Note": {"zh": "附注", "ar": "إيضاح", "fr": "Note"},
    "Current": {"zh": "本期", "ar": "الحالية", "fr": "Actuel"},
    "Prior": {"zh": "上期", "ar": "السابقة", "fr": "Précédent"},
    "Conf": {"zh": "置信度", "ar": "الثقة", "fr": "Conf."},
    "Source": {"zh": "来源", "ar": "المصدر", "fr": "Source"},
}


def _col(term: str, locale: str) -> str:
    return term if locale == "en" else _COL.get(term, {}).get(locale, term)


# The validation banner's vocabulary. A sheet that says nothing was checked has to say it in the
# reader's language, or the one warning on the workbook is the one line they cannot read.
_EXPORT_TR = {
    "Validation status is unknown for this export.": {
        "zh": "本次导出的校验状态未知。", "ar": "حالة التحقق غير معروفة لهذا التصدير.",
        "fr": "Le statut de validation est inconnu pour cet export."},
    "No template relations were checked on this filing.": {
        "zh": "本报表未执行任何模板关系校验。",
        "ar": "لم يتم التحقق من أي علاقات في القالب لهذا الملف.",
        "fr": "Aucune relation du modèle n'a été vérifiée sur ce dépôt."},
    "A blocking rule could not be enforced on this filing.": {
        "zh": "本报表存在无法执行的阻断性规则。",
        "ar": "تعذّر تطبيق قاعدة مانعة على هذا الملف.",
        "fr": "Une règle bloquante n'a pu être appliquée à ce dépôt."},
    "The template declares no relations for this filing, so none was checked.": {
        "zh": "模板未为本报表声明任何关系，因此未做校验。",
        "ar": "لا يعلن القالب أي علاقات لهذا الملف، فلم يتم التحقق من شيء.",
        "fr": "Le modèle ne déclare aucune relation pour ce dépôt : rien n'a été vérifié."},
    "No template relation could be evaluated — nothing here is validated.": {
        "zh": "没有任何模板关系可被评估——此处内容均未经校验。",
        "ar": "لم يكن بالإمكان تقييم أي علاقة في القالب — لا شيء هنا مُتحقَّق منه.",
        "fr": "Aucune relation du modèle n'a pu être évaluée — rien ici n'est validé."},
    "of": {"zh": "/", "ar": "من", "fr": "sur"},
    "relations checked did not hold; see the review queue.": {
        "zh": "项已校验的关系不成立；请见审核队列。",
        "ar": "علاقة تم التحقق منها لم تتحقق؛ راجع قائمة المراجعة.",
        "fr": "relations vérifiées ne tiennent pas ; voir la file de revue."},
    "relations could be checked; the rest were not evaluable.": {
        "zh": "项关系可被校验；其余无法评估。",
        "ar": "علاقة أمكن التحقق منها؛ الباقي غير قابل للتقييم.",
        "fr": "relations ont pu être vérifiées ; les autres n'étaient pas évaluables."},
    "relations checked held.": {
        "zh": "项已校验的关系成立。", "ar": "علاقة تم التحقق منها وتحققت.",
        "fr": "relations vérifiées tiennent."},
    "Validation": {"zh": "校验", "ar": "التحقق", "fr": "Validation"},
}


def _prov_str(prov: dict | None) -> str:
    if not prov:
        return ""
    if prov.get("source_kind") == "spreadsheet" and prov.get("sheet"):
        return f"{prov['sheet']}!{prov.get('cell', '')}"
    return f"p.{(prov.get('page_index', 0) or 0) + 1}"


def _netting_block(rows: list[dict], netting_rules: list | None) -> list[dict]:
    if not netting_rules:
        return []
    from app.services.netting import compute_netting

    cur = compute_netting(rows, netting_rules, basis="consolidated", period="current")
    prior = compute_netting(rows, netting_rules, basis="consolidated", period="prior")
    out = []
    for key, c in cur.items():
        out.append({**c, "prior_net": prior.get(key, {}).get("net")})
    return out


def build_rows_json(rows: list[dict], *, filename: str, disclosures: list[dict] | None = None,
                    note_details: list[dict] | None = None, reconciliation: list[dict] | None = None,
                    locale: str = "en", credit_narrative: dict | None = None,
                    netting_rules: list | None = None, coverage: dict | None = None) -> bytes:
    """JSON export of a REAL extraction: every line item with its mapping, confidence, any
    edited formula, and the exact source location of each value (sheet/cell or page/bbox),
    plus a derived-analysis block (ratios / disclosures / credit) and the note detail +
    reconciliation — so the JSON carries the same information the UI and the Excel show."""
    from app.services.derived import build_credit_analysis, compute_ratios, localize_disclosures

    disc = localize_disclosures(disclosures or [], locale)
    credit = build_credit_analysis(rows, disc, locale=locale)
    if credit_narrative and credit_narrative.get("text"):
        credit = {**credit, "narrative": credit_narrative}

    payload = {
        "source_document": filename,
        "line_item_count": len(rows),
        "line_items": [
            {
                "source_label": r.get("source_label"),
                "canonical_key": r.get("canonical_key"),
                "note": r.get("note"),
                "mapping_method": r.get("mapping_method"),
                "mapping_confidence": r.get("mapping_confidence"),
                "flags": r.get("flags") or [],
                "formula": r.get("formula"),
                "edited": bool(r.get("edited")),
                "values": [
                    {"period": v.get("period_label"), "basis": v.get("basis"),
                     "value": v.get("value"), "source": _prov_str(v.get("provenance"))}
                    for v in (r.get("values") or [])
                ],
            }
            for r in rows
        ],
        "analysis": {
            "ratios": [{"key": x["key"], "label": x["label"], "category": x.get("category"),
                        "value": x["value"], "display": x["display"], "available": x["available"]}
                       for x in compute_ratios(rows, locale=locale)],
            "disclosures": disc,
            "credit": credit,
            # Face-line containment netting (target line net of contained lines) + the formula.
            "netting": _netting_block(rows, netting_rules),
        },
        "note_details": note_details or [],
        "reconciliation": reconciliation or [],
        # What these figures were verified against, verbatim from the report the review screen is
        # served — a machine-readable consumer must be able to tell a validated extraction from one
        # where no relation could be evaluated, which it could not before.
        "validation": {
            "coverage": coverage,
            "caption": (lambda b: None if b is None else {"text": b[0], "severe": b[1]})(
                validation_caption(coverage, locale)),
        },
    }
    return json.dumps(payload, indent=2, ensure_ascii=False).encode("utf-8")


_TARGET_UNIT_SCALE = {"absolute": 1.0, "thousands": 1e3, "lakh": 1e5, "lakhs": 1e5,
                      "millions": 1e6, "million": 1e6, "crore": 1e7, "crores": 1e7,
                      "billions": 1e9, "billion": 1e9}


def units_scale(source_units: dict | None, target: str | None) -> tuple[float, str | None]:
    """Factor to present source-unit figures in `target` units, and the label to show.

    Only converts when the source scale was actually detected AND a target is requested — we
    never guess a scale for an undeclared document (that would silently corrupt figures)."""
    label = (source_units or {}).get("units_label")
    if not source_units or not target:
        return 1.0, label
    try:
        src = float(source_units.get("scale_factor") or 1.0)
    except (TypeError, ValueError):
        src = 1.0
    tgt = _TARGET_UNIT_SCALE.get(target.lower(), 1.0)
    return (src / tgt if tgt else 1.0), target


def build_rows_xlsx(rows: list[dict], *, filename: str, scale: float = 1.0) -> bytes:
    """Excel export of a REAL extraction — one row per extracted line item, with mapping,
    confidence, and the source cell/page of the first value for traceability."""
    import openpyxl

    def _sc(v):
        n = _num(v)
        return round(n * scale) if (n is not None and scale != 1.0) else v

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extraction"
    headers = ["#", "Line item", "Note", "Mapped to", "Method", "Confidence",
               "Current", "Prior", "Formula", "Source"]
    ws.append(headers)
    formula_col = headers.index("Formula") + 1
    for i, r in enumerate(rows, start=1):
        values = r.get("values") or []
        cur_v, prior_v = split_current_prior(values)
        current = (cur_v or {}).get("value")
        prior = (prior_v or {}).get("value")
        conf = r.get("mapping_confidence")
        ws.append([
            i,
            r.get("source_label", ""),
            r.get("note") or "",
            r.get("canonical_key") or "",
            r.get("mapping_method") or "",
            f"{round(conf * 100)}%" if isinstance(conf, (int, float)) else "",
            _sc(current), _sc(prior),
            r.get("formula") or "",                       # edited-item formula, if any
            _prov_str(values[0].get("provenance")) if values else "",
        ])
        # The formula is carried for AUDIT, as the analyst typed it — it is not a live cell.
        # openpyxl promotes any string starting with "=" to a real formula, and this one's
        # references are canonical line-item keys resolved server-side (services/formula.py), not
        # cell addresses, so Excel opened the workbook showing #NAME? exactly where an audit trail
        # was intended. Forcing the string type keeps the expression readable.
        ws.cell(row=ws.max_row, column=formula_col).data_type = "s"
    ws.freeze_panes = "A2"
    for col, width in zip("ABCDEFGHIJ", (5, 34, 8, 28, 12, 11, 14, 14, 18, 16)):
        ws.column_dimensions[col].width = width
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Formatted statement export — mirrors the TEMPLATE's structure (sections,
# subtotals, totals, ordering, localized labels) with consolidated + standalone
# side by side. Driven entirely by the template, so any template of the same
# schema produces the same high-quality output.
# ---------------------------------------------------------------------------
def _label(node: dict, locale: str) -> str:
    return (node.get("label_i18n") or {}).get(locale) or node.get("label") or node.get("canonical_key") or ""


def _cell_value(group, basis: str, period: str):
    """The figure for one concept in one (basis, period) — read through the same resolver the
    statement view and the accounting checks use, so the workbook cannot disagree with the
    screen: the SUM when several printed lines map to the concept, or the analyst's manual
    value when one replaced it."""
    if not group:
        return None
    rows = [r for r in (group if isinstance(group, list) else [group]) if r]
    return concept_value(rows, basis, period)


def _contribution_note(group: list[dict], basis: str, period: str) -> str | None:
    """The audit trail for a combined figure: every contributing caption with its own amount and
    the page it was printed on. A combined figure matches no single line in the document, so
    without this the workbook shows a number the reader cannot find anywhere."""
    if not group or len(group) <= 1:
        return None
    lines = [f"Combined from {len(group)} printed lines:"]
    for row in group:
        amount = _cell_value([row], basis, period)
        vals = row.get("values") or []
        where = _prov_str(vals[0].get("provenance")) if vals else ""
        shown = "—" if amount is None else f"{amount:,.0f}"
        lines.append(f"  {row.get('source_label') or ''} = {shown}"
                     + (f"  [{where}]" if where else ""))
    total = _cell_value(group, basis, period)
    lines.append(f"  Total = {'—' if total is None else f'{total:,.0f}'}")
    return "\n".join(lines)


def _num(s):
    if s is None:
        return None
    try:
        return float(str(s).replace(",", ""))
    except ValueError:
        return None


def _bases_present(rows: list[dict]) -> list[str]:
    found = {(v.get("basis") or "consolidated") for r in rows for v in (r.get("values") or [])}
    return [b for b in ("consolidated", "standalone") if b in found] or ["consolidated"]


def validation_caption(coverage: dict | None, locale: str = "en") -> tuple[str, bool] | None:
    """What this workbook's figures were actually verified against — or None when everything held.

    A filing whose template relations could not be evaluated used to export as a workbook
    indistinguishable from a fully validated one: the numbers look the same, the formatting is the
    same, and nothing on the sheet says the arithmetic behind them was never checked. That is the
    export equivalent of reading "3 relations passed" as "the statement is verified", which the
    coverage contract exists to prevent on screen.

    Every figure here is read from the coverage report the review screen is served — no second
    computation, so the sheet and the queue cannot disagree. `severe` marks the cases where nothing
    was verified, or a BLOCKING rule could not be enforced, as opposed to a partial check.
    """
    def L(s: str) -> str:
        return _EXPORT_TR.get(s, {}).get(locale, s)

    if coverage is None:
        # The flat row layout, or a run stored before coverage existed. Silence here would be the
        # defect: an absent report is not a clean one.
        return L("Validation status is unknown for this export."), True
    if not coverage.get("available"):
        return (str(coverage.get("reason_label")
                    or L("No template relations were checked on this filing.")), True)

    agg = coverage.get("aggregate") or {}
    status = str(agg.get("status") or "")
    passed, failed = int(agg.get("passed") or 0), int(agg.get("failed") or 0)
    evaluated, declarable = int(agg.get("evaluated") or 0), int(agg.get("declarable") or 0)
    unenforceable = [a for a in (coverage.get("alarms") or [])
                     if a.get("code") == "BLOCKING_RULE_UNENFORCEABLE"]

    parts: list[str] = []
    severe = False
    if unenforceable:
        parts.append(L("A blocking rule could not be enforced on this filing."))
        severe = True
    if status == "ABSENT" or declarable == 0:
        parts.append(L("The template declares no relations for this filing, so none was checked."))
        severe = True
    elif status == "UNVALIDATED" or evaluated == 0:
        parts.append(L("No template relation could be evaluated — nothing here is validated."))
        severe = True
    elif status == "FAILED":
        parts.append(f"{failed} {L('of')} {evaluated} "
                     f"{L('relations checked did not hold; see the review queue.')}")
        severe = True
    elif status == "PARTIAL":
        parts.append(f"{evaluated} {L('of')} {declarable} "
                     f"{L('relations could be checked; the rest were not evaluable.')}")
    elif status == "PASSED" and not unenforceable:
        return None                          # every declarable relation ran and held
    else:
        parts.append(f"{passed} {L('of')} {evaluated} {L('relations checked held.')}")
    return " ".join(parts), severe


def build_statement_workbook(rows: list[dict], template_def: dict, *, locale: str = "en",
                             filename: str = "", disclosures: list[dict] | None = None,
                             note_details: list[dict] | None = None,
                             reconciliation: list[dict] | None = None,
                             include: set[str] | None = None,
                             scale: float = 1.0, units_caption: str | None = None,
                             credit_narrative: dict | None = None,
                             coverage: dict | None = None) -> bytes:
    """A formatted, statement-shaped workbook: one sheet per statement in the template, with
    its sections / subtotals / totals, localized line labels, and consolidated + standalone
    columns side by side, plus Note details / Ratios / Disclosures sheets. Purely
    template-driven — the same code renders any template."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    # What these figures were verified against, derived once from the report the review screen is
    # served. None means every declarable relation ran and held, and the sheet needs no warning.
    banner = validation_caption(coverage, locale)

    # Grouped, not overwritten. Several printed lines legitimately share one concept — three
    # depreciation lines, two tax payments, a section's residual "Others" bucket — and keeping
    # only the last silently dropped the rest from the exported statement.
    by_key: dict[str, list[dict]] = {}
    for r in rows:
        if r.get("canonical_key"):
            by_key.setdefault(r["canonical_key"], []).append(r)
    bases = _bases_present(rows)

    # The template's CALCULATED lines, evaluated from their components — the same evaluation the
    # statement view uses, so the workbook and the screen cannot show a different subtotal. What
    # the document printed goes into the cell's comment instead of the cell: a subtotal that
    # contradicts its own components is a finding, not the figure to hand to a reader.
    from app.services.rollups import evaluate_rows

    calc = {(b, p): evaluate_rows(template_def, rows, b, p, locale)
            for b in bases for p in ("current", "prior")}

    ink = "1f2937"
    section_fill = PatternFill("solid", fgColor="EEF1F6")
    total_fill = PatternFill("solid", fgColor="E7ECF5")
    band_fill = PatternFill("solid", fgColor="243044")
    thin_top = Border(top=Side(style="thin", color="9AA4B2"))
    dbl_top = Border(top=Side(style="double", color="6B7686"))
    right = Alignment(horizontal="right")
    num_fmt = "#,##0;(#,##0)"

    # Column layout: Line item | Note | [basis × (Current, Prior)] | Conf | Source
    period_cols = [(b, p) for b in bases for p in ("current", "prior")]
    n_val = len(period_cols)
    first_val = 3                                   # col C
    conf_col = first_val + n_val
    src_col = conf_col + 1
    last_col = src_col

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    statements = [s for s in template_def.get("statements", []) if s.get("type")]

    any_sheet = False
    for stmt in statements:
        stype = stmt.get("type")
        # The FULL template is written out — every section and every line item, in template
        # order — with extracted values filled in and the rest left blank. The download is a
        # complete, template-shaped statement, not just the lines we happened to find.
        any_sheet = True
        ws = wb.create_sheet(_STMT_SHEET.get(stype, stype)[:31])

        ws.cell(1, 1, filename).font = Font(size=9, color="8A94A6")
        ws.cell(2, 1, _STMT_TITLE.get(stype, {}).get(locale, stype)).font = Font(bold=True, size=13, color=ink)
        nxt = 3
        if units_caption:
            ws.cell(nxt, 1, units_caption).font = Font(size=9, italic=True, color="6B7280")
            nxt += 1
        if banner:
            # Above the figures, not on a sheet of its own: a reader who opens the balance sheet and
            # scrolls has to pass it. A workbook whose arithmetic was never checked looked exactly
            # like one that was, and this is the only line that says otherwise.
            text, severe = banner
            cell = ws.cell(nxt, 1, f"⚠ {_EXPORT_TR.get('Validation', {}).get(locale, 'Validation')}"
                                   f": {text}")
            cell.font = Font(bold=True, size=9, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="B42318" if severe else "B54708")
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            nxt += 1

        # Header band (basis groups, then column names).
        hb, hc = nxt, nxt + 1
        ws.cell(hb, 1, filename and "" or "")
        for idx, b in enumerate(bases):
            c0 = first_val + idx * 2
            cell = ws.cell(hb, c0, _BASIS_LABEL[b][locale] if b in _BASIS_LABEL else b)
            cell.font = Font(bold=True, color="FFFFFF"); cell.alignment = Alignment(horizontal="center")
            cell.fill = band_fill
            ws.merge_cells(start_row=hb, start_column=c0, end_row=hb, end_column=c0 + 1)
            ws.cell(hb, c0 + 1).fill = band_fill
        for col, name in [(1, "Line item"), (2, "Note"), (conf_col, "Conf"), (src_col, "Source")]:
            hcell = ws.cell(hc, col, _col(name, locale)); hcell.font = Font(bold=True, size=9, color="5A6472")
        for (b, p) in period_cols:
            ci = first_val + period_cols.index((b, p))
            hcell = ws.cell(hc, ci, _col("Current" if p == "current" else "Prior", locale))
            hcell.font = Font(bold=True, size=9, color="5A6472"); hcell.alignment = right

        r = hc + 1
        r = _emit_nodes(ws, stmt.get("sections", []), by_key, period_cols, first_val, conf_col,
                        src_col, locale, r, section_fill, total_fill, thin_top, dbl_top, right,
                        num_fmt, ink, scale, calc)

        ws.freeze_panes = ws.cell(hc + 1, 1)
        ws.column_dimensions["A"].width = 46
        ws.column_dimensions["B"].width = 7
        for i in range(n_val):
            ws.column_dimensions[get_column_letter(first_val + i)].width = 14
        ws.column_dimensions[get_column_letter(conf_col)].width = 7
        ws.column_dimensions[get_column_letter(src_col)].width = 16

    if not any_sheet:                                # nothing extracted → a friendly stub sheet
        ws = wb.create_sheet("Extraction")
        ws.cell(1, 1, "No extracted line items for this document.")

    _add_analysis_sheets(wb, rows, disclosures or [], note_details or [], locale,
                         reconciliation or [], include, credit_narrative)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _recon_by_note(reconciliation: list[dict]) -> dict[str, dict]:
    """First (consolidated/current-preferred) reconciliation entry per note number."""
    out: dict[str, dict] = {}
    for e in sorted(reconciliation, key=lambda e: (e.get("basis") != "consolidated",
                                                    e.get("period_label") != "current")):
        out.setdefault(str(e.get("note_number")), e)
    return out


def _add_analysis_sheets(wb, rows: list[dict], disclosures: list[dict],
                         note_details: list[dict], locale: str,
                         reconciliation: list[dict] | None = None,
                         include: set[str] | None = None,
                         credit_narrative: dict | None = None) -> None:
    """Note details / Ratios / Disclosures / Credit Analysis sheets, each gated by the Include
    set (all on when include is None). ``credit_narrative`` is the optional stored LLM narrative
    (``run.result['credit_narrative']``) — folded into the Credit Analysis sheet when present."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    from app.services.derived import build_credit_analysis, compute_ratios, localize_disclosures

    def on(key: str) -> bool:
        return include is None or key in include

    recon = _recon_by_note(reconciliation or [])
    disclosures = localize_disclosures(disclosures, locale)

    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="243044")
    wrap = Alignment(wrap_text=True, vertical="top")

    def _header(ws, names, widths):
        for i, (n, w) in enumerate(zip(names, widths), start=1):
            c = ws.cell(1, i, n); c.font = head_font; c.fill = head_fill
            ws.column_dimensions[chr(64 + i)].width = w
        ws.freeze_panes = "A2"

    num_fmt = "#,##0;(#,##0)"
    right = Alignment(horizontal="right")

    # Note details — the extracted breakdown tables behind the face figures, laid out
    # note by note: each note's title on its own row, then its sub-items beneath it, with
    # current + prior columns. Note 1 (and its detail), then Note 2, and so on.
    ws = wb.create_sheet("Note details")
    _header(ws, [_col("Note", locale), _col("Current", locale), _col("Prior", locale)],
            [52, 16, 16])
    ws.cell(1, 2).alignment = right
    ws.cell(1, 3).alignment = right
    note_fill = PatternFill("solid", fgColor="EEF1F6")
    note_border = Border(top=Side(style="thin", color="9AA4B2"))
    ink = "1f2937"
    ri = 2

    def _note_sort_key(note):
        try:
            return (0, int(str(note.get("no")).strip()))
        except (TypeError, ValueError):
            return (1, str(note.get("no")))

    for note in sorted(note_details, key=_note_sort_key):
        title = note.get("title") or ""
        heading = f"Note {note.get('no')}" + (f" — {title}" if title else "")
        hc = ws.cell(ri, 1, heading); hc.font = Font(bold=True, color=ink)
        for col in (1, 2, 3):
            cell = ws.cell(ri, col)
            cell.fill = note_fill
            cell.border = note_border
        ri += 1
        for row in note.get("rows", []):
            vals = row.get("values") or []
            cur_v, prior_v = split_current_prior(vals)
            cur, prior = cur_v or {}, prior_v or {}
            lab = ws.cell(ri, 1, row.get("label", "")); lab.alignment = Alignment(indent=1)
            c = ws.cell(ri, 2, _num(cur.get("value"))); c.number_format = num_fmt; c.alignment = right
            c = ws.cell(ri, 3, _num(prior.get("value"))); c.number_format = num_fmt; c.alignment = right
            ri += 1
        e = recon.get(str(note.get("no")))               # note→face reconciliation (§20)
        if e is not None:
            from app.services.reconcile import tie_status

            status = tie_status(e)
            resid = _num(e.get("residual")) or 0
            if status == "unconfirmed":
                # Saying "does NOT tie" about an analysis or segment note reads as an error in
                # the filing; it only means the note is not a breakdown of that figure.
                txt = ("Reconciliation: this note is not a breakdown of the face figure it is "
                       "cited from, so no tie is asserted.")
            else:
                tie = ("ties to the face figure" if status == "tied"
                       else "does NOT tie to the face figure")
                txt = f"Reconciliation: note total {tie} (residual {resid:,.0f})."
            rc = ws.cell(ri, 1, txt)
            rc.font = Font(italic=True, size=9,
                           color=("2E7D52" if status == "tied"
                                  else "8A8F98" if status == "unconfirmed" else "C0362C"))
            ws.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=3)
            ri += 1
        ri += 1                                          # blank spacer row between notes
    if ri == 2:
        ws.cell(2, 1, "No note detail tables were parsed for this document.")

    # Ratios — grouped by category (Liquidity / Leverage / Coverage / Efficiency /
    # Profitability), a category header row before each block. Credit metrics lead.
    ws = wb.create_sheet("Ratios")
    _header(ws, ["Ratio", "Value", "Formula"], [34, 12, 60])
    cat_fill = PatternFill("solid", fgColor="EEF1F6")
    ri = 2
    last_cat = None
    for r in compute_ratios(rows, locale=locale):
        cat = r.get("category") or ""
        if cat != last_cat:
            hc = ws.cell(ri, 1, cat); hc.font = Font(bold=True, color="1f2937")
            for c in (1, 2, 3):
                ws.cell(ri, c).fill = cat_fill
            last_cat = cat
            ri += 1
        ws.cell(ri, 1, r["label"]).alignment = Alignment(indent=1)
        ws.cell(ri, 2, r["display"]).alignment = right
        ws.cell(ri, 3, r["formula"]).font = Font(size=9, color="6B7280")
        ri += 1

    # Disclosures
    ws = wb.create_sheet("Disclosures")
    _header(ws, ["Disclosure", "Present", "Page", "Where found"], [30, 10, 8, 70])
    for i, d in enumerate(disclosures, start=2):
        ws.cell(i, 1, d.get("label", ""))
        ws.cell(i, 2, "Yes" if d.get("present") else "—")
        ws.cell(i, 3, d.get("page") or "")
        ws.cell(i, 4, d.get("snippet", "")).alignment = wrap

    # Credit Analysis — the deterministic credit view (stance + rating factors from the
    # extracted ratios + report signals), plus the LLM narrative if one was generated.
    credit = build_credit_analysis(rows, disclosures, locale=locale)
    ws = wb.create_sheet("Credit Analysis")
    _header(ws, ["Credit analysis", "Value", "Rating"], [46, 18, 14])
    ws.cell(1, 2).alignment = right
    cat_fill2 = PatternFill("solid", fgColor="EEF1F6")
    ink2 = "1f2937"
    ri = 2
    sc = ws.cell(ri, 1, "Overall stance"); sc.font = Font(bold=True, color=ink2)
    ws.cell(ri, 3, credit.get("stance_label", "")).font = Font(bold=True, color=ink2)
    ri += 1
    ws.cell(ri, 1, credit.get("summary", "")).alignment = wrap
    ws.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=3)
    ri += 2
    if credit_narrative and credit_narrative.get("text"):
        model = credit_narrative.get("model") or ""
        hc = ws.cell(ri, 1, f"Narrative{f' — {model}' if model else ''}")
        hc.font = Font(bold=True, color=ink2)
        for c in (1, 2, 3):
            ws.cell(ri, c).fill = cat_fill2
        ri += 1
        nc = ws.cell(ri, 1, credit_narrative["text"]); nc.alignment = wrap
        ws.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=3)
        ws.row_dimensions[ri].height = 84
        ri += 2
    # Rating factors
    fh = ws.cell(ri, 1, "Rating factors"); fh.font = Font(bold=True, color=ink2)
    for c in (1, 2, 3):
        ws.cell(ri, c).fill = cat_fill2
    ri += 1
    for f in credit.get("factors", []):
        ws.cell(ri, 1, f"{f.get('category', '')} · {f.get('label', '')}").alignment = Alignment(indent=1)
        ws.cell(ri, 2, f.get("display", "")).alignment = right
        ws.cell(ri, 3, f.get("tone_label", ""))
        ri += 1
    # Report signals (narrative flags)
    ri += 1
    gh = ws.cell(ri, 1, "Report signals"); gh.font = Font(bold=True, color=ink2)
    for c in (1, 2, 3):
        ws.cell(ri, c).fill = cat_fill2
    ri += 1
    flags = credit.get("flags", [])
    if not flags:
        ws.cell(ri, 1, "No adverse narrative signals found in the report.").font = Font(
            italic=True, size=9, color="6B7280")
    else:
        for fl in flags:
            lab = fl.get("label", "") + (f" (p.{fl['page']})" if fl.get("page") else "")
            ws.cell(ri, 1, lab).alignment = Alignment(indent=1)
            ws.cell(ri, 2, fl.get("severity", ""))
            ws.cell(ri, 3, fl.get("implication", "")).alignment = wrap
            ri += 1

    # Honor the Include selection: drop any analysis sheet the caller didn't ask for.
    for key, name in (("note_details", "Note details"), ("ratios", "Ratios"),
                      ("disclosures", "Disclosures"), ("credit", "Credit Analysis")):
        if not on(key) and name in wb.sheetnames:
            wb.remove(wb[name])


def _emit_nodes(ws, nodes, by_key, period_cols, first_val, conf_col, src_col, locale, r,
                section_fill, total_fill, thin_top, dbl_top, right, num_fmt, ink, scale=1.0,
                calc=None):
    from openpyxl.styles import Alignment, Font

    for node in nodes:
        role = node.get("role")
        key = node.get("canonical_key")
        group = by_key.get(key) or []
        row = group[0] if group else None
        label = _label(node, locale)

        if role == "header":
            hc = ws.cell(r, 1, label)
            hc.font = Font(bold=True, color=ink)
            for c in range(1, src_col + 1):
                ws.cell(r, c).fill = section_fill
            r += 1
            # Emit EVERY child (all line items in the template), extracted or not.
            for child in node.get("children", []):
                r = _emit_nodes(ws, [child], by_key, period_cols, first_val, conf_col, src_col,
                                locale, r, section_fill, total_fill, thin_top, dbl_top, right,
                                num_fmt, ink, scale, calc)
            continue

        is_bold = role in ("subtotal", "total")
        lab = ws.cell(r, 1, label)
        lab.font = Font(bold=is_bold, color=ink)
        lab.alignment = Alignment(indent=0 if is_bold else 1)
        ws.cell(r, 2, (row or {}).get("note") or "")

        calc_notes: list[str] = []
        for (b, p) in period_cols:
            ci = first_val + period_cols.index((b, p))
            printed = _cell_value(group, b, p)
            computed = ((calc or {}).get((b, p), {}) or {}).get(key)
            if computed is not None and computed.computable and not (row or {}).get("edited"):
                # A calculated line carries its computed figure. The printed one is recorded in
                # the comment, and named as a difference when it disagrees.
                val = computed.value
                label_p = _col("Current" if p == "current" else "Prior", locale)
                bits = [f"{label_p}: computed {val:,.0f} = {computed.formula}"]
                if printed is not None and abs(printed - val) > 0.5:
                    bits.append(f"document printed {printed:,.0f} "
                                f"(difference {printed - val:,.0f})")
                elif printed is not None:
                    bits.append(f"agrees with the printed {printed:,.0f}")
                else:
                    bits.append("the document did not print this subtotal")
                calc_notes.append(" · ".join(bits))
            else:
                val = printed
                if computed is not None and not computed.computable and printed is not None:
                    calc_notes.append(f"{_col('Current' if p == 'current' else 'Prior', locale)}: "
                                      f"printed {printed:,.0f}; none of this line's components "
                                      f"were extracted, so it could not be recomputed")
            if val is not None and scale != 1.0:
                val = round(val * scale)
            cell = ws.cell(r, ci, val)
            cell.number_format = num_fmt
            cell.alignment = right
            if is_bold:
                cell.font = Font(bold=True)

        if row is not None:
            confs = [x.get("mapping_confidence") for x in group
                     if isinstance(x.get("mapping_confidence"), (int, float))]
            conf = min(confs) if confs else None
            ws.cell(r, conf_col, f"{round(conf * 100)}%" if conf is not None else "")
            # Every page a contributing line came from, so the source column stays truthful for
            # a combined figure instead of naming only the first.
            srcs = []
            for x in group:
                vals = x.get("values") or []
                where = _prov_str(vals[0].get("provenance")) if vals else ""
                if where and where not in srcs:
                    srcs.append(where)
            ws.cell(r, src_col, " · ".join(srcs))
            notes = []
            # Edited items carry their formula into the workbook as a cell note (the value is
            # the applied result; the formula is preserved for audit).
            if row.get("edited") and row.get("formula"):
                notes.append(f"Edited · formula: {row['formula']}")
            for slot, meta in sorted((row.get("edit_comments") or {}).items()):
                if (meta or {}).get("text"):
                    who = f" — {meta['by']}" if meta.get("by") else ""
                    when = f" ({meta['at']})" if meta.get("at") else ""
                    notes.append(f"Edit note [{slot}]{who}{when}: {meta['text']}")
            trace = _contribution_note(group, *period_cols[0]) if period_cols else None
            if trace:
                notes.append(trace)
            notes += calc_notes
            if notes:
                from openpyxl.comments import Comment
                lab.comment = Comment("\n\n".join(notes), "FinExtract")
        elif calc_notes:
            # A calculated line the document never printed has no extracted row of its own, so
            # this is the only place its arithmetic can be recorded.
            from openpyxl.comments import Comment
            lab.comment = Comment("\n\n".join(calc_notes), "FinExtract")

        if role == "subtotal":
            for c in range(1, src_col + 1):
                ws.cell(r, c).border = thin_top
        elif role == "total":
            for c in range(1, src_col + 1):
                ws.cell(r, c).border = dbl_top
                ws.cell(r, c).fill = total_fill
        r += 1
    return r


def _apply_units(v, scale: float):
    if v is None:
        return None
    return round(v * scale)


def build_json(statements: dict, notes_index: list, note_detail: dict, *,
               basis: str = "consolidated", currency: str = "INR", units: str = "crore") -> bytes:
    payload = {
        "entity": "Reliance Industries Ltd",
        "period": "FY2024-25",
        "dataset": basis,
        "units": f"{currency}_{units}".lower(),
        "statements": {},
        "notes": [],
    }
    for key, st in statements.items():
        rows = []
        for r in st["rows"]:
            if r.get("kind") in ("section", "subhead"):
                continue
            rows.append({
                "item": r["label"],
                "value": r.get("v1"),
                "value_prior": r.get("v2"),
                "note_ref": r.get("note"),
                "confidence": (CONF_PCT.get(r.get("conf"), None) or 0) / 100 if r.get("conf") else None,
                "kind": r.get("kind", "item"),
            })
        payload["statements"][key] = rows
    for n in notes_index:
        entry = {"note": n["no"], "title": n["title"]}
        if n["no"] in note_detail:
            entry["reconciliation"] = note_detail[n["no"]]["reconciliation"]
        payload["notes"].append(entry)
    return json.dumps(payload, indent=2).encode("utf-8")


def build_xlsx(statements: dict, notes_index: list, note_detail: dict, *,
               basis: str = "consolidated", currency: str = "INR", units: str = "crore",
               include: dict | None = None) -> bytes:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    include = include or {}
    wb = openpyxl.Workbook()
    header_fill = PatternFill("solid", fgColor="217346")
    header_font = Font(color="FFFFFF", bold=True)
    total_font = Font(bold=True)

    first = True
    for key, st in statements.items():
        ws = wb.active if first else wb.create_sheet()
        ws.title = st["label"][:31]
        first = False

        cols = ["#", "Line item", "FY25", "FY24", "Note"]
        if include.get("confidence", True):
            cols.append("Conf.")
        ws.append(cols)
        for ci, _ in enumerate(cols, start=1):
            c = ws.cell(row=1, column=ci)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="left" if ci <= 2 else "right")

        n = 0
        for r in st["rows"]:
            kind = r.get("kind", "item")
            if kind in ("section", "subhead"):
                ws.append(["", r["label"]])
                ws.cell(row=ws.max_row, column=2).font = total_font
                continue
            n += 1
            conf_pct = CONF_PCT.get(r.get("conf")) if r.get("conf") else None
            row = [n, r["label"], r.get("v1"), r.get("v2"), r.get("note", "")]
            if include.get("confidence", True):
                row.append(f"{conf_pct}%" if conf_pct else "ƒ")
            ws.append(row)
            if kind in ("subtotal", "total"):
                for ci in range(1, len(cols) + 1):
                    ws.cell(row=ws.max_row, column=ci).font = total_font

        widths = [5, 42, 12, 12, 8, 8]
        for ci, w in enumerate(widths[:len(cols)], start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    # Separate All-notes sheet.
    if include.get("notes_sheet", True):
        ws = wb.create_sheet("All notes")
        ws.append(["Note", "Particulars", "FY25", "FY24"])
        for ci in range(1, 5):
            ws.cell(row=1, column=ci).fill = header_fill
            ws.cell(row=1, column=ci).font = header_font
        for n in notes_index:
            detail = note_detail.get(n["no"])
            if detail:
                for dr in detail["rows"]:
                    ws.append([f"N{n['no']}", dr["label"], dr.get("v1"), dr.get("v2")])
            else:
                ws.append([f"N{n['no']}", n["title"], "", ""])
        ws.column_dimensions["B"].width = 48

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
