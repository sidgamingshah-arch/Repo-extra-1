"""The output template as a workbook an analyst can edit, and read back again.

A template is a section→line tree with rollups — JSON, which is the right shape for the
pipeline and the wrong shape for the person who decides what the spread should contain. This
module is the round trip: one flat sheet, one row per line, with the two facts that decide how a
line behaves stated in their own columns:

* **Kind** — ``extracted`` (read off the document by the mapper) or ``calculated`` (computed from
  other lines and never mapped). This is the distinction that makes a template reviewable: a
  calculated line that is mistakenly marked extracted gets a figure from the page instead of from
  its components, and a subtotal that stops tying is the first anyone hears of it.
* **Calculated from** — the exact lines a calculated one is made of, which is also what the
  structural checks recompute and compare, so editing this column changes what gets validated.

The workbook is authoritative on the way back in: ``extracted`` drops any rollup, ``calculated``
requires one, and every referenced key must exist in the sheet. Nothing is inferred silently —
an ambiguous edit is an error with the row number on it, not a guess.
"""
from __future__ import annotations

import io
import re

from app.services.statements import TITLES as STATEMENT_TITLES

# Column order of the Template sheet. Kept as data so the reader and the writer cannot drift.
COLUMNS = [
    ("statement", "Statement"),
    ("section", "Section"),
    ("node_id", "Node ID"),
    ("canonical_key", "Canonical key"),
    ("label", "Label (en)"),
    ("label_zh", "Label (zh)"),
    ("label_ar", "Label (ar)"),
    ("label_fr", "Label (fr)"),
    ("role", "Role"),
    ("kind", "Kind"),
    ("op", "Calculation"),
    ("children", "Calculated from"),
    ("sign", "Sign"),
    ("expects_note", "Expects note"),
    ("required", "Required"),
]
_WIDTHS = [20, 34, 40, 44, 40, 26, 26, 30, 11, 12, 12, 60, 14, 13, 10]

KIND_EXTRACTED = "extracted"
KIND_CALCULATED = "calculated"
KIND_HEADING = "heading"
_KINDS = {KIND_EXTRACTED, KIND_CALCULATED, KIND_HEADING}
# Exactly the ops ``schemas.template.Rollup`` accepts. ``weighted_sum`` was still admitted here
# after the JSON gate dropped it, so the two authoring routes disagreed about what is legal: a
# workbook naming it was accepted and published a template whose op the JSON upload refuses and
# whose relation ``structural_checks`` can only report as ``unsupported_op`` — the calculated line
# it was authored on silently lost its arithmetic guard. A workbook that names it is now refused on
# the row that names it, which is the same answer the JSON route gives.
_OPS = {"sum", "diff"}
_ROLES = {"header", "line", "subtotal", "total"}
_LOCALES = ("zh", "ar", "fr")

# The vocabulary lives in ``services.statements`` because the API serves it too — a screen
# offering a statement this importer would refuse is a disagreement with no owner.
_STATEMENT_TITLE = STATEMENT_TITLES
_TITLE_STATEMENT = {v.lower(): k for k, v in _STATEMENT_TITLE.items()}


class TemplateSheetError(ValueError):
    """An edited workbook that cannot be read as a template, with the offending row named."""


def _kind_of(node: dict) -> str:
    if (node.get("role") or "line") == "header":
        return KIND_HEADING
    return KIND_CALCULATED if node.get("rollup") else KIND_EXTRACTED


def _yes(v) -> str:
    return "yes" if v else ""


def build_template_xlsx(definition: dict, *, filename_hint: str = "template") -> bytes:
    """The template as an editable workbook: one row per line, plus its identity checks."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.Workbook()
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="34405A")
    sec_font = Font(bold=True)
    sec_fill = PatternFill("solid", fgColor="EEF1F6")
    calc_fill = PatternFill("solid", fgColor="FFF6E5")
    wrap = Alignment(vertical="top", wrap_text=True)

    ws = wb.active
    ws.title = "Template"
    ws.append([h for _k, h in COLUMNS])
    for i, w in enumerate(_WIDTHS, start=1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w
    for c in ws[1]:
        c.font, c.fill = head_font, head_fill
        c.alignment = Alignment(vertical="center", wrap_text=True)

    def write(node: dict, statement: str, section: str) -> None:
        i18n = node.get("label_i18n") or {}
        rollup = node.get("rollup") or {}
        kind = _kind_of(node)
        row = {
            "statement": statement, "section": section,
            "node_id": node.get("node_id") or "",
            "canonical_key": node.get("canonical_key") or "",
            "label": i18n.get("en") or node.get("label") or "",
            "role": node.get("role") or "line",
            "kind": kind,
            "op": rollup.get("op") or ("sum" if kind == KIND_CALCULATED else ""),
            # One key per line inside the cell: a 20-child subtotal is unreadable on one line,
            # and newline-separated is what an editor can actually work with.
            "children": "\n".join(rollup.get("children") or []),
            "sign": node.get("sign") or "natural",
            "expects_note": _yes(node.get("expects_note")),
            "required": _yes(node.get("required")),
        }
        for loc in _LOCALES:
            row[f"label_{loc}"] = i18n.get(loc, "")
        ws.append([row.get(k, "") for k, _h in COLUMNS])
        r = ws.max_row
        for c in ws[r]:
            c.alignment = wrap
        if kind == KIND_HEADING:
            for c in ws[r]:
                c.font, c.fill = sec_font, sec_fill
        elif kind == KIND_CALCULATED:
            for c in ws[r]:
                c.fill = calc_fill

    for stmt in definition.get("statements", []):
        st = str(stmt.get("type") or "")
        title = _STATEMENT_TITLE.get(st, st.replace("_", " ").capitalize())
        for sec in stmt.get("sections") or []:
            write(sec, title, "")
            for child in sec.get("children") or []:
                write(child, title, sec.get("node_id") or "")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(1, len(COLUMNS)).column_letter}{ws.max_row}"
    kind_col = ws.cell(1, 1 + [k for k, _h in COLUMNS].index("kind")).column_letter
    dv = DataValidation(type="list", formula1=f'"{",".join(sorted(_KINDS))}"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{kind_col}2:{kind_col}{max(2, ws.max_row)}")

    # --- identity checks (statement-level, also calculated relationships) ---
    ids = wb.create_sheet("Identities")
    ids.append(["Statement", "Identity ID", "Left (canonical key)", "Calculation",
                "Right (canonical keys)", "Tolerance (abs)", "Tolerance (rel)"])
    for c in ids[1]:
        c.font, c.fill = head_font, head_fill
    for stmt in definition.get("statements", []):
        st = str(stmt.get("type") or "")
        title = _STATEMENT_TITLE.get(st, st)
        for ident in stmt.get("identities", []) or []:
            rhs = ident.get("rhs") or {}
            ids.append([title, ident.get("id") or "", ident.get("lhs") or "",
                        rhs.get("op") or "sum", "\n".join(rhs.get("children") or []),
                        ident.get("tolerance_abs", 1.0), ident.get("tolerance_rel", 0.001)])
    for col, w in zip("ABCDEFG", (20, 26, 44, 14, 60, 15, 15)):
        ids.column_dimensions[col].width = w
    for row in ids.iter_rows(min_row=2):
        for c in row:
            c.alignment = wrap
    ids.freeze_panes = "A2"

    # --- the contract, in the file itself ---
    rm = wb.create_sheet("Read me")
    rm.column_dimensions["A"].width = 26
    rm.column_dimensions["B"].width = 118
    lines = [
        ("Template", f"{definition.get('name') or filename_hint} "
                     f"(key: {definition.get('template_key') or '—'})"),
        ("", ""),
        ("How to use this", "Edit the Template sheet, then upload it back on the Template & "
                            "Ontology screen. A new template VERSION is created — earlier runs "
                            "keep explaining themselves against the version they used."),
        ("", ""),
        ("Statement", "Which statement the line belongs to. Keep the spelling used in the "
                      "existing rows; a new statement name is rejected rather than guessed."),
        ("Section", "The Node ID of the heading this line sits under. Leave BLANK for a row that "
                    "is itself a section heading or a statement-level total."),
        ("Node ID", "Stable identifier. Leave as-is for existing lines; for a new line use a "
                    "unique lowercase_with_underscores name (the canonical key is fine)."),
        ("Canonical key", "What extraction maps to and what every check, export and formula "
                          "refers to. Must be unique. Renaming one breaks the ontology entry and "
                          "any rollup that names it, so change both together."),
        ("Label (en/zh/ar/fr)", "What the line is called on screen and in the export, per output "
                                "language. English is required; the others fall back to it."),
        ("Role", "header | line | subtotal | total — presentation emphasis in the grid."),
        ("Kind", f"{KIND_EXTRACTED} — read off the document by the mapper.\n"
                 f"{KIND_CALCULATED} — computed from other lines and never mapped; requires "
                 f"'Calculated from'.\n"
                 f"{KIND_HEADING} — a section heading; carries no figure."),
        ("Calculation", "sum or diff — how the lines in 'Calculated from' combine (diff = first "
                        "minus the rest). Required for a calculated line."),
        ("Calculated from", "One canonical key per line. These are recomputed and compared with "
                            "the extracted figure, which is what makes a mis-mapping show up as a "
                            "failed check instead of a wrong number."),
        ("Sign", "natural (as reported) or contra (an expense/outflow shown positive)."),
        ("Expects note / Required", "'yes' to flag the line as normally note-referenced, or as "
                                    "one whose absence should be raised."),
        ("", ""),
        ("Shading", "Grey rows are section headings. Amber rows are calculated lines."),
        ("Identities sheet", "Statement-level equalities (e.g. total assets = total equity and "
                             "liabilities) checked after extraction, with their tolerances."),
    ]
    for a, b in lines:
        rm.append([a, b])
        rm.cell(rm.max_row, 1).font = sec_font
        rm.cell(rm.max_row, 2).alignment = wrap
    wb.move_sheet("Read me", offset=-2)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _cells(ws) -> list[dict]:
    """Sheet rows as dicts keyed by our column names, matched on the HEADER TEXT so a reordered
    or extra column in an edited workbook doesn't shift every value one to the left."""
    header = [str(c.value or "").strip().lower() for c in ws[1]]
    by_head = {h.lower(): k for k, h in COLUMNS}
    idx = {}
    for i, h in enumerate(header):
        key = by_head.get(h)
        if key:
            idx[key] = i
    missing = [h for k, h in COLUMNS
               if k in ("statement", "canonical_key", "label", "kind") and k not in idx]
    if missing:
        raise TemplateSheetError(
            f"The Template sheet is missing required column(s): {', '.join(missing)}")
    out = []
    for n, row in enumerate(ws.iter_rows(min_row=2), start=2):
        vals = {k: row[i].value if i < len(row) else None for k, i in idx.items()}
        if not any(str(v or "").strip() for v in vals.values()):
            continue                                   # a blank spacer row
        vals["_row"] = n
        out.append(vals)
    return out


def _s(v) -> str:
    return str(v).strip() if v is not None else ""


def _keys(v) -> list[str]:
    """A 'Calculated from' cell → canonical keys. Accepts one per line, or comma/plus separated,
    because that is how people actually type a list into a cell."""
    return [p for p in (x.strip() for x in re.split(r"[\n,;+]+", _s(v))) if p]


def _truthy(v) -> bool:
    return _s(v).lower() in {"yes", "y", "true", "1", "x", "✓"}


def _slug(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", s.lower()).strip("_")


def _statement_type(title: str) -> str | None:
    """A Statement cell → a statement type, accepting the heading we wrote, the type itself, or
    the obvious respellings ("Profit and loss"). Anything else is refused, not guessed."""
    t = _s(title).lower()
    if t in _TITLE_STATEMENT:
        return _TITLE_STATEMENT[t]
    slug = _slug(t)
    if slug in _STATEMENT_TITLE:
        return slug
    return next((st for st, disp in _STATEMENT_TITLE.items() if _slug(disp) == slug), None)


def parse_template_xlsx(data: bytes, *, template_key: str, name: str) -> dict:
    """An edited workbook → a template definition, or a TemplateSheetError naming the bad row.

    Deliberately strict. A template drives what every extraction maps to and what every check
    recomputes, so a row this reader is unsure about is a row it refuses: guessing here would
    show up much later as a figure on the wrong line.
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    if "Template" not in wb.sheetnames:
        raise TemplateSheetError("The workbook has no 'Template' sheet. Download the current "
                                 "template, edit that, and upload it back.")
    rows = _cells(wb["Template"])
    if not rows:
        raise TemplateSheetError("The Template sheet has no rows.")

    # Pass 1: validate every row, collect the keys that exist so rollups can be checked.
    known: set[str] = set()
    for r in rows:
        key = _s(r.get("canonical_key"))
        if not key:
            raise TemplateSheetError(f"Row {r['_row']}: Canonical key is required.")
        if key in known:
            raise TemplateSheetError(f"Row {r['_row']}: duplicate canonical key '{key}'.")
        known.add(key)

    statements: dict[str, dict] = {}
    order: list[str] = []
    sections: dict[tuple[str, str], dict] = {}

    for r in rows:
        row_no = r["_row"]
        title = _s(r.get("statement"))
        if not title:
            raise TemplateSheetError(f"Row {row_no}: Statement is required.")
        st = _statement_type(title)
        if st is None:
            raise TemplateSheetError(
                f"Row {row_no}: unknown Statement '{title}'. Use one of: "
                f"{', '.join(sorted(_STATEMENT_TITLE.values()))}.")

        key = _s(r.get("canonical_key"))
        label = _s(r.get("label"))
        if not label:
            raise TemplateSheetError(f"Row {row_no}: Label (en) is required for '{key}'.")
        kind = _s(r.get("kind")).lower() or KIND_EXTRACTED
        if kind not in _KINDS:
            raise TemplateSheetError(
                f"Row {row_no}: Kind must be one of {', '.join(sorted(_KINDS))}, not '{kind}'.")
        role = _s(r.get("role")).lower() or ("header" if kind == KIND_HEADING else "line")
        if role not in _ROLES:
            raise TemplateSheetError(
                f"Row {row_no}: Role must be one of {', '.join(sorted(_ROLES))}, not '{role}'.")
        children = _keys(r.get("children"))
        op = _s(r.get("op")).lower() or "sum"
        if kind == KIND_CALCULATED:
            if not children:
                raise TemplateSheetError(
                    f"Row {row_no}: '{key}' is marked {KIND_CALCULATED} but 'Calculated from' is "
                    f"empty — a calculated line has to say what it is calculated from.")
            if op not in _OPS:
                raise TemplateSheetError(
                    f"Row {row_no}: Calculation must be one of {', '.join(sorted(_OPS))}.")
            unknown = [c for c in children if c not in known]
            if unknown:
                raise TemplateSheetError(
                    f"Row {row_no}: '{key}' is calculated from key(s) that are not in this "
                    f"template: {', '.join(unknown)}.")
            if key in children:
                raise TemplateSheetError(f"Row {row_no}: '{key}' cannot be calculated from itself.")
        elif children:
            raise TemplateSheetError(
                f"Row {row_no}: '{key}' is marked {kind} but has 'Calculated from' set. Mark it "
                f"{KIND_CALCULATED}, or clear that column.")

        node: dict = {
            "node_id": _s(r.get("node_id")) or key,
            "canonical_key": key, "label": label, "role": role,
            "label_i18n": {"en": label,
                           **{loc: _s(r.get(f"label_{loc}")) for loc in _LOCALES
                              if _s(r.get(f"label_{loc}"))}},
            "sign": _s(r.get("sign")).lower() or "natural",
        }
        if _truthy(r.get("expects_note")):
            node["expects_note"] = True
        if _truthy(r.get("required")):
            node["required"] = True
        if kind == KIND_CALCULATED:
            node["rollup"] = {"op": op, "children": children}

        if st not in statements:
            statements[st] = {"type": st, "sections": [], "identities": []}
            order.append(st)
        parent_id = _s(r.get("section"))
        if not parent_id:
            node["children"] = []
            statements[st]["sections"].append(node)
            sections[(st, node["node_id"])] = node
        else:
            parent = sections.get((st, parent_id))
            if parent is None:
                raise TemplateSheetError(
                    f"Row {row_no}: Section '{parent_id}' has no heading row above it in "
                    f"{title}. A section row (blank Section) must come before its lines.")
            parent["children"].append(node)

    # --- identities ---
    if "Identities" in wb.sheetnames:
        ws = wb["Identities"]
        for n, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            cells = list(row) + [None] * (7 - len(row))
            title, ident_id, lhs, op, rhs, tol_abs, tol_rel = cells[:7]
            if not _s(ident_id) and not _s(lhs):
                continue
            st = _statement_type(_s(title)) or _slug(_s(title))
            if st not in statements:
                raise TemplateSheetError(
                    f"Identities row {n}: statement '{_s(title)}' is not in the Template sheet.")
            terms = _keys(rhs)
            for k in [_s(lhs), *terms]:
                if k and k not in known:
                    raise TemplateSheetError(
                        f"Identities row {n}: '{k}' is not a canonical key in this template.")
            ident = {"id": _s(ident_id) or f"{st}_identity_{n}", "lhs": _s(lhs),
                     "rhs": {"op": _s(op).lower() or "sum", "children": terms}}
            if isinstance(tol_abs, (int, float)):
                ident["tolerance_abs"] = float(tol_abs)
            if isinstance(tol_rel, (int, float)):
                ident["tolerance_rel"] = float(tol_rel)
            statements[st]["identities"].append(ident)

    return {"schema_version": 1, "template_key": template_key, "name": name,
            "statements": [statements[st] for st in order]}
