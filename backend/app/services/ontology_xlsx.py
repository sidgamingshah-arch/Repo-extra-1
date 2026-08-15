"""The rulebook as a workbook an admin can edit, and read back again without losing anything.

An ontology is 185 concepts plus a dozen blocks of policy — the right shape for the matcher and a
poor one for the person deciding what a caption means. This module is the round trip, built on the
same conventions as ``template_xlsx``: newline-separated lists inside a cell, one row per thing, and
an edit that cannot be read unambiguously is an error with the row number on it rather than a guess.

WHAT GETS A GRID, AND WHY NOT EVERYTHING. The concepts, the section layer and the netting rules are
what an admin actually tunes, so those are real sheets with one row each. The remaining top-level
blocks — ``normalisation``, ``binding``, ``scope_selection``, ``global_rules``,
``residual_framework``, ``validation``, ``worked_examples``, ``metadata`` — are nested policy and
prose. Inventing a grid for each would mean inventing a validation story for each, and an admin
editing them is editing structure, not vocabulary. They travel VERBATIM as JSON on the Config sheet:
lossless, still editable, and honest about being the advanced end of the file.

BLANK IS NOT THE SAME AS EMPTY, and this is not a nicety. Thirteen concepts — every residual bucket —
carry ``aliases: []`` on purpose, because their ``alias_matching`` is disabled and an empty alias list
is the statement of that. One netting rule carries ``add_keys: []`` the same way. If a blank cell
meant "empty list" there would be no way to say "this field is not set", and if blank meant "absent"
those thirteen deliberate declarations would vanish on the first round trip. So: a blank cell means
the field is ABSENT, and a cell holding ``[]`` means the field is present and empty. The README sheet
says so where an editor will see it.
"""
from __future__ import annotations

import io
import json

# Per-concept columns, in workbook order. Kept as data so the writer and the reader cannot drift.
# `kind` is how the value is encoded in the cell: "s" scalar string, "i" int, "b" boolean,
# "l" newline-separated list, "j" JSON.
CONCEPT_COLUMNS: list[tuple[str, str, str]] = [
    ("canonical_key", "Canonical key", "s"),
    ("label", "Label", "s"),
    ("inherits", "Inherits (section)", "s"),
    ("definition", "Definition — what this line MEANS", "s"),
    ("aliases", "Aliases", "l"),
    ("keyword_hints", "Keyword hints", "l"),
    ("regex_hints", "Regex hints", "l"),
    ("exclude_hints", "Exclude hints", "l"),
    ("include", "Include (belongs here)", "l"),
    ("exclude", "Exclude (does NOT belong here)", "l"),
    ("confusable_with", "Confusable with", "l"),
    ("sign_convention", "Sign convention", "s"),
    ("value_scope", "Value scope", "s"),
    ("match_priority", "Match priority", "i"),
    ("alias_matching", "Alias matching", "s"),
    ("extraction_mode", "Extraction mode", "s"),
    ("unit_of_account", "Unit of account", "s"),
    ("temporality", "Temporality", "s"),
    ("note_use", "Note use", "s"),
    ("face_only", "Face only", "b"),
    ("is_gross_parent", "Is gross parent", "b"),
    ("children_if_decomposed", "Children if decomposed", "l"),
    ("derivation", "Derivation", "s"),
    ("decomposition_rule", "Decomposition rule", "s"),
    ("aggregation_note", "Aggregation note", "s"),
    ("section_disambiguation", "Section disambiguation", "s"),
    ("sole_component_of", "Sole component of", "s"),
    ("never_sweep", "Never sweep", "l"),
    ("expected_components", "Expected components", "l"),
    ("template_note", "Template note", "s"),
    ("notes_as_source_rationale", "Notes-as-source rationale", "s"),
    ("residual_policy", "Residual policy (JSON)", "j"),
    ("equivalence", "Equivalence (JSON)", "j"),
]
_CONCEPT_WIDTHS = [46, 34, 30, 64, 40, 30, 30, 30, 40, 40, 40, 20, 20, 14, 15,
                   16, 18, 14, 16, 11, 15, 34, 40, 40, 40, 40, 34, 34, 40, 40, 40, 44, 34]

# Per-section columns (`section_defaults`), same encoding vocabulary.
SECTION_COLUMNS: list[tuple[str, str, str]] = [
    ("section_key", "Section key", "s"),
    ("statement", "Statement", "s"),
    ("section_scope", "Section scope", "l"),
    ("temporality", "Temporality", "s"),
    ("unit_of_account", "Unit of account", "s"),
    ("value_scope", "Value scope", "s"),
    ("extraction_mode", "Extraction mode", "s"),
    ("face_only", "Face only", "b"),
    ("note_use", "Note use", "s"),
    ("note_use_rationale", "Note-use rationale", "s"),
    ("sign_convention", "Sign convention", "s"),
    ("match_priority", "Match priority", "i"),
    ("include", "Include", "l"),
    ("exclude", "Exclude", "l"),
]
_SECTION_WIDTHS = [34, 18, 34, 14, 18, 20, 16, 11, 14, 48, 20, 14, 40, 40]

NETTING_COLUMNS: list[tuple[str, str, str]] = [
    ("id", "Rule id", "s"),
    ("target_key", "Target key", "s"),
    ("subtract_keys", "Subtract keys", "l"),
    ("add_keys", "Add keys", "l"),
    ("condition", "Condition", "s"),
    ("evidence_required", "Evidence required", "b"),
    ("on_apply", "On apply", "s"),
    ("decompose_into", "Decompose into", "l"),
    ("label", "Label", "s"),
]
_NETTING_WIDTHS = [30, 46, 52, 40, 60, 15, 40, 40, 34]

# Everything that is NOT one of the three grids. Order preserved so the rebuilt file reads like the
# authored one; anything the source carries beyond this list still travels (see `build_ontology_xlsx`).
CONFIG_KEYS = [
    "schema_version", "ontology_key", "target_template_key", "locale", "supported_locales",
    "normalisation", "binding", "scope_selection", "global_rules", "residual_framework",
    "validation", "worked_examples", "metadata",
]

# Any key a grid's columns do not cover travels here, as JSON, so a field this module has never heard
# of still round-trips instead of being dropped by a column list that went stale. The round-trip test
# found three netting fields missing this way; a catch-all is the fix that does not need repeating.
OTHER_COLUMN = "Other (JSON)"

SHEET_CONCEPTS = "Concepts"
SHEET_SECTIONS = "Section defaults"
SHEET_NETTING = "Netting rules"
SHEET_CONFIG = "Config (JSON)"
SHEET_README = "README"

# A cell holding exactly this means "present, and empty" — see the module docstring.
EMPTY_MARKER = "[]"


class OntologySheetError(ValueError):
    """A workbook that cannot be read as a rulebook. Carries the sheet and row."""


def _enc(value, kind: str) -> str:
    """One value → one cell. Absent is handled by the caller; this never sees None."""
    if kind == "l":
        items = list(value or [])
        return "\n".join(str(x) for x in items) if items else EMPTY_MARKER
    if kind == "j":
        return json.dumps(value, indent=2, ensure_ascii=False)
    if kind == "b":
        return "yes" if value else "no"
    if kind == "i":
        return value
    return "" if value is None else str(value)


def _dec(raw, kind: str, *, sheet: str, row: int, header: str):
    """One cell → one value, or ``None`` for "the field is absent"."""
    if raw is None:
        return None
    text = str(raw).strip() if not isinstance(raw, bool) else raw
    if text == "":
        return None
    if kind == "l":
        if text == EMPTY_MARKER:
            return []
        return [ln.strip() for ln in str(text).splitlines() if ln.strip()]
    if kind == "j":
        try:
            return json.loads(str(text))
        except json.JSONDecodeError as exc:
            raise OntologySheetError(
                f"{sheet} row {row}: {header!r} is not valid JSON ({exc.msg} at position "
                f"{exc.pos}). Leave the cell blank to remove it.") from exc
    if kind == "b":
        if isinstance(text, bool):
            return text
        low = str(text).strip().lower()
        if low in ("yes", "true", "y", "1"):
            return True
        if low in ("no", "false", "n", "0"):
            return False
        raise OntologySheetError(
            f"{sheet} row {row}: {header!r} must be yes or no, not {text!r}.")
    if kind == "i":
        try:
            return int(float(str(text)))
        except ValueError as exc:
            raise OntologySheetError(
                f"{sheet} row {row}: {header!r} must be a whole number, not {text!r}.") from exc
    return str(text)


def _leftovers(src: dict, columns: list[tuple[str, str, str]], handled: set[str]) -> str:
    """Whatever the grid's own columns do not carry, as JSON — or "" when there is nothing left."""
    known = {k for k, _h, _t in columns} | handled
    rest = {k: v for k, v in src.items() if k not in known}
    return json.dumps(rest, indent=2, ensure_ascii=False) if rest else ""


def _locale_columns(definition: dict) -> list[str]:
    """The locales `aliases_i18n` gets a column for: those the rulebook says it supports, plus any
    a concept actually carries. Derived, so a rulebook that gains a locale gains a column instead of
    quietly dropping it."""
    locales = list(definition.get("supported_locales") or [])
    for m in definition.get("mappings") or []:
        for loc in (m.get("aliases_i18n") or {}):
            if loc not in locales:
                locales.append(loc)
    return locales or ["en"]


def build_ontology_xlsx(definition: dict, *, filename_hint: str = "ontology") -> bytes:
    """The rulebook as an editable workbook. Lossless: see ``parse_ontology_xlsx``."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="34405A")
    residual_fill = PatternFill("solid", fgColor="FFF6E5")
    wrap = Alignment(vertical="top", wrap_text=True)
    locales = _locale_columns(definition)

    def head(ws, columns: list[tuple[str, str, str]], widths: list[int], extra: list[str]) -> None:
        ws.append([h for _k, h, _t in columns] + extra + [OTHER_COLUMN])
        for i, w in enumerate([*widths, *([34] * len(extra)), 46], start=1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = w
        for c in ws[1]:
            c.font, c.fill = head_font, head_fill
            c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.freeze_panes = "B2"

    # --- Concepts -------------------------------------------------------------------------
    ws = wb.active
    ws.title = SHEET_CONCEPTS
    alias_headers = [f"Aliases ({loc})" for loc in locales]
    head(ws, CONCEPT_COLUMNS, _CONCEPT_WIDTHS, alias_headers)
    for m in definition.get("mappings") or []:
        row = [_enc(m[k], t) if k in m else "" for k, _h, t in CONCEPT_COLUMNS]
        i18n = m.get("aliases_i18n")
        for loc in locales:
            if i18n is None or loc not in i18n:
                row.append("")
            else:
                row.append(_enc(i18n[loc], "l"))
        row.append(_leftovers(m, CONCEPT_COLUMNS, {"aliases_i18n"}))
        ws.append(row)
        for c in ws[ws.max_row]:
            c.alignment = wrap
        # The residual buckets read differently from every other concept — they match nothing and
        # are populated by the sweep — so they are shaded rather than left to be told apart by
        # reading two columns.
        if m.get("value_scope") == "exclusive_residual":
            for c in ws[ws.max_row]:
                c.fill = residual_fill

    # --- Section defaults -----------------------------------------------------------------
    ws2 = wb.create_sheet(SHEET_SECTIONS)
    head(ws2, SECTION_COLUMNS, _SECTION_WIDTHS, [])
    for key, sec in (definition.get("section_defaults") or {}).items():
        src = {"section_key": key, **sec}
        ws2.append([*[_enc(src[k], t) if k in src else "" for k, _h, t in SECTION_COLUMNS],
                    _leftovers(src, SECTION_COLUMNS, set())])
        for c in ws2[ws2.max_row]:
            c.alignment = wrap

    # --- Netting rules --------------------------------------------------------------------
    ws3 = wb.create_sheet(SHEET_NETTING)
    head(ws3, NETTING_COLUMNS, _NETTING_WIDTHS, [])
    for rule in definition.get("netting_rules") or []:
        ws3.append([*[_enc(rule[k], t) if k in rule else "" for k, _h, t in NETTING_COLUMNS],
                    _leftovers(rule, NETTING_COLUMNS, set())])
        for c in ws3[ws3.max_row]:
            c.alignment = wrap

    # --- Config ---------------------------------------------------------------------------
    # Every top-level key that is not one of the grids above, verbatim. Driven off the definition
    # rather than CONFIG_KEYS alone, so a block this module has never heard of still round-trips
    # instead of being dropped by a list that went stale.
    ws4 = wb.create_sheet(SHEET_CONFIG)
    ws4.append(["Key", "Value (JSON)"])
    ws4.column_dimensions["A"].width = 26
    ws4.column_dimensions["B"].width = 140
    for c in ws4[1]:
        c.font, c.fill = head_font, head_fill
    gridded = {"mappings", "section_defaults", "netting_rules"}
    ordered = [k for k in CONFIG_KEYS if k in definition]
    ordered += [k for k in definition if k not in gridded and k not in ordered]
    for key in ordered:
        ws4.append([key, json.dumps(definition[key], indent=2, ensure_ascii=False)])
        ws4.cell(ws4.max_row, 2).alignment = wrap
        ws4.cell(ws4.max_row, 1).alignment = Alignment(vertical="top")

    # --- README ---------------------------------------------------------------------------
    ws5 = wb.create_sheet(SHEET_README)
    ws5.column_dimensions["A"].width = 118
    for line in _readme(definition, locales):
        ws5.append([line])
        ws5.cell(ws5.max_row, 1).alignment = Alignment(vertical="top", wrap_text=True)
    ws5["A1"].font = Font(bold=True, size=13)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _readme(definition: dict, locales: list[str]) -> list[str]:
    key = definition.get("ontology_key") or "ontology"
    n = len(definition.get("mappings") or [])
    return [
        f"{key} — editable rulebook ({n} concepts)",
        "",
        "HOW TO EDIT",
        f"  • {SHEET_CONCEPTS}: one row per concept. This is the vocabulary — what a printed caption",
        "    has to look like, or mean, to be recognised as this line.",
        f"  • {SHEET_SECTIONS}: the section layer every concept inherits from (see the 'Inherits'",
        "    column on Concepts). A value set on a concept overrides the section's.",
        f"  • {SHEET_NETTING}: containment netting — a face line restated net of lines inside it.",
        f"  • {SHEET_CONFIG}: the structural policy blocks, verbatim JSON. Edit with care; the",
        "    upload validates them and refuses anything the schema does not accept.",
        "",
        "A LIST IN A CELL is one item per line (use Alt+Enter in Excel to add a line).",
        "",
        "BLANK versus EMPTY — this distinction is load-bearing:",
        "  • a BLANK cell means the field is not set on that row;",
        f"  • a cell holding exactly {EMPTY_MARKER} means the field IS set, to an empty list.",
        "    The residual buckets (shaded on the Concepts sheet) rely on this: their alias list is",
        "    deliberately empty, because they match nothing and are filled by the sweep instead.",
        "    Clearing those cells to blank would change what the rulebook says.",
        "",
        "DO NOT rename a sheet, reorder or rename a header, or change a Canonical key you did not",
        "intend to rename — the keys are what tie a concept to the template and to every past run.",
        "",
        f"ALIASES per locale have their own columns ({', '.join(f'Aliases ({l})' for l in locales)}).",
        "The plain 'Aliases' column is what non-localized consumers read; keep it in step with the",
        "default locale's column.",
        "",
        "UPLOADING: POST the edited file to /api/v1/ontologies/xlsx (admin only). It is validated",
        "against the target template and published as a NEW VERSION — nothing is overwritten, and",
        "the latest version stored is the one the next run maps against.",
    ]


def _merge_other(target: dict, rec: dict, *, sheet: str, row: int) -> None:
    """Fold the catch-all column back in, refusing an overlap rather than picking a winner.

    A key present both as its own column and inside the JSON is ambiguous — two cells claiming one
    field — and guessing which the editor meant is how an edit gets silently discarded.
    """
    extra = _dec(rec.get(OTHER_COLUMN), "j", sheet=sheet, row=row, header=OTHER_COLUMN)
    if extra is None:
        return
    if not isinstance(extra, dict):
        raise OntologySheetError(
            f"{sheet} row {row}: {OTHER_COLUMN!r} must be a JSON object of field: value.")
    clash = sorted(set(extra) & set(target))
    if clash:
        raise OntologySheetError(
            f"{sheet} row {row}: {', '.join(repr(c) for c in clash)} appears both in its own column "
            f"and inside {OTHER_COLUMN!r}. Keep each field in one place.")
    target.update(extra)


def _rows(ws, columns: list[tuple[str, str, str]], extra: list[str]) -> list[dict]:
    """Sheet → list of {header: raw cell}, header-driven so column ORDER may move even though the
    names may not. A missing expected column is an error, not a silently absent field."""
    headers = [(c.value or "").strip() if isinstance(c.value, str) else c.value for c in ws[1]]
    expected = [h for _k, h, _t in columns]
    missing = [h for h in expected if h not in headers]
    if missing:
        raise OntologySheetError(
            f"{ws.title}: missing column(s) {', '.join(repr(m) for m in missing)}. "
            f"Headers must not be renamed.")
    index = {h: i for i, h in enumerate(headers) if isinstance(h, str)}
    out = []
    for r in range(2, ws.max_row + 1):
        cells = [c.value for c in ws[r]]
        def at(h: str):
            i = index.get(h)
            return cells[i] if i is not None and i < len(cells) else None
        wanted = expected + [h for h in extra if h in index]
        if OTHER_COLUMN in index:
            wanted = wanted + [OTHER_COLUMN]
        record = {h: at(h) for h in wanted}
        if any(v not in (None, "") for v in record.values()):
            record["__row__"] = r
            out.append(record)
    return out


def parse_ontology_xlsx(data: bytes) -> dict:
    """An edited workbook → the ontology definition, ready for the same gate the JSON upload uses.

    Lossless against ``build_ontology_xlsx``: round-tripping a stored rulebook reproduces its
    definition, which is what ``tests/test_ontology_xlsx.py`` asserts on the shipped 185-concept
    file rather than on a fixture. Nothing is inferred — a cell that cannot be read is an error
    naming the sheet, the row and the column.
    """
    import openpyxl

    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    except Exception as exc:  # noqa: BLE001 — anything openpyxl refuses is a bad upload
        raise OntologySheetError(f"Not a readable .xlsx workbook ({exc}).") from exc

    for name in (SHEET_CONCEPTS, SHEET_CONFIG):
        if name not in wb.sheetnames:
            raise OntologySheetError(
                f"Missing the {name!r} sheet. Download the rulebook workbook and edit that file, "
                f"rather than assembling one.")

    definition: dict = {}

    # --- Config first: it carries ontology_key / target_template_key, which everything else is for.
    ws4 = wb[SHEET_CONFIG]
    for r in range(2, ws4.max_row + 1):
        key = ws4.cell(r, 1).value
        raw = ws4.cell(r, 2).value
        if key is None or str(key).strip() == "":
            continue
        key = str(key).strip()
        if raw is None or str(raw).strip() == "":
            continue
        try:
            definition[key] = json.loads(str(raw))
        except json.JSONDecodeError as exc:
            raise OntologySheetError(
                f"{SHEET_CONFIG} row {r}: {key!r} is not valid JSON ({exc.msg} at position "
                f"{exc.pos}).") from exc

    # --- Concepts
    ws = wb[SHEET_CONCEPTS]
    locales = _locale_columns(definition)
    alias_headers = [f"Aliases ({loc})" for loc in locales]
    mappings = []
    seen: set[str] = set()
    for rec in _rows(ws, CONCEPT_COLUMNS, alias_headers):
        row = rec["__row__"]
        m: dict = {}
        for k, h, t in CONCEPT_COLUMNS:
            v = _dec(rec.get(h), t, sheet=SHEET_CONCEPTS, row=row, header=h)
            if v is not None:
                m[k] = v
        key = m.get("canonical_key")
        if not key:
            raise OntologySheetError(
                f"{SHEET_CONCEPTS} row {row}: no Canonical key. Every concept needs one — it is "
                f"what ties the row to the template.")
        if key in seen:
            raise OntologySheetError(
                f"{SHEET_CONCEPTS} row {row}: {key!r} appears more than once. Two rows claiming one "
                f"concept cannot both be applied.")
        seen.add(key)
        i18n = {}
        for loc, h in zip(locales, alias_headers):
            v = _dec(rec.get(h), "l", sheet=SHEET_CONCEPTS, row=row, header=h)
            if v is not None:
                i18n[loc] = v
        if i18n:
            m["aliases_i18n"] = i18n
        _merge_other(m, rec, sheet=SHEET_CONCEPTS, row=row)
        mappings.append(m)
    if not mappings:
        raise OntologySheetError(f"{SHEET_CONCEPTS} has no concept rows.")
    definition["mappings"] = mappings

    # --- Section defaults
    if SHEET_SECTIONS in wb.sheetnames:
        sections: dict = {}
        for rec in _rows(wb[SHEET_SECTIONS], SECTION_COLUMNS, []):
            row = rec["__row__"]
            entry: dict = {}
            for k, h, t in SECTION_COLUMNS:
                v = _dec(rec.get(h), t, sheet=SHEET_SECTIONS, row=row, header=h)
                if v is not None and k != "section_key":
                    entry[k] = v
            skey = _dec(rec.get("Section key"), "s", sheet=SHEET_SECTIONS, row=row,
                        header="Section key")
            if not skey:
                raise OntologySheetError(f"{SHEET_SECTIONS} row {row}: no Section key.")
            _merge_other(entry, rec, sheet=SHEET_SECTIONS, row=row)
            sections[str(skey)] = entry
        if sections:
            definition["section_defaults"] = sections

    # --- Netting rules
    if SHEET_NETTING in wb.sheetnames:
        rules = []
        for rec in _rows(wb[SHEET_NETTING], NETTING_COLUMNS, []):
            row = rec["__row__"]
            rule: dict = {}
            for k, h, t in NETTING_COLUMNS:
                v = _dec(rec.get(h), t, sheet=SHEET_NETTING, row=row, header=h)
                if v is not None:
                    rule[k] = v
            if not rule.get("target_key"):
                raise OntologySheetError(
                    f"{SHEET_NETTING} row {row}: a netting rule needs a Target key.")
            _merge_other(rule, rec, sheet=SHEET_NETTING, row=row)
            rules.append(rule)
        if rules:
            definition["netting_rules"] = rules

    if not definition.get("ontology_key"):
        raise OntologySheetError(
            f"{SHEET_CONFIG}: no 'ontology_key'. It names the rulebook, so it cannot be blank.")
    if not definition.get("target_template_key"):
        raise OntologySheetError(
            f"{SHEET_CONFIG}: no 'target_template_key'. A rulebook maps onto ONE template's "
            f"canonical keys, and the upload is validated against it.")
    return definition
